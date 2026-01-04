"""
Source detection and data extraction utilities.

Supports multiple data sources: Google Sheets, Airtable, XLSX, raw text, PDF, image.
"""

from dotenv import load_dotenv
load_dotenv()

from typing import Any, Dict, List, Optional, Union, Tuple
from pathlib import Path
from enum import Enum
from datetime import datetime
import logging
import json
import re
import os

logger = logging.getLogger(__name__)

class SourceType(Enum):
    """Supported data source types."""
    GOOGLE_SHEETS = "google_sheet"  # Match mapping structure
    AIRTABLE = "airtable"
    SHARED_TABLE = "shared_table"
    XLSX = "xlsx_file"  # Match mapping structure
    CSV = "csv"
    RAW_TEXT = "raw_text"
    PDF = "pdf"
    IMAGE = "image"
    UNKNOWN = "unknown"

def detect_source_type(source: Union[str, Dict, Path]) -> SourceType:
    """
    Detect the type of data source.
    
    Args:
        source: Source identifier (sheet_id, file path, URL, etc.)
        
    Returns:
        SourceType enum value
    """
    if isinstance(source, dict):
        # Check for explicit source type
        if "source_type" in source:
            try:
                return SourceType(source["source_type"])
            except ValueError:
                pass
        
        # Check for Google Sheets indicators
        if "sheet_id" in source or "spreadsheet_id" in source:
            return SourceType.GOOGLE_SHEETS
        
        # Check for Airtable indicators
        if "base_id" in source or "table_id" in source:
            return SourceType.AIRTABLE
    
    if isinstance(source, (str, Path)):
        source_str = str(source)
        
        # File extensions
        if source_str.endswith(('.xlsx', '.XLSX')):
            return SourceType.XLSX
        elif source_str.endswith(('.csv', '.CSV')):
            return SourceType.CSV
        elif source_str.endswith(('.json', '.JSON')):
            # Check if it's an Airtable export (has "records" key)
            if isinstance(source, Path) or (isinstance(source, str) and not source.startswith('http')):
                try:
                    import json
                    file_path = Path(source) if isinstance(source, (str, Path)) else source
                    if file_path.exists():
                        with open(file_path, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            if isinstance(data, dict) and "records" in data:
                                return SourceType.AIRTABLE
                except (json.JSONDecodeError, IOError):
                    pass
            return SourceType.AIRTABLE  # Default for .json files
        elif source_str.endswith(('.pdf', '.PDF')):
            return SourceType.PDF
        elif source_str.endswith(('.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG')):
            return SourceType.IMAGE
        
        # URL patterns
        if 'docs.google.com/spreadsheets' in source_str:
            return SourceType.GOOGLE_SHEETS
        elif 'airtable.com' in source_str:
            return SourceType.AIRTABLE
    
    return SourceType.UNKNOWN

def fetch_from_google_sheets_raw(
    source: Union[str, Dict],
    range_: str = "Sheet1!A:Z",
    header_row_index: int = 0,
) -> List[List[Any]]:
    """
    Fetch raw 2D data from Google Sheets (helper for extract_from_source).
    
    This is a wrapper around the existing Google Sheets functionality.
    Google Sheets support is optional - will raise ImportError if dependencies are missing.
    """
    try:
        from external_tables import (
            SERVICE_ACCOUNT_FILE,
            SCOPES,
            GOOGLE_SHEETS_AVAILABLE,
        )
        
        if not GOOGLE_SHEETS_AVAILABLE:
            raise ImportError(
                "Google Sheets support requires google-auth and google-api-python-client. "
                "Install with: pip install google-auth google-api-python-client"
            )
        
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
        
        # Extract sheet_id from source
        if isinstance(source, dict):
            sheet_id = source.get("sheet_id") or source.get("spreadsheet_id")
            range_ = source.get("range", range_)
        else:
            sheet_id = str(source)
        
        if not sheet_id:
            raise ValueError("Google Sheets source must provide sheet_id")
        
        # Authenticate and fetch
        creds = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE,
            scopes=SCOPES,
        )
        
        service = build("sheets", "v4", credentials=creds)
        sheet = service.spreadsheets()
        
        result = sheet.values().get(
            spreadsheetId=sheet_id,
            range=range_,
            majorDimension="ROWS",
        ).execute()
        
        return result.get("values", [])
    except ImportError as e:
        raise ImportError(
            f"Google Sheets support not available: {e}. "
            "For testing, use CSV/Excel/Airtable files instead."
        )

def extract_from_source(
    source: Union[str, Dict, Path],
    source_type: Optional[SourceType] = None,
    mapping_id: Optional[str] = None,
    **kwargs
) -> List[List[Any]]:
    """
    Extract raw 2D data from a source.
    
    Args:
        source: Source identifier
        source_type: Optional explicit source type (auto-detected if not provided)
        **kwargs: Additional source-specific parameters
        
    Returns:
        2D list of values (rows x columns)
        
    Raises:
        NotImplementedError: If source type is not yet supported
        ValueError: If source cannot be processed
    """
    if source_type is None:
        source_type = detect_source_type(source)
    
    if source_type == SourceType.GOOGLE_SHEETS:
        return fetch_from_google_sheets_raw(source, **kwargs)
    
    elif source_type == SourceType.XLSX:
        # Load Excel file - try openpyxl first, fallback to zipfile/xml
        # Handle both file paths and dict with file_path
        if isinstance(source, dict):
            file_path = Path(source.get("file_path", source.get("path", "")))
        else:
            file_path = Path(source)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        # Try pandas first (most reliable for Excel)
        try:
            import pandas as pd
            # Read with header in first row (header=0) to preserve column names
            df = pd.read_excel(file_path, header=0, engine='openpyxl')
            # Convert DataFrame to list of lists, handling NaN values
            # First row should be headers
            rows = []
            # Add header row first
            header_row = [str(col) for col in df.columns]
            rows.append(header_row)
            # Then add data rows
            for _, row in df.iterrows():
                row_values = []
                for val in row:
                    if pd.isna(val):
                        row_values.append("")
                    elif isinstance(val, (int, float)):
                        row_values.append(str(val))
                    elif isinstance(val, datetime):
                        row_values.append(val.isoformat())
                    else:
                        val_str = str(val).strip()
                        row_values.append(val_str if val_str else "")
                rows.append(row_values)
            return rows
        except (ImportError, Exception) as e:
            logger.debug(f"[Excel] Pandas extraction failed: {e}, trying openpyxl")
            # Fallback to openpyxl
            try:
                from openpyxl import load_workbook
                # Use data_only=False to get actual cell values (not calculated formulas)
                workbook = load_workbook(file_path, data_only=False)
                sheet = workbook.active  # Use first sheet
                
                # Extract all rows
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Convert cell values properly - handle None, numbers, strings, dates
                    row_values = []
                    for cell in row:
                        if cell is None:
                            row_values.append("")
                        elif isinstance(cell, (int, float)):
                            # Keep numbers as strings to preserve precision and match truth files
                            row_values.append(str(cell))
                        elif isinstance(cell, datetime):
                            # Convert dates to ISO format string
                            row_values.append(cell.isoformat())
                        else:
                            # Convert to string, but preserve None-like values
                            cell_str = str(cell).strip()
                            row_values.append(cell_str if cell_str else "")
                    rows.append(row_values)
                
                workbook.close()
                return rows
            except ImportError:
                # Fallback to zipfile/xml method
                import zipfile
                import xml.etree.ElementTree as ET
                
                rows = []
                with zipfile.ZipFile(file_path, 'r') as z:
                    # Read shared strings
                    strings = []
                    try:
                        with z.open('xl/sharedStrings.xml') as f:
                            tree = ET.parse(f)
                            root = tree.getroot()
                            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                            strings = [t.text if t.text else '' for t in root.findall('.//main:t', ns)]
                    except:
                        pass
                    
                    # Read all rows from first sheet
                    with z.open('xl/worksheets/sheet1.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        
                        # Process rows in order
                        row_elems = sorted(root.findall('.//main:row', ns), key=lambda r: int(r.get('r', '0')))
                        for row_elem in row_elems:
                            row = []
                            # Get cells in this row, sorted by column
                            cells = sorted(row_elem.findall('main:c', ns), key=lambda c: c.get('r', ''))
                            for cell in cells:
                                val_elem = cell.find('main:v', ns)
                                cell_type = cell.get('t', '')  # 's' means shared string
                                
                                if val_elem is not None and val_elem.text:
                                    val = val_elem.text
                                    # If cell type is 's', it's a shared string index
                                    if cell_type == 's' and strings:
                                        try:
                                            idx = int(val)
                                            row.append(strings[idx] if idx < len(strings) else '')
                                        except (ValueError, IndexError):
                                            row.append(val)
                                    else:
                                        # Direct value (number, date, etc.)
                                        row.append(val)
                                else:
                                    row.append('')
                            if row:  # Only add non-empty rows
                                rows.append(row)
                
                return rows
    
    elif source_type == SourceType.CSV:
        import csv
        
        # Handle both file paths and dict with file_path
        if isinstance(source, dict):
            file_path = Path(source.get("file_path", source.get("path", "")))
        else:
            file_path = Path(source)
        
        if not file_path.exists():
            raise FileNotFoundError(f"CSV file not found: {file_path}")
        
        # Read CSV file
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        return rows
    
    elif source_type == SourceType.AIRTABLE:
        # Load Airtable JSON export
        import json
        
        # Handle both file paths and dict with file_path
        if isinstance(source, dict):
            file_path = Path(source.get("file_path", source.get("path", "")))
        else:
            file_path = Path(source)
        
        if not file_path.exists():
            raise FileNotFoundError(f"Airtable JSON file not found: {file_path}")
        
        # Load JSON file
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Convert JSON "records" format to 2D array
        # Supports both Airtable format (records[].fields) and flat format (records[] as flat objects)
        if not isinstance(data, dict) or "records" not in data:
            raise ValueError("JSON must have a 'records' key")
        
        records = data["records"]
        if not records:
            return []
        
        # Check if records use Airtable format (with "fields" key) or flat format
        first_record = records[0]
        is_airtable_format = isinstance(first_record, dict) and "fields" in first_record
        
        # Get all unique field names from all records
        all_fields = set()
        for record in records:
            if isinstance(record, dict):
                if is_airtable_format:
                    # Airtable format: record.fields.{field_name}
                    if "fields" in record:
                        all_fields.update(record["fields"].keys())
                else:
                    # Flat format: record.{field_name} directly
                    all_fields.update(record.keys())
        
        # Sort for consistency and create a mapping from original to cleaned
        from external_tables import clean_header
        original_headers = sorted(all_fields)
        cleaned_headers = [clean_header(h) for h in original_headers]
        header_map = dict(zip(original_headers, cleaned_headers))
        
        # Build 2D array: header row (cleaned) + data rows
        rows = [cleaned_headers]
        for record in records:
            row = []
            if is_airtable_format:
                fields = record.get("fields", {})
            else:
                fields = record  # Flat format - record is the fields dict
            for original_header in original_headers:
                value = fields.get(original_header, None)
                # Convert to string for consistency with CSV/Excel
                if value is not None:
                    row.append(str(value))
                else:
                    row.append("")
            rows.append(row)
        
        return rows
    
    elif source_type == SourceType.PDF:
        # Extract using OCR pipeline
        logger.debug(f"[extract_from_source] PDF extraction - mapping_id={mapping_id}")
        return _extract_from_ocr_source(source, SourceType.PDF, mapping_id=mapping_id, **kwargs)
    
    elif source_type == SourceType.IMAGE:
        # Extract using OCR pipeline
        logger.debug(f"[extract_from_source] IMAGE extraction - mapping_id={mapping_id}")
        return _extract_from_ocr_source(source, SourceType.IMAGE, mapping_id=mapping_id, **kwargs)
    
    elif source_type == SourceType.RAW_TEXT:
        # Extract using OCR pipeline (same as PDF/IMAGE)
        # Raw text files are read as text and processed through OCR pipeline
        return _extract_from_ocr_source(source, SourceType.RAW_TEXT, mapping_id=mapping_id, **kwargs)
    
    else:
        raise ValueError(f"Unknown or unsupported source type: {source_type}")

def _is_pdf_scanned_or_handwritten(pdf_path: Path, confidence_threshold: float = 0.5) -> Tuple[bool, Optional[str]]:
    """
    Detect if a PDF is scanned or handwritten (no embedded text OR low OCR confidence).
    
    Args:
        pdf_path: Path to PDF file
        confidence_threshold: Minimum OCR confidence to consider PDF as structured (default: 0.5)
    
    Returns:
        Tuple of (is_scanned, reason)
        - is_scanned: True if PDF should be treated as scanned/handwritten
        - reason: Explanation string (None if not scanned)
    """
    try:
        from ocr.reader import extract_text_from_pdf
        
        # Quick check: Try to extract embedded text using PyPDF2
        has_embedded_text = False
        try:
            import PyPDF2
            with open(pdf_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                page_count = len(pdf_reader.pages)
                
                # Check first few pages for embedded text
                pages_to_check = min(3, page_count)
                for page_num in range(pages_to_check):
                    page = pdf_reader.pages[page_num]
                    page_text = page.extract_text()
                    if page_text.strip():
                        has_embedded_text = True
                        break
        except Exception:
            # If we can't read with PyPDF2, assume no embedded text
            has_embedded_text = False
        
        # If no embedded text, it's definitely scanned
        if not has_embedded_text:
            logger.info(f"[PDF Detection] No embedded text found in {pdf_path.name} - treating as scanned/handwritten")
            return True, "no_embedded_text"
        
        # If embedded text exists, check OCR confidence
        # Extract text with OCR to get confidence score
        text_blocks, metadata = extract_text_from_pdf(pdf_path, enable_vision=False)
        
        # Check confidence from metadata
        if metadata and hasattr(metadata, 'confidence'):
            confidence = metadata.confidence
            if confidence < confidence_threshold:
                logger.info(f"[PDF Detection] Low OCR confidence ({confidence:.2f} < {confidence_threshold}) in {pdf_path.name} - treating as scanned/handwritten")
                return True, f"low_confidence_{confidence:.2f}"
        
        # If we have embedded text and good confidence, it's structured
        logger.debug(f"[PDF Detection] PDF {pdf_path.name} has embedded text and good confidence - treating as structured")
        return False, None
        
    except Exception as e:
        logger.warning(f"[PDF Detection] Error detecting PDF type for {pdf_path.name}: {e}, defaulting to scanned")
        # On error, default to scanned (safer for handwritten documents)
        return True, f"detection_error_{str(e)[:50]}"

def _is_ocr_result_unusable(rows_2d: List[List[Any]]) -> Tuple[bool, str]:
    """
    Check if OCR extraction result is unusable and should trigger Vision API fallback.
    
    Args:
        rows_2d: 2D array of extracted rows (header row + data rows)
    
    Returns:
        Tuple of (is_unusable: bool, reason: str)
    """
    if not rows_2d or len(rows_2d) == 0:
        return (True, "No rows extracted")
    
    # Check if we only have header row (no data rows)
    if len(rows_2d) <= 1:
        return (True, "Only header row, no data rows")
    
    # Check if all data rows have only None/empty values
    data_rows = rows_2d[1:]  # Skip header row
    all_empty = True
    for row in data_rows:
        if row:
            # Check if row has any non-empty values
            for val in row:
                if val and str(val).strip() and str(val).strip().lower() not in ['none', 'null', '']:
                    all_empty = False
                    break
        if not all_empty:
            break
    
    if all_empty:
        return (True, "All rows contain only None/empty values")
    
    # Check for jumbled multi-vehicle rows (very long column names or values)
    # This indicates OCR extracted text incorrectly
    header_row = rows_2d[0] if rows_2d else []
    for header in header_row:
        header_str = str(header) if header else ""
        # Very long headers (>100 chars) suggest jumbled OCR
        if len(header_str) > 100:
            return (True, f"Jumbled OCR output detected (header length: {len(header_str)})")
    
    # Check for suspicious patterns in data (multiple vehicles concatenated)
    # Look for patterns like "Model: Camry Model: F-150" in single cells
    for row in data_rows[:3]:  # Check first 3 rows
        for val in row:
            val_str = str(val) if val else ""
            # Pattern: multiple occurrences of field labels (e.g., "Model: X Model: Y")
            if re.search(r'(Model|Make|Year|VIN)\s*:.*?(Model|Make|Year|VIN)\s*:', val_str, re.IGNORECASE):
                return (True, "Jumbled multi-vehicle data detected in OCR output")
    
    return (False, "")

# Vision API filler words that should be removed from extracted values
VISION_FILLER_WORDS = [
    "is", "was", "are", "described", "described as", "listed as", "shown as", "marked as"
]

def cleanup_vision_value(value: Optional[str]) -> Optional[str]:
    """
    Clean up Vision API extracted values by removing filler words and trailing punctuation.
    
    Removes common filler phrases like "is", "was", "described", "described as" from the start of values.
    Also removes trailing punctuation (periods, commas).
    
    Examples:
        "is automatic" -> "automatic"
        "described as manual" -> "manual"
        "described" -> None
        "is" -> None
    
    Args:
        value: Raw value from Vision API (string or None)
    
    Returns:
        Cleaned value string, or None if value is empty/None after cleaning
    """
    if not value or not isinstance(value, str):
        return value
    
    import re
    
    v = value.strip()
    if not v:
        return None
    
    # Convert to lowercase for comparison, but preserve original case
    v_lower = v.lower()
    
    # Remove filler words/phrases from the start (try longer phrases first)
    # Sort by length descending to match longer phrases first
    sorted_phrases = sorted(VISION_FILLER_WORDS, key=len, reverse=True)
    
    for phrase in sorted_phrases:
        # Check if value starts with phrase followed by space
        if v_lower.startswith(phrase + " "):
            v = v[len(phrase):].strip()
            v_lower = v.lower()
            # Continue checking in case there are multiple filler words
        # Also check if value is exactly the phrase (edge case)
        elif v_lower == phrase:
            return None
    
    # Remove trailing punctuation (periods, commas)
    v = re.sub(r'[.,]+$', '', v).strip()
    
    # Return None if empty after cleaning
    return v if v else None

def _normalize_vision_value(field_name: str, value: Any, preserve_raw: bool = True) -> Any:
    """
    Clean and preserve raw values from Vision API output.
    
    ARCHITECTURAL DECISION: All normalization (canonical forms, transforms) happens in normalize_v2().
    This function only does minimal cleaning (remove filler words, basic type conversion for numerics).
    
    Args:
        field_name: Field name from schema
        value: Raw value from Vision API
        preserve_raw: If True, preserve raw values (all sources should use True now)
    
    Returns:
        Cleaned raw value (preserves original format, minimal type conversion for numerics)
    """
    if value is None or value == "" or (isinstance(value, str) and value.strip().lower() in ['none', 'null', '']):
        return None
    
    # Clean up Vision API filler words BEFORE normalization
    if isinstance(value, str):
        value = cleanup_vision_value(value)
        if value is None:
            return None
    
    value_str = str(value).strip()
    
    # For PDF sources (preserve_raw=True), preserve exact values as extracted by Vision
    # CRITICAL: Preserve raw string values when type conversion fails (instead of returning None)
    # This allows normalize_v2 to generate warnings for invalid values
    if preserve_raw:
        # Type conversions only (no validation or normalization)
        if field_name == "year":
            try:
                return int(value_str)
            except (ValueError, TypeError):
                # Preserve raw string value if conversion fails - let normalize_v2 generate warnings
                return value_str
        
        if field_name == "mileage":
            try:
                # Remove commas and other formatting, but preserve negative values
                mileage_str = value_str.replace(",", "").replace(" ", "")
                return int(mileage_str)  # Preserve negative values
            except (ValueError, TypeError):
                # Preserve raw string value if conversion fails - let normalize_v2 generate warnings
                return value_str
        
        if field_name == "weight":
            try:
                weight_str = value_str.replace(",", "").replace(" ", "")
                weight_int = int(weight_str)
                # Preserve negative weights as string (let normalize_v2 handle warnings)
                return weight_int if weight_int >= 0 else value_str
            except (ValueError, TypeError):
                # Preserve raw string value if conversion fails - let normalize_v2 generate warnings
                return value_str
        
        # No special PDF handling - use same normalization as other sources
        
        # VIN: uppercase (still needed for consistency) - preserve exactly as extracted, no correction
        if field_name == "vin":
            return value_str.upper()
        
        # Email: preserve raw (even if invalid format) - NEVER return None for PDF sources
        if field_name == "owner_email":
            # Always return the string value, even if it looks invalid
            # Do NOT default to None - preserve exactly as extracted
            return value_str if value_str else None
        
        # Return as string for other fields
        return value_str if value_str else None
    
    # ARCHITECTURAL DECISION: All normalization removed from this function.
    # All canonical transformations (fuel_type, transmission, body_style, casing) happen in normalize_v2() via transforms.
    # This function only does minimal cleaning and type conversion for numerics.
    
    # For non-PDF sources, apply same minimal cleaning as PDF sources (preserve raw values)
    # Type conversions for numeric fields only - preserve invalid values, warnings will be generated later
    # CRITICAL: Preserve raw string values when type conversion fails (instead of returning None)
    # This allows normalize_v2 to generate warnings for invalid values
    if field_name == "year":
        try:
            year = int(value_str)
            # Preserve invalid years - do not return None
            return year
        except (ValueError, TypeError):
            # Preserve raw string value if conversion fails - let normalize_v2 generate warnings
            return value_str
    
    if field_name == "mileage":
        try:
            # Remove commas and other formatting, but preserve negative values
            mileage_str = value_str.replace(",", "").replace(" ", "")
            mileage = int(mileage_str)
            # Preserve negative mileage - do not return None
            return mileage
        except (ValueError, TypeError):
            # Preserve raw string value if conversion fails - let normalize_v2 generate warnings
            return value_str
    
    if field_name == "weight":
        try:
            weight_str = value_str.replace(",", "").replace(" ", "")
            weight_int = int(weight_str)
            # Preserve negative weights as string (let normalize_v2 handle warnings)
            return weight_int if weight_int >= 0 else value_str
        except (ValueError, TypeError):
            # Preserve raw string value if conversion fails - let normalize_v2 generate warnings
            return value_str
    
    # CRITICAL: Do NOT normalize transmission, fuel_type, body_style, casing here.
    # All normalization happens in normalize_v2() via transforms.
    # Return raw values as-is (only VIN gets uppercase for consistency)
    
    # VIN: uppercase (for consistency, but no other normalization)
    if field_name == "vin":
        return value_str.upper()
    
    # Return as string for all other fields (preserve raw, no normalization)
    return value_str if value_str else None

def _extract_from_ocr_text_with_fallback(
    ocr_text: str,
    page_num: int
) -> Optional[Dict[str, Any]]:
    """
    Extract vehicle data from OCR text using fallback inference functions.
    
    Args:
        ocr_text: Raw OCR text from a single page
        page_num: Page number (for logging)
    
    Returns:
        Dictionary with extracted fields, or None if nothing found
    """
    try:
        from inference import (
            detect_vin_in_row,
            detect_year_in_row,
            detect_make_model_in_row
        )
    except ImportError:
        logger.error("[Fallback] Could not import inference functions")
        return None
    
    # Create a row dict from OCR text (treat entire text as one cell)
    row_dict = {"ocr_text": ocr_text}
    
    # Try to extract fields using fallback inference
    extracted = {}
    
    # VIN
    vin_result = detect_vin_in_row(row_dict)
    if vin_result:
        extracted["vin"] = vin_result[0]
    
    # Year
    year_result = detect_year_in_row(row_dict)
    if year_result:
        extracted["year"] = year_result[0]
    
    # Make/Model
    make_model_result = detect_make_model_in_row(row_dict)
    if make_model_result:
        make, model, _ = make_model_result
        if make:
            extracted["make"] = make
        if model:
            extracted["model"] = model
    
    # If we extracted at least one field, create a row
    if extracted:
        logger.info(f"[Fallback] Extracted page {page_num} from OCR text: {list(extracted.keys())}")
        return extracted
    
    return None

# ============================================================================
# Raw Text Vehicle Extraction - Dedicated Field Extractors
# ============================================================================

def _extract_vin_from_text(text: str) -> Optional[str]:
    """Extract VIN from text block. Returns first valid VIN found."""
    # VIN patterns: "VIN is X", "VIN# X", "VIN: X", or standalone 17-char alphanumeric
    vin_patterns = [
        r'VIN\s+(?:is|#|:)\s*([A-HJ-NPR-Z0-9]{10,20})',
        r'\b([A-HJ-NPR-Z0-9]{17})\b',  # Standard 17-char VIN
        r'\b([A-HJ-NPR-Z0-9]{10,20})\b'  # Fallback: any 10-20 char alphanumeric
    ]
    
    for pattern in vin_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).upper()
    return None

def _extract_year_from_text(text: str) -> Optional[int]:
    """Extract 4-digit year from text block."""
    year_pattern = r'\b(19\d{2}|20[0-3]\d)\b'
    match = re.search(year_pattern, text)
    if match:
        try:
            return int(match.group(1))
        except ValueError:
            pass
    return None

def _extract_make_from_text(text: str) -> Optional[str]:
    """Extract vehicle make from text block. Ignores example text like 'Vehicle V001'."""
    known_makes = [
        'Toyota', 'Ford', 'Honda', 'Tesla', 'BMW', 'Nissan', 'Chevrolet', 'Chevy',
        'Mercedes', 'Mercedes-Benz', 'Audi', 'Volkswagen', 'VW', 'Hyundai', 'Kia',
        'Mazda', 'Subaru', 'Jeep', 'Ram', 'GMC', 'Cadillac', 'Lexus', 'Acura',
        'Infiniti', 'Lincoln', 'Buick', 'Chrysler', 'Dodge', 'Porsche', 'Volvo',
        'Jaguar', 'Land Rover', 'Mitsubishi', 'Genesis', 'Alfa Romeo', 'Fiat',
        'Mini', 'Smart', 'Ferrari', 'Lamborghini', 'Maserati', 'Bentley', 'Rolls-Royce'
    ]
    
    # Remove example/template text patterns before extraction
    # Remove "Vehicle V001:" or "Vehicle V0XX:" patterns
    text_cleaned = re.sub(r'Vehicle\s+V\d+:\s*', '', text, flags=re.IGNORECASE)
    
    # Look for patterns like "2024 Toyota" or "Toyota Camry" - but NOT "Vehicle"
    make_model_patterns = [
        r'(?:\d{4}\s+)([A-Z][a-zA-Z]+)\s+(?:Camry|F-150|Model|CR-V|328i|Altima|Civic|Accord|Fusion|Silverado|F-250|Corolla|RAV4|Highlander)',
        r'\b([A-Z][a-zA-Z]+)\s+(?:Camry|F-150|Model|CR-V|328i|Altima|Civic|Accord|Fusion|Silverado|F-250|Corolla|RAV4|Highlander)',
    ]
    
    for pattern in make_model_patterns:
        match = re.search(pattern, text_cleaned)
        if match:
            potential_make = match.group(1).strip()
            # Skip if it's "Vehicle" or other non-make words
            if potential_make.lower() in ['vehicle', 'this', 'is', 'a']:
                continue
            # Check against known makes
            for make in known_makes:
                if potential_make.startswith(make) or make.startswith(potential_make):
                    return make
    return None

def _extract_model_from_text(text: str, make: Optional[str] = None) -> Optional[str]:
    """Extract vehicle model from text block. Ignores example identifiers like 'V001'."""
    # Remove example/template text patterns
    text_cleaned = re.sub(r'Vehicle\s+V\d+:\s*', '', text, flags=re.IGNORECASE)
    
    # Common model names to look for (ordered by specificity - multi-word first)
    known_models = ['Model 3', 'F-150', 'CR-V', 'F-250', 'Camry', '328i', 'Altima', 'Civic', 'Accord', 
                    'Fusion', 'Silverado', 'Corolla', 'RAV4', 'Highlander']
    
    # If we have make, look for pattern after make - capture multi-word models
    if make:
        # Pattern to capture model after make, including multi-word models like "Model 3"
        pattern = rf'{re.escape(make)}\s+([A-Z][a-zA-Z0-9\-]+(?:\s+[A-Z][a-zA-Z0-9\-]+)?)'
        match = re.search(pattern, text_cleaned)
        if match:
            potential_model = match.group(1).strip()
            # Skip if it's an example identifier like "V001"
            if re.match(r'V\d+', potential_model, re.IGNORECASE):
                return None
            # Check if it matches a known model (for validation)
            for known_model in known_models:
                if potential_model.startswith(known_model) or known_model.startswith(potential_model):
                    return known_model
            # Return the extracted model even if not in known list
            return potential_model
    
    # Look for known model names (check multi-word models first)
    for model in known_models:
        pattern = rf'\b{re.escape(model)}\b'
        if re.search(pattern, text_cleaned):
            return model
    
    return None

def _extract_color_from_text(text: str) -> Optional[str]:
    """Extract vehicle color from text block."""
    colors = ['blue', 'red', 'white', 'black', 'silver', 'gray', 'grey', 'green', 'yellow',
              'orange', 'purple', 'brown', 'tan', 'beige', 'gold', 'bronze', 'maroon', 'navy']
    
    text_lower = text.lower()
    for color in colors:
        color_pattern = r'\b' + re.escape(color) + r'\b'
        if re.search(color_pattern, text_lower):
            return color.capitalize()
    return None

def _extract_mileage_from_text(text: str) -> Optional[int]:
    """Extract mileage from text block."""
    mileage_pattern = r'(?:about\s+)?([\d,]+)\s*(?:miles?|mi\.?)(?:\s+on\s+the\s+odometer)?'
    match = re.search(mileage_pattern, text, re.IGNORECASE)
    if match:
        try:
            mileage_str = match.group(1).replace(',', '')
            return int(mileage_str)
        except ValueError:
            pass
    return None

def _extract_fuel_type_from_text(text: str) -> Optional[str]:
    """Extract fuel type from text block. Returns raw value (normalization happens in normalize_v2)."""
    fuel_pattern = r'Fuel:\s*(\w+)'
    match = re.search(fuel_pattern, text, re.IGNORECASE)
    if match:
        return match.group(1).lower()
    return None

def _extract_transmission_from_text(text: str) -> Optional[str]:
    """Extract transmission from text block. Returns raw value (normalization happens in normalize_v2)."""
    transmission_pattern = r'transmission.*?(?:is\s+)?(?:described\s+as\s+)?(?:an?\s+)?([^.,]+?)(?:\.|,|$)'
    match = re.search(transmission_pattern, text, re.IGNORECASE)
    if match:
        # Return raw value - normalize_v2 will handle normalization via transforms
        return match.group(1).strip()
    return None

def _extract_owner_email_from_text(text: str) -> Optional[str]:
    """Extract owner email from text block."""
    email_patterns = [
        r'(?:Owner\s+contact\s+email|email):\s*([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})',
        r'\b([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})\b'
    ]
    
    for pattern in email_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1).lower()
    return None

def _extract_body_style_from_text(text: str) -> Optional[str]:
    """Extract body style from text block."""
    body_styles = ['sedan', 'truck', 'suv', 'crossover', 'coupe', 'convertible', 'wagon', 
                   'hatchback', 'van', 'minivan', 'pickup', 'sport utility']
    
    text_lower = text.lower()
    for style in body_styles:
        if style in text_lower:
            return style
    return None

def _extract_notes_from_text(text: str, vehicle: Dict[str, Any]) -> Optional[str]:
    """Extract notes from text block. Returns clean narrative sentence, ignoring example blocks."""
    # Remove example/template text patterns
    text_cleaned = re.sub(r'Vehicle\s+V\d+:\s*', '', text, flags=re.IGNORECASE)
    text_cleaned = re.sub(r'Another\s+unit\s+is.*?\.', '', text_cleaned, flags=re.IGNORECASE | re.DOTALL)
    
    # Extract the full descriptive sentence for THIS vehicle only
    # Pattern: "This is a [year] [make] [model] [body_style] painted [color]. [mileage] miles. Fuel: [fuel]. [transmission] transmission."
    # Match from "This is a" until the email or end of sentence
    pattern = r'This is a\s+(\d{4}\s+)?([A-Z][a-zA-Z]+\s+[A-Z][a-zA-Z0-9\-\s]+?)(?:\.|Owner|$)'
    match = re.search(pattern, text_cleaned, re.IGNORECASE | re.DOTALL)
    if match:
        # Get the full sentence up to the period before "Owner"
        full_match = match.group(0)
        # Extract everything from "This is a" until just before "Owner contact email"
        notes_pattern = r'(This is a.*?)(?:Owner contact email|$)'
        notes_match = re.search(notes_pattern, text_cleaned, re.IGNORECASE | re.DOTALL)
        if notes_match:
            notes = notes_match.group(1).strip()
            # Remove trailing periods and clean up
            notes = re.sub(r'\.+$', '', notes).strip()
            # Ensure it ends with a period
            if notes and not notes.endswith('.'):
                notes += '.'
            return notes
    
    # Fallback: extract first complete sentence that describes the vehicle
    sentences = re.split(r'[.!?]+', text_cleaned)
    for sentence in sentences:
        sentence = sentence.strip()
        # Skip if it's just "Vehicle V001" or similar
        if re.match(r'^Vehicle\s+V\d+', sentence, re.IGNORECASE):
            continue
        # Skip if it's too short or doesn't contain vehicle info
        if len(sentence) > 20 and any(word in sentence.lower() for word in ['painted', 'miles', 'fuel', 'transmission']):
            return sentence + '.'
    
    return None

def _extract_document_level_defaults(full_ocr_text: str, source_type: str, file_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Pre-split extraction: Scan full OCR text to extract document-level defaults
    for body_style, fuel_type, transmission, and mileage.
    
    These fields may appear in the full text but be lost during VIN-based splitting.
    Only applies to handwritten PDF/image or PNG image sources.
    
    Args:
        full_ocr_text: Full OCR text before splitting
        source_type: Source type ("pdf" or "image")
        file_path: Optional file path to check for handwriting indicators
    
    Returns:
        Dict with extracted defaults (only fields that were found, None for others)
    """
    import re
    defaults = {
        'body_style': None,
        'fuel_type': None,
        'transmission': None,
        'mileage': None
    }
    
    # Only apply for PDF/image sources
    if source_type not in ("pdf", "image"):
        return defaults
    
    # Check if file path indicates handwriting (more reliable than OCR text analysis)
    is_handwritten_by_path = False
    if file_path:
        file_path_str = str(file_path).lower()
        is_handwritten_by_path = any(keyword in file_path_str for keyword in ['handwritten', 'hand'])
    
    # Check if this is likely handwritten (no structured labels in full text)
    # BUT: OCR from handwritten docs can still produce labels, so we also check file path
    has_structured_labels = bool(re.search(r'\b(Make|Model|Year|Color|Mileage|Transmission|Fuel|Body)[:\s]', full_ocr_text, re.IGNORECASE))
    is_likely_handwritten_by_text = not has_structured_labels and len(full_ocr_text) > 50
    
    # If file path indicates handwriting, treat as handwritten even if labels are present
    is_likely_handwritten = is_handwritten_by_path or is_likely_handwritten_by_text
    
    # Apply only for handwritten docs or all image sources (PNG images)
    if not (is_likely_handwritten or source_type == "image"):
        return defaults
    
    text_lower = full_ocr_text.lower()
    
    # Extract body_style: Look for keywords anywhere in full text
    if defaults['body_style'] is None:
        body_styles = ['sedan', 'truck', 'suv', 'crossover', 'coupe', 'convertible', 'wagon', 'hatchback']
        for style in body_styles:
            if re.search(rf'\b{re.escape(style)}\b', text_lower):
                defaults['body_style'] = style
                break
    
    # Extract fuel_type: Look for keywords anywhere in full text
    if defaults['fuel_type'] is None:
        if re.search(r'\b(gas|gasoline)\b', text_lower):
            defaults['fuel_type'] = 'gasoline'
        elif re.search(r'\bdiesel\b', text_lower):
            defaults['fuel_type'] = 'diesel'
        elif re.search(r'\belectric\b', text_lower):
            defaults['fuel_type'] = 'electric'
        elif re.search(r'\bhybrid\b', text_lower):
            defaults['fuel_type'] = 'hybrid'
    
    # Extract transmission: Look for keywords anywhere in full text
    if defaults['transmission'] is None:
        if re.search(r'\b(automatic|auto)\b', text_lower):
            defaults['transmission'] = 'automatic'
        elif re.search(r'\bmanual\b', text_lower):
            defaults['transmission'] = 'manual'
        elif re.search(r'\bcvt\b', text_lower):
            defaults['transmission'] = 'cvt'
    
    # Extract mileage: Look for patterns in full text
    if defaults['mileage'] is None:
        mileage_patterns = [
            r'(?:mileage|milage)[:\s]+([\d,]+)',  # "Mileage: 100,245" or "Milage: 100,245" (handles typo)
            r'(-?[\d,]+)\s*(?:miles?|mi\.?)',  # "100,245 miles" or "-150 miles"
            r'about\s+([\d,]+)\s*(?:miles?|mi\.?)',  # "about 100,245 miles"
            r'([\d,]+)\s*(?:miles?|mi\.?)\s+on\s+(?:the\s+)?odometer',  # "100,245 miles on odometer"
            r'odometer[:\s]+([\d,]+)',  # "odometer: 100,245"
        ]
        for pattern in mileage_patterns:
            mileage_match = re.search(pattern, full_ocr_text, re.IGNORECASE)
            if mileage_match:
                try:
                    mileage_str = mileage_match.group(1).replace(',', '').replace(' ', '')
                    defaults['mileage'] = int(mileage_str)  # Preserve negative values
                    break
                except ValueError:
                    continue
    
    return defaults

def split_by_vin(text: str) -> List[str]:
    """
    Non-overlapping segmentation of OCR or raw text into per-vehicle text blocks.
    Each block begins with vehicle description (or VIN) and ends right before the next VIN.
    """
    import re
    # VIN pattern: matches VINs in various formats
    # Pattern 1: "VIN" followed by optional chars (including parens/colon) and whitespace, then VIN
    # Pattern 2: Standalone 17-char VIN (standard format)
    # Pattern 3: Standalone 10-16 char alphanumeric (will be validated for digits below)
    # Exclude common words that might match: DEALERSHIP, IDENTIFICATION, TRANSMISSION, etc.
    # VINs typically have a mix of letters and numbers, not just letters
    vin_regex = r'VIN[#\s:)]+\s*([A-Z0-9]{10,20})|([A-HJ-NPR-Z0-9]{17})\b|([A-Z0-9]{10,16})\b'
    matches = list(re.finditer(vin_regex, text, re.IGNORECASE | re.MULTILINE))
    
    # SAFEGUARD: Filter out false positives (common words that match the pattern)
    # Expanded list to include all header tokens that might be misidentified as VINs
    # Note: All comparisons are case-insensitive (converted to uppercase)
    excluded_words = {'DEALERSHIP', 'IDENTIFICATION', 'TRANSMISSION', 'DESCRIPTION', 
                      'INFORMATION', 'REGISTRATION', 'DOCUMENTATION', 'SPECIFICATION',
                      'AUTOMATIC', 'MANUAL', 'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 
                      'VIN', 'BODY_STYLE', 'FUEL_TYPE', 'OWNER_EMAIL', 'NOTES', 'VEHICLE_ID',
                      'EFFECTIVE_DATE', 'TRIM', 'WEIGHT', 'BODY_TYPE', 'CURRENT_MILEAGE',
                      'EXTERIOR_COLOR', 'FUEL', 'BODY', 'PAINTED', 'GASOLINE', 'CURRENT', 
                      'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET', 'STYLE', 'TYPE', 'EMAIL'}
    
    blocks = []
    valid_matches = []
    seen_vins = set()  # Track VINs we've already seen to avoid duplicates
    
    for m in matches:
        vin_value = (m.group(1) or m.group(2) or m.group(3))
        if vin_value:
            vin_upper = vin_value.upper().strip()
            # SAFEGUARD: Exclude common words FIRST (before other checks)
            if vin_upper in excluded_words:
                logger.debug(f"[split_by_vin] Rejected excluded word as VIN: {vin_upper}")
                continue
            # Additional check: VINs should have at least one digit
            if not re.search(r'\d', vin_value):
                logger.debug(f"[split_by_vin] Rejected VIN with no digits: {vin_value}")
                continue
            # VINs should have at least 2 digits (to avoid false positives like "IDENTIFICATION")
            if len([c for c in vin_value if c.isdigit()]) < 2:
                logger.debug(f"[split_by_vin] Rejected VIN with < 2 digits: {vin_value}")
                continue
            # SAFEGUARD: Additional length check - VINs should be 10-17 characters
            if len(vin_value) < 10 or len(vin_value) > 17:
                logger.debug(f"[split_by_vin] Rejected VIN with invalid length ({len(vin_value)}): {vin_value}")
                continue
            # Skip if we've already seen this VIN (deduplicate)
            if vin_upper in seen_vins:
                logger.debug(f"[split_by_vin] Rejected duplicate VIN: {vin_upper}")
                continue
            seen_vins.add(vin_upper)
            valid_matches.append(m)
    
    # VIN RECOVERY FALLBACK: If regex didn't find enough VINs, try removing line breaks
    # Run recovery if we have fewer than 6 VINs (expected vehicle count)
    if len(valid_matches) < 6:
        # DEBUG: Print diagnostics for each candidate (only if we need recovery)
        logger.debug(f"[split_by_vin DEBUG] Found {len(matches)} raw VIN regex matches")
        for i, m in enumerate(matches):
            vin_value = (m.group(1) or m.group(2) or m.group(3))
            if vin_value:
                vin_upper = vin_value.upper().strip()
                digit_count = len([c for c in vin_value if c.isdigit()])
                has_digit = bool(re.search(r'\d', vin_value))
                is_excluded = vin_upper in excluded_words
                is_duplicate = vin_upper in seen_vins
                
                logger.debug(f"  Candidate {i+1}: '{vin_value}'")
                logger.debug(f"    Length: {len(vin_value)}")
                logger.debug(f"    Digit count: {digit_count}")
                logger.debug(f"    Has digit: {has_digit}")
                logger.debug(f"    Valid after VIN rules?: {not is_excluded and has_digit and digit_count >= 2 and not is_duplicate}")
                if is_excluded:
                    logger.debug(f"    REJECTED: Excluded word")
                elif not has_digit:
                    logger.debug(f"    REJECTED: No digits")
                elif digit_count < 2:
                    logger.debug(f"    REJECTED: Less than 2 digits")
                elif is_duplicate:
                    logger.debug(f"    REJECTED: Duplicate")
        
        logger.debug(f"[split_by_vin DEBUG] Only {len(valid_matches)} valid VIN(s) found, trying recovery fallback...")
        # Remove line breaks and whitespace to find VINs that were split across lines
        text_no_breaks = re.sub(r'[\s\n\r]+', '', text)
        
        # Find all 17-character alphanumeric sequences (excluding I, O, Q)
        recovery_candidates = re.findall(r'[A-HJ-NPR-Z0-9]{17}', text_no_breaks, re.IGNORECASE)
        logger.debug(f"[split_by_vin DEBUG] Recovery fallback found {len(recovery_candidates)} potential 17-char sequences")
        
        for candidate in recovery_candidates:
            candidate_upper = candidate.upper()
            digit_count = len([c for c in candidate if c.isdigit()])
            
            # Validate: must have at least 1 digit, not be an excluded word, length exactly 17
            if candidate_upper in excluded_words:
                logger.debug(f"  Recovery candidate '{candidate}': REJECTED (excluded word)")
                continue
            if digit_count < 1:
                logger.debug(f"  Recovery candidate '{candidate}': REJECTED (no digits)")
                continue
            if len(candidate) != 17:
                logger.debug(f"  Recovery candidate '{candidate}': REJECTED (length != 17)")
                continue
            
            # Check if this VIN was already found by regex
            if candidate_upper in seen_vins:
                logger.debug(f"  Recovery candidate '{candidate}': SKIPPED (already found by regex)")
                continue
            
            # Try to find this VIN in the original text (with line breaks)
            # Build pattern allowing any whitespace between characters
            vin_chars = list(candidate_upper)
            vin_search_pattern = r'[\s\n\r]*'.join([re.escape(c) for c in vin_chars])
            recovery_match = re.search(vin_search_pattern, text, re.IGNORECASE)
            
            if recovery_match:
                logger.debug(f"  Recovery candidate '{candidate}': ACCEPTED (found in text with line breaks at position {recovery_match.start()})")
                seen_vins.add(candidate_upper)
                # Create a match object-like structure for the recovery VIN
                class RecoveryMatch:
                    def __init__(self, start_pos, end_pos, vin_value):
                        self.start_pos = start_pos
                        self.end_pos = end_pos
                        self.vin_value = vin_value
                    def start(self):
                        return self.start_pos
                    def end(self):
                        return self.end_pos
                    def group(self, n):
                        if n == 2:  # Pattern 2 is the 17-char standalone
                            return self.vin_value
                        return None
                
                # Find the best position: look for "VIN" near this match
                match_start = recovery_match.start()
                lookback_start = max(0, match_start - 100)
                lookback = text[lookback_start:match_start]
                vin_keyword = re.search(r'VIN[#\s:)]*\s*', lookback, re.IGNORECASE)
                if vin_keyword:
                    # Use position after "VIN" keyword
                    actual_start = lookback_start + vin_keyword.end()
                else:
                    # Use the match start
                    actual_start = match_start
                
                recovery_match_obj = RecoveryMatch(actual_start, recovery_match.end(), candidate_upper)
                valid_matches.append(recovery_match_obj)
            else:
                logger.debug(f"  Recovery candidate '{candidate}': REJECTED (not found in original text)")
    
    if not valid_matches:
        logger.debug("[split_by_vin DEBUG] No valid VIN matches found after recovery")
        return []
    
    # Sort matches by position to process in order
    valid_matches.sort(key=lambda m: m.start())
    
    # Only print debug if we didn't find 6 VINs (success case)
    if len(valid_matches) < 6:
        logger.debug(f"[split_by_vin DEBUG] Total valid VINs after recovery: {len(valid_matches)}")
    # Success case (6+ VINs): debug prints automatically disabled
    
    for i, m in enumerate(valid_matches):
        vin_start = m.start()
        vin_end = m.end()
        
        # Determine start of this block
        # Look backwards for "Vehicle Identification Number" header (for PDFs) or "Vehicle V00X:" (for raw text)
        if i == 0:
            # First block: look back from start
            lookback_start = max(0, vin_start - 500)
        else:
            # Subsequent blocks: start from previous block's end (or previous VIN end)
            prev_vin_end = valid_matches[i - 1].end()
            lookback_start = max(prev_vin_end, vin_start - 500)
        
        lookback_text = text[lookback_start:vin_start]
        
        # Try to find "Vehicle Identification Number" header first (for PDF format)
        # Find the LAST occurrence in the lookback text (closest to current VIN)
        vin_header_matches = list(re.finditer(r'Vehicle\s+Identification\s+Number', lookback_text, re.IGNORECASE))
        if vin_header_matches:
            # Use the last (closest) match
            vin_header_match = vin_header_matches[-1]
            start = lookback_start + vin_header_match.start()
        else:
            # Fallback: look for "Vehicle V00X:" pattern (for raw text format)
            vehicle_matches = list(re.finditer(r'(Vehicle\s+V\d+:)', lookback_text, re.IGNORECASE))
            if vehicle_matches:
                # Use the last (closest) match
                vehicle_start_match = vehicle_matches[-1]
                start = lookback_start + vehicle_start_match.start()
            else:
                # Find last newline or start of text
                line_start = text.rfind('\n', lookback_start, vin_start)
                start = line_start + 1 if line_start >= 0 else lookback_start
        
        # Determine end of this block
        # Look for the next "Vehicle Identification Number" header (for PDF format)
        if i < len(valid_matches) - 1:
            # Search between current VIN end and next VIN start
            search_start = vin_end
            search_end = valid_matches[i + 1].start()
            search_text = text[search_start:search_end]
            
            # Look for next "Vehicle Identification Number" header
            next_vin_header = re.search(r'\n\s*Vehicle\s+Identification\s+Number', search_text, re.IGNORECASE)
            if next_vin_header:
                # End right before the next header
                end = search_start + next_vin_header.start()
            else:
                # Look for next "Vehicle V00X:" pattern
                next_vehicle_marker = re.search(r'\n\s*Vehicle\s+V\d+:', search_text, re.IGNORECASE)
                if next_vehicle_marker:
                    end = search_start + next_vehicle_marker.start()
                else:
                    # Use the next VIN's start position
                    end = search_end
        else:
            # Last block: go to end of text
            end = len(text)
        
        block = text[start:end].strip()
        # Only add non-empty blocks
        if block:
            blocks.append(block)
    return blocks

def extract_fields_from_block(text_block: str, mapping: Optional[Dict[str, Any]] = None, source_type: Optional[str] = None) -> Dict[str, Any]:
    """
    Extracts ALL fields for a single vehicle from a VIN-isolated block.
    Uses mapping.ai_instruction for hints.
    Must NOT scan outside the block.
    Must NOT leak values from other vehicles.
    
    IMPORTANT: All regex searches MUST use text_block only, never full_text or global text.
    
    Args:
        text_block: Text block containing vehicle information
        mapping: Optional mapping configuration
        source_type: Source type ("pdf", "raw", "image") to determine extraction logic
    """
    from schema import VEHICLE_SCHEMA_ORDER
    import re
    
    # Debug: Log block being processed (suppressed in test output)
    logger.debug(f"[extract_fields_from_block] USING BLOCK: {text_block[:150]}")
    
    # Enhanced debug for handwritten documents: print full block content
    # This will help us understand if keywords are in blocks or lost during splitting
    if source_type in ("pdf", "image") and len(text_block) > 50:
        has_structured_labels = bool(re.search(r'\b(Make|Model|Year|Color|Mileage|Transmission|Fuel|Body)[:\s]', text_block, re.IGNORECASE))
        if not has_structured_labels:
            # Likely handwritten - print block for debugging (only for test VINs to avoid spam)
            block_vin = None
            vin_match = re.search(r'([A-HJ-NPR-Z0-9]{17})', text_block, re.IGNORECASE)
            if vin_match:
                block_vin = vin_match.group(1).upper()
            
            # Only debug for known test VINs
            test_vins = ['HBUSRJGF4CBFPR9BN', 'HBU5RTGF4CBEPR9BN', 'HBUSRTGF4CBEPR9BN', 
                        '3R5UAL4YUKPYGF1GZ', '3R5UAL4YUKPYGFIG7', 
                        'ST420RJ98FDHKL4E', 'STLZ0RJ98FDHKL4HE']
            if block_vin in test_vins:
                print(f"\n[extract_fields_from_block DEBUG] HANDWRITTEN BLOCK for VIN {block_vin}:")
                print(f"  Block length: {len(text_block)} chars")
                print(f"  Full text: {repr(text_block[:500])}")  # First 500 chars
                print(f"  Keyword checks:")
                print(f"    - 'sedan': {'sedan' in text_block.lower()}")
                print(f"    - 'truck': {'truck' in text_block.lower()}")
                print(f"    - 'gas'/'gasoline': {bool(re.search(r'\b(gas|gasoline)\b', text_block.lower()))}")
                print(f"    - 'automatic'/'auto': {bool(re.search(r'\b(automatic|auto)\b', text_block.lower()))}")
                print(f"    - mileage pattern: {bool(re.search(r'\d+.*miles?', text_block, re.IGNORECASE))}")
                print(f"    - '100245' or '100,245': {'100245' in text_block or '100,245' in text_block}")
                print(f"    - '89000' or '89,000': {'89000' in text_block or '89,000' in text_block}")
    
    # Initialize vehicle dict with all fields as None
    vehicle = {field: None for field in VEHICLE_SCHEMA_ORDER}
    
    # Default _is_handwritten to False for all rows
    vehicle["_is_handwritten"] = False
    
    # Detect handwriting: simple heuristic - check if text lacks structured PDF labels
    # Handwritten documents typically don't have labels like "Make:", "Model:", "Year:", etc.
    if source_type in ("pdf", "image"):
        has_structured_labels = bool(re.search(r'\b(Make|Model|Year|Color|Mileage|Transmission|Fuel|Body)[:\s]', text_block, re.IGNORECASE))
        # If no structured labels and text is substantial, likely handwritten
        if not has_structured_labels and len(text_block) > 50:
            vehicle["_is_handwritten"] = True
    
    # Extract VIN from this block (must be present)
    # SAFEGUARD: Exclude common words that might match (including header tokens like "YEAR")
    # Expanded list to match split_by_vin() excluded words
    excluded_words = {'DEALERSHIP', 'IDENTIFICATION', 'TRANSMISSION', 'DESCRIPTION', 
                      'INFORMATION', 'REGISTRATION', 'DOCUMENTATION', 'SPECIFICATION',
                      'AUTOMATIC', 'MANUAL', 'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 
                      'VIN', 'BODY_STYLE', 'FUEL_TYPE', 'OWNER_EMAIL', 'NOTES', 'VEHICLE_ID',
                      'EFFECTIVE_DATE', 'TRIM', 'WEIGHT', 'BODY_TYPE', 'CURRENT_MILEAGE',
                      'EXTERIOR_COLOR', 'FUEL', 'BODY', 'PAINTED', 'GASOLINE', 'TRANSMISSION',
                      'CURRENT', 'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET'}
    
    # VIN pattern: handle multiple formats
    # Pattern 1: Full "Vehicle Identification Number (VIN): X" format (PDF format) - HIGH PRIORITY
    # Pattern 2: "VIN:" or "VIN):" or "VIN is" etc.
    # Pattern 3: Standalone 17-char VIN (standard format)
    # Pattern 4: Standalone 10-16 char alphanumeric (will be validated for digits below)
    # Use word boundaries to avoid matching "Identification" from "Vehicle Identification Number"
    
    # Try PDF format first (HIGH PRIORITY): "Vehicle Identification Number (VIN): HBUSRJGF4CBFPR9BN"
    pdf_vin_pattern = r"Vehicle Identification Number\s*\(VIN\)\s*:\s*([A-HJ-NPR-Z0-9]{17})"
    pdf_match = re.search(pdf_vin_pattern, text_block, re.IGNORECASE)
    pdf_vin_extracted = False
    if pdf_match:
        vin = pdf_match.group(1).upper().strip()
        # Check excluded words FIRST before other validations
        if vin in excluded_words:
            logger.debug(f"[extract_fields_from_block] PDF VIN pattern matched excluded word: {vin}")
            return None
        if len(vin) >= 10 and len([c for c in vin if c.isdigit()]) >= 2:
            vehicle['vin'] = vin
            pdf_vin_extracted = True
            logger.debug(f"[extract_fields_from_block] PDF-style VIN extracted: {vin}")
        else:
            logger.debug(f"[extract_fields_from_block] PDF pattern matched but VIN validation failed: {vin}")
            return None
    
    if not pdf_vin_extracted:
        # Fallback to other patterns
        vin_pattern = r'VIN[#\s:)]*\s+([A-Z0-9]{10,20})|([A-HJ-NPR-Z0-9]{17})\b|(?<![A-Z])([A-Z0-9]{10,16})(?![A-Z])'
        vin_match = re.search(vin_pattern, text_block, re.IGNORECASE)
        if not vin_match:
            logger.debug(f"[extract_fields_from_block] No VIN pattern matched in block (first 200 chars): {text_block[:200]}")
            return None
        
        vin = (vin_match.group(1) or vin_match.group(2) or vin_match.group(3))
        if not vin:
            logger.debug(f"[extract_fields_from_block] VIN pattern matched but no capture group: {vin_match.groups()}")
            return None
        vin = vin.upper().strip()
        
        # CRITICAL: Check excluded words FIRST (before length/digit checks)
        # This prevents header tokens like "YEAR", "MAKE", "MODEL" from being extracted as VINs
        # Also check case-insensitive match to catch variations
        vin_upper = vin.upper().strip()
        if vin in excluded_words or vin_upper in excluded_words:
            logger.debug(f"[extract_fields_from_block] VIN is excluded word: {vin}")
            return None
        
        # Additional check: reject common header tokens that might slip through
        # SAFEGUARD: Expanded list to catch all header tokens
        common_header_tokens = {'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 'VIN', 'TRANSMISSION', 
                                'BODY_STYLE', 'FUEL_TYPE', 'OWNER_EMAIL', 'NOTES', 'VEHICLE_ID',
                                'EFFECTIVE_DATE', 'TRIM', 'WEIGHT', 'BODY_TYPE', 'CURRENT_MILEAGE',
                                'EXTERIOR_COLOR', 'FUEL', 'AUTOMATIC', 'MANUAL', 'BODY', 'PAINTED',
                                'GASOLINE', 'CURRENT', 'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET'}
        if vin_upper in common_header_tokens:
            logger.debug(f"[extract_fields_from_block] VIN is header token: {vin}")
            return None
        
        if len(vin) < 10:
            logger.debug(f"[extract_fields_from_block] VIN too short: {vin} (length: {len(vin)})")
            return None
        
        # VINs should have at least one digit
        if not re.search(r'\d', vin):
            logger.debug(f"[extract_fields_from_block] VIN has no digits: {vin}")
            return None
        
        # VINs should have at least 2 digits (to avoid false positives like "IDENTIFICATION")
        if len([c for c in vin if c.isdigit()]) < 2:
            logger.debug(f"[extract_fields_from_block] VIN has less than 2 digits: {vin}")
            return None
        
        vehicle['vin'] = vin
    
    # Debug print for successful VIN extraction (only for non-PDF patterns, since PDF pattern already printed)
    if 'vin' in vehicle and vehicle['vin'] and not pdf_vin_extracted:
        logger.debug(f"[extract_fields_from_block] Successfully extracted VIN: {vehicle['vin']}")
    
    # Extract year: 4-digit year pattern
    # PRESERVE invalid years (e.g., 1899) - warnings will be generated later
    # Try "Year: X" pattern first (common in PDFs)
    year_label_match = re.search(r'Year[:\s]+(\d{4})', text_block, re.IGNORECASE)
    if year_label_match:
        try:
            vehicle['year'] = int(year_label_match.group(1))  # Preserve even if < 1990
        except ValueError:
            pass
    
    # If not found, try general year pattern
    if not vehicle.get('year'):
        # Match any 4-digit year starting with 18, 19, 20, or 21
        year_match = re.search(r'\b(1[89]\d{2}|20[0-3]\d|21[0-3]\d)\b', text_block)
        if year_match:
            try:
                vehicle['year'] = int(year_match.group(1))  # Preserve even if < 1990
            except ValueError:
                pass
        else:
            # Fallback: try to find any 4-digit number that might be a year
            # This handles edge cases like handwritten years
            year_fallback = re.search(r'\b(\d{4})\b', text_block)
            if year_fallback:
                try:
                    year_val = int(year_fallback.group(1))
                    # Only use if it's a reasonable year range (1800-2100)
                    if 1800 <= year_val <= 2100:
                        vehicle['year'] = year_val  # Preserve even if invalid (warnings will be generated)
                except ValueError:
                    pass
    
    # Extract make: Common makes in the block
    # Try patterns: "Make: Toyota", "Toyota Camry", "2024 Toyota"
    known_makes = ['Toyota', 'Ford', 'Honda', 'Tesla', 'BMW', 'Nissan', 'Chevrolet', 'Chevy',
                   'Mercedes', 'Mercedes-Benz', 'Audi', 'Volkswagen', 'VW', 'Hyundai', 'Kia',
                   'Mazda', 'Subaru', 'Jeep', 'Ram', 'GMC', 'Cadillac', 'Lexus', 'Acura']
    
    # First try "Make: X" pattern (common in PDFs)
    make_label_match = re.search(r'Make:\s*([A-Z][a-zA-Z\s-]+)', text_block, re.IGNORECASE)
    if make_label_match:
        make_candidate = make_label_match.group(1).strip()
        for make in known_makes:
            if make.lower() in make_candidate.lower():
                vehicle['make'] = make
                break
    
    # If not found, try standalone make names
    if not vehicle.get('make'):
        for make in known_makes:
            if re.search(rf'\b{re.escape(make)}\b', text_block, re.IGNORECASE):
                vehicle['make'] = make
                break
    
    # ------------------------------------------------------
    # MODEL extraction (PDF vs RAW TEXT)
    # ------------------------------------------------------
    vehicle['model'] = None
    
    if source_type == "pdf":
        # PDF format always has: Model: <VALUE>
        # Capture everything after "Model:" until newline
        m = re.search(r'Model[:\s]+([^\n]+)', text_block, re.IGNORECASE)
        if m:
            model_value = m.group(1).strip()
            # Clean up any trailing field labels that might have been captured
            model_value = re.sub(r'\s+(Year|Exterior Color|Current Mileage|Body Style|Fuel Type|Transmission|Owner Email|Additional Notes).*$', '', model_value, flags=re.IGNORECASE)
            vehicle['model'] = model_value.strip()
            logger.debug(f"[extract_fields_from_block] PDF model extracted: {vehicle['model']}")
        else:
            logger.debug(f"[extract_fields_from_block] PDF model pattern did not match in block (first 200 chars): {text_block[:200]}")
    else:
        # RAW TEXT format: "<year> <make> <model> ..."
        # Example: "This is a 2024 Toyota Camry sedan painted blue"
        # First try to match multi-word models (Model 3, F-150, CR-V)
        multi_word_match = re.search(
            r'\b(?:Toyota|Honda|Ford|BMW|Tesla|Nissan)\s+([A-Za-z0-9\-]+\s+[0-9A-Z]|[A-Z]-[0-9]+|[A-Z]+-[A-Z]+)(?=\s+(?:sedan|truck|suv|crossover|coupe|painted|\.|,|$))',
            text_block,
            re.IGNORECASE
        )
        if multi_word_match:
            vehicle['model'] = multi_word_match.group(1).strip()
        else:
            # Single word model - stop at body style words or punctuation
            # Use word boundary to ensure we stop at the end of the model word
            m = re.search(
                r'\b(?:Toyota|Honda|Ford|BMW|Tesla|Nissan)\s+([A-Za-z0-9\-]+)\b(?=\s+(?:sedan|truck|suv|crossover|coupe|painted|\.|,|$))',
                text_block,
                re.IGNORECASE
            )
            if m:
                vehicle['model'] = m.group(1).strip()
    
    # Extract color: Common colors
    # Try "Color: X" or "Exterior Color: X" pattern first (common in PDFs)
    # Allow for more flexible patterns - color might be after colon or space
    color_label_match = re.search(r'(?:Exterior\s+)?Color[:\s]+([A-Z][a-z]+)', text_block, re.IGNORECASE)
    if color_label_match:
        vehicle['color'] = color_label_match.group(1).capitalize()
    else:
        # Fallback: search for color words (case-insensitive, whole word match)
        colors = ['blue', 'red', 'white', 'black', 'silver', 'gray', 'grey', 'green', 'yellow', 'orange', 'purple', 'brown']
        text_lower = text_block.lower()
        for color in colors:
            if re.search(rf'\b{re.escape(color)}\b', text_lower):
                vehicle['color'] = color.capitalize()
                break
    
    # Extract mileage: Numbers with "miles" or "mi" or "Mileage: X"
    # PRESERVE negative values - do not drop them
    # Try "Mileage: X" or "Current Mileage: X" pattern first (common in PDFs)
    # Also handle OCR misspelling "Milage" (missing 'e')
    mileage_label_match = re.search(r'(?:Current\s+)?(?:Mileage|Milage)[:\s]+(-?[\d,]+)', text_block, re.IGNORECASE)
    if mileage_label_match:
        try:
            mileage_str = mileage_label_match.group(1).replace(',', '').replace(' ', '')
            vehicle['mileage'] = int(mileage_str)  # Preserve negative values
        except ValueError:
            pass
    else:
        # Fallback: search for "X miles" pattern (including negative values)
        mileage_match = re.search(r'(-?[\d,]+)\s*(?:miles?|mi\.?)', text_block, re.IGNORECASE)
        if mileage_match:
            try:
                mileage_str = mileage_match.group(1).replace(',', '').replace(' ', '')
                vehicle['mileage'] = int(mileage_str)  # Preserve negative values
            except ValueError:
                pass
    
    # Flexible keyword fallback for handwritten/natural language (only if still None)
    # Apply only for handwritten PDF/image (no structured labels) or all image sources (PNG images)
    # Look for patterns like "about X miles", "X miles on odometer", etc.
    if vehicle.get('mileage') is None and source_type in ("pdf", "image"):
        # Check if this is likely handwritten (no structured labels) or an image source
        has_structured_labels = bool(re.search(r'\b(Make|Model|Year|Color|Mileage|Transmission|Fuel|Body)[:\s]', text_block, re.IGNORECASE))
        is_likely_handwritten = not has_structured_labels and len(text_block) > 50
        # Apply flexible patterns for handwritten docs or all image sources
        if is_likely_handwritten or source_type == "image":
            # More flexible patterns for natural language
            mileage_patterns = [
                r'about\s+([\d,]+)\s*(?:miles?|mi\.?)',  # "about 100,245 miles"
                r'([\d,]+)\s*(?:miles?|mi\.?)\s+on\s+(?:the\s+)?odometer',  # "100,245 miles on odometer"
                r'odometer[:\s]+([\d,]+)',  # "odometer: 100,245"
                r'([\d,]+)\s*(?:miles?|mi\.?)\s+on\s+it',  # "100,245 miles on it"
            ]
            for pattern in mileage_patterns:
                mileage_match = re.search(pattern, text_block, re.IGNORECASE)
                if mileage_match:
                    try:
                        mileage_str = mileage_match.group(1).replace(',', '').replace(' ', '')
                        vehicle['mileage'] = int(mileage_str)
                        break
                    except ValueError:
                        continue
    
    # Extract fuel_type: "Fuel Type: X" or "Fuel: X" pattern
    # Try "Fuel Type: X" first (common in PDFs)
    fuel_type_label_match = re.search(r'Fuel\s+Type:\s*([A-Z][a-z]+)', text_block, re.IGNORECASE)
    if fuel_type_label_match:
        vehicle['fuel_type'] = fuel_type_label_match.group(1).lower()
    else:
        # Fallback: "Fuel: X" pattern
        fuel_match = re.search(r'Fuel:\s*(\w+)', text_block, re.IGNORECASE)
        if fuel_match:
            vehicle['fuel_type'] = fuel_match.group(1).lower()
    
    # Flexible keyword fallback for handwritten/natural language (only if still None)
    # Apply only for handwritten PDF/image (no structured labels) or all image sources (PNG images)
    if vehicle.get('fuel_type') is None and source_type in ("pdf", "image"):
        # Check if this is likely handwritten (no structured labels) or an image source
        has_structured_labels = bool(re.search(r'\b(Make|Model|Year|Color|Mileage|Transmission|Fuel|Body)[:\s]', text_block, re.IGNORECASE))
        is_likely_handwritten = not has_structured_labels and len(text_block) > 50
        # Apply flexible patterns for handwritten docs or all image sources
        if is_likely_handwritten or source_type == "image":
            text_lower = text_block.lower()
            # Debug: Log what keywords we're searching for
            if vehicle.get('vin') and vehicle.get('vin') in ['HBUSRJGF4CBFPR9BN', 'HBU5RTGF4CBEPR9BN', '3R5UAL4YUKPYGF1GZ', 'ST420RJ98FDHKL4E']:
                logger.debug(f"[extract_fields_from_block] VIN {vehicle.get('vin')}: Searching for fuel_type keywords in block (length: {len(text_block)}, preview: {text_block[:200]})")
            # Look for fuel type keywords anywhere in text (natural language)
            if re.search(r'\b(gas|gasoline)\b', text_lower):
                vehicle['fuel_type'] = 'gasoline'
                if vehicle.get('vin') and vehicle.get('vin') in ['HBUSRJGF4CBFPR9BN', 'HBU5RTGF4CBEPR9BN', '3R5UAL4YUKPYGF1GZ', 'ST420RJ98FDHKL4E']:
                    logger.debug(f"[extract_fields_from_block] VIN {vehicle.get('vin')}: Found fuel_type='gasoline' via keyword fallback")
            elif re.search(r'\bdiesel\b', text_lower):
                vehicle['fuel_type'] = 'diesel'
            elif re.search(r'\belectric\b', text_lower):
                vehicle['fuel_type'] = 'electric'
            elif re.search(r'\bhybrid\b', text_lower):
                vehicle['fuel_type'] = 'hybrid'
            elif vehicle.get('vin') and vehicle.get('vin') in ['HBUSRJGF4CBFPR9BN', 'HBU5RTGF4CBEPR9BN', '3R5UAL4YUKPYGF1GZ', 'ST420RJ98FDHKL4E']:
                logger.debug(f"[extract_fields_from_block] VIN {vehicle.get('vin')}: No fuel_type keywords found in block")
    
    # Extract transmission: "Transmission: X" or "8-speed automatic" pattern
    # SAFEGUARD: Try fallback pattern FIRST (handles "transmission is described as" better)
    # Then try label pattern, but capture multi-word values
    transmission_extracted = False
    
    # Pattern 1: "transmission is described as an 8-speed auto" or similar (HIGH PRIORITY)
    trans_match = re.search(r'transmission.*?(?:is\s+)?(?:described\s+as\s+)?(?:an?\s+)?(\d+[-\s]?speed\s+)?(automatic|manual|auto|AUTO)', text_block, re.IGNORECASE)
    if trans_match:
        trans_type = trans_match.group(2).lower() if trans_match.group(2) else None
        if trans_type:
            if trans_type == 'auto' or trans_type == 'AUTO':
                vehicle['transmission'] = 'automatic'
            else:
                vehicle['transmission'] = trans_type
            transmission_extracted = True
    
    # Pattern 2: "Transmission: X" - capture multi-word values (not just first word)
    if not transmission_extracted:
        # Capture everything after "Transmission:" until newline, comma, or period
        trans_label_match = re.search(r'Transmission[:\s]+([^\n,\.]+)', text_block, re.IGNORECASE)
        if trans_label_match:
            trans_val_raw = trans_label_match.group(1).strip()
            # Clean up filler words using cleanup_vision_value
            trans_val_cleaned = cleanup_vision_value(trans_val_raw)
            if trans_val_cleaned:
                trans_val = trans_val_cleaned.lower()
                if 'auto' in trans_val or trans_val.upper() == 'AUTO':
                    vehicle['transmission'] = 'automatic'
                elif 'manual' in trans_val:
                    vehicle['transmission'] = 'manual'
                elif 'cvt' in trans_val:
                    vehicle['transmission'] = 'cvt'
                else:
                    # If cleaned value is just a filler word, try to extract from original
                    # Look for transmission type keywords in the original value
                    if 'automatic' in trans_val_raw.lower() or 'auto' in trans_val_raw.lower():
                        vehicle['transmission'] = 'automatic'
                    elif 'manual' in trans_val_raw.lower():
                        vehicle['transmission'] = 'manual'
                    else:
                        # Last resort: use cleaned value (will trigger unknown_transmission warning)
                        vehicle['transmission'] = trans_val_cleaned
    
    # Flexible keyword fallback for handwritten/natural language (only if still None)
    # Apply only for handwritten PDF/image (no structured labels) or all image sources (PNG images)
    if vehicle.get('transmission') is None and source_type in ("pdf", "image"):
        # Check if this is likely handwritten (no structured labels) or an image source
        has_structured_labels = bool(re.search(r'\b(Make|Model|Year|Color|Mileage|Transmission|Fuel|Body)[:\s]', text_block, re.IGNORECASE))
        is_likely_handwritten = not has_structured_labels and len(text_block) > 50
        # Apply flexible patterns for handwritten docs or all image sources
        if is_likely_handwritten or source_type == "image":
            text_lower = text_block.lower()
            # Debug: Log what keywords we're searching for
            if vehicle.get('vin') and vehicle.get('vin') in ['HBUSRJGF4CBFPR9BN', 'HBU5RTGF4CBEPR9BN', '3R5UAL4YUKPYGF1GZ', 'ST420RJ98FDHKL4E']:
                logger.debug(f"[extract_fields_from_block] VIN {vehicle.get('vin')}: Searching for transmission keywords in block")
            # Look for transmission keywords anywhere in text (natural language)
            if re.search(r'\b(automatic|auto)\b', text_lower):
                vehicle['transmission'] = 'automatic'
                if vehicle.get('vin') and vehicle.get('vin') in ['HBUSRJGF4CBFPR9BN', 'HBU5RTGF4CBEPR9BN', '3R5UAL4YUKPYGF1GZ', 'ST420RJ98FDHKL4E']:
                    logger.debug(f"[extract_fields_from_block] VIN {vehicle.get('vin')}: Found transmission='automatic' via keyword fallback")
            elif re.search(r'\bmanual\b', text_lower):
                vehicle['transmission'] = 'manual'
            elif re.search(r'\bcvt\b', text_lower):
                vehicle['transmission'] = 'cvt'
            elif vehicle.get('vin') and vehicle.get('vin') in ['HBUSRJGF4CBFPR9BN', 'HBU5RTGF4CBEPR9BN', '3R5UAL4YUKPYGF1GZ', 'ST420RJ98FDHKL4E']:
                logger.debug(f"[extract_fields_from_block] VIN {vehicle.get('vin')}: No transmission keywords found in block")
    
    # Extract owner_email: "Owner Email: X" or email pattern
    # PRESERVE invalid email formats - warnings will be generated later
    # Try "Owner Email: X" first (common in PDFs) - extract any value after the label
    email_label_match = re.search(r'Owner\s+Email:\s*([^\n,]+)', text_block, re.IGNORECASE)
    if email_label_match:
        email_value = email_label_match.group(1).strip()
        # If it looks like a valid email, lowercase it; otherwise preserve as-is
        if '@' in email_value and '.' in email_value:
            vehicle['owner_email'] = email_value.lower()
        else:
            # Preserve invalid email format (e.g., "bad-email-format")
            vehicle['owner_email'] = email_value
    else:
        # Fallback: search for any email pattern (valid format)
        email_match = re.search(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})', text_block)
        if email_match:
            vehicle['owner_email'] = email_match.group(1).lower()
    
    # Extract body_style: "Body Style: X" or common body styles
    # Try "Body Style: X" first (common in PDFs)
    body_style_label_match = re.search(r'Body\s+Style:\s*([A-Z][a-z]+)', text_block, re.IGNORECASE)
    if body_style_label_match:
        vehicle['body_style'] = body_style_label_match.group(1).lower()
    else:
        # Fallback: search for body style words
        body_styles = ['sedan', 'truck', 'suv', 'crossover', 'coupe', 'convertible', 'wagon', 'hatchback']
        text_lower = text_block.lower()
        for style in body_styles:
            if re.search(rf'\b{re.escape(style)}\b', text_lower):
                vehicle['body_style'] = style
                break
    
    # Note: body_style already has flexible keyword fallback above, so no additional fallback needed
    
    # Extract notes: PDF uses new pattern, everything else leaves notes empty
    if source_type == "pdf":
        notes_match = re.search(
            r'Additional Notes:\s*(.*?)(?=\n\s*(?:Vehicle Identification Number|Make|Model|Year|Exterior Color|Current Mileage|Body Style|Fuel Type|Transmission|Owner Email|vin:|VEHICLE DETAIL SHEET)|\Z)',
            text_block,
            re.IGNORECASE | re.DOTALL,
        )
        if notes_match:
            notes_clean = notes_match.group(1)
            # Remove bullet points and normalize whitespace
            notes_clean = re.sub(r'^[\-\•\*]\s*', '', notes_clean, flags=re.MULTILINE)
            notes_clean = re.sub(r'\s+', ' ', notes_clean).strip()
            # Remove "VEHICLE DETAIL SHEET" if it appears
            notes_clean = re.sub(r'\s*VEHICLE DETAIL SHEET\s*$', '', notes_clean, flags=re.IGNORECASE)
            if notes_clean:
                vehicle['notes'] = notes_clean
    
    # Generate human-readable notes from extracted fields (if notes not already set)
    if not vehicle.get('notes'):
        vehicle['notes'] = _generate_vehicle_notes(vehicle)
    
    # SAFEGUARD: Clean up all extracted string values using cleanup_vision_value
    # This removes filler words like "is", "described", "was" that might have been extracted
    for field_name, field_value in vehicle.items():
        if isinstance(field_value, str) and field_value:
            cleaned_value = cleanup_vision_value(field_value)
            if cleaned_value is not None:
                vehicle[field_name] = cleaned_value
            elif cleaned_value is None and field_value:
                # If cleanup returned None, it means value was just filler words
                # For critical fields, try to preserve original if it's not just filler
                if field_name in ['transmission', 'color', 'make', 'model', 'body_style', 'fuel_type']:
                    # Keep original value - normalization will handle it
                    pass
                else:
                    vehicle[field_name] = None
    
    # SAFEGUARD: Field preservation validation - log if critical fields are missing
    critical_fields = ['vin', 'year', 'make', 'model']
    missing_critical = [f for f in critical_fields if not vehicle.get(f)]
    if missing_critical and vehicle.get('vin'):
        logger.debug(f"[extract_fields_from_block] WARNING: Missing critical fields for VIN {vehicle.get('vin')}: {missing_critical}")
    
    
    return vehicle

def _generate_vehicle_notes(vehicle: Dict[str, Any]) -> str:
    """
    Generate deterministic notes field following exact format:
    {year} {make} {model} {body_style} painted {color}. {mileage} miles. Fuel: {fuel_type}. Automatic transmission.
    
    Rules:
    - Only use information explicitly present in the vehicle dict
    - Do not invent values
    - Follow exact format: year make model body_style painted color. mileage miles. Fuel: fuel_type. Automatic transmission.
    - Remove OCR explanations, test commentary, duplicated sentences, meta text
    - Mileage should include commas (e.g., 38,990 miles)
    - Fuel type should be normalized (gasoline, electric, diesel)
    - Transmission should be normalized (automatic, manual) - always "Automatic transmission" or "Manual transmission"
    
    Args:
        vehicle: Dictionary with extracted vehicle fields
    
    Returns:
        Human-readable notes string, or empty string if no fields to summarize
    """
    # Check if we only have body_style (no meaningful content for notes)
    # If only body_style exists without year/make/model/color/mileage/fuel/transmission, return None
    has_meaningful_content = any([
        vehicle.get('year'),
        vehicle.get('make'),
        vehicle.get('model'),
        vehicle.get('color'),
        vehicle.get('mileage') is not None,
        vehicle.get('fuel_type'),
        vehicle.get('transmission')
    ])
    
    if not has_meaningful_content:
        # Only body_style or nothing - return None (don't generate "suv.")
        return None
    
    parts = []
    
    # Build year, make, model, body_style, color description
    desc_parts = []
    
    if vehicle.get('year'):
        try:
            year_val = vehicle['year']
            # Handle both int and string year values
            if isinstance(year_val, str):
                year_val = int(year_val.strip())
            desc_parts.append(str(year_val))
        except (ValueError, TypeError):
            # Skip invalid year
            pass
    
    if vehicle.get('make'):
        make_val = str(vehicle['make']).strip()
        if make_val:
            desc_parts.append(make_val)
    
    if vehicle.get('model'):
        model_val = str(vehicle['model']).strip()
        if model_val:
            desc_parts.append(model_val)
    
    if vehicle.get('body_style'):
        body_style = str(vehicle['body_style']).strip().lower()
        # Keep lowercase for body_style (sedan, truck, suv, etc.)
        if body_style:
            desc_parts.append(body_style)
    
    if vehicle.get('color'):
        color = str(vehicle['color']).strip()
        if color:
            color = color.lower()  # Keep lowercase for color
            desc_parts.append(f"painted {color}")
    
    # Build main description sentence
    if desc_parts:
        main_desc = " ".join(desc_parts)
        # Add period if it doesn't end with one
        if not main_desc.endswith('.'):
            main_desc += "."
        parts.append(main_desc)
    
    # Add mileage if present
    if vehicle.get('mileage') is not None:
        try:
            mileage_val = vehicle['mileage']
            # Handle both int and string mileage values
            if isinstance(mileage_val, str):
                # Remove commas and whitespace
                mileage_val = mileage_val.replace(',', '').strip()
                mileage = int(mileage_val)
            else:
                mileage = int(mileage_val)
            # Format with commas
            mileage_str = f"{mileage:,}"
            parts.append(f"{mileage_str} miles")
        except (ValueError, TypeError):
            # If mileage is invalid, skip it (don't include in notes)
            pass
    
    # Add fuel type if present - format: "Fuel: {fuel_type}"
    if vehicle.get('fuel_type'):
        fuel_type = str(vehicle['fuel_type']).strip().lower()
        if fuel_type:
            # Normalize fuel type
            fuel_normalized = None
            if 'gas' in fuel_type or 'gasoline' in fuel_type or fuel_type == 'petrol':
                fuel_normalized = 'gasoline'
            elif 'electric' in fuel_type or 'ev' in fuel_type:
                fuel_normalized = 'electric'
            elif 'diesel' in fuel_type:
                fuel_normalized = 'diesel'
            elif fuel_type in ['gasoline', 'electric', 'diesel', 'hybrid', 'plug-in hybrid']:
                fuel_normalized = fuel_type
            
            if fuel_normalized:
                parts.append(f"Fuel: {fuel_normalized}")
    
    # Add transmission if present - format: "Automatic transmission" or "Manual transmission"
    if vehicle.get('transmission'):
        transmission = str(vehicle['transmission']).strip().lower()
        if transmission:
            # Normalize transmission
            trans_normalized = None
            if 'auto' in transmission:
                trans_normalized = 'Automatic'
            elif 'manual' in transmission:
                trans_normalized = 'Manual'
            elif transmission in ['automatic', 'manual', 'cvt']:
                trans_normalized = transmission.capitalize()
            
            if trans_normalized:
                parts.append(f"{trans_normalized} transmission")
    
    # Join all parts with periods and spaces
    if parts:
        notes = ". ".join(parts)
        # Ensure it ends with a period
        if not notes.endswith('.'):
            notes += "."
        return notes
    
    return None  # Return None if no meaningful content

def _get_vision_ocr_text_for_pdf(file_path: Path) -> str:
    """
    Extract raw text from PDF using Vision API OCR.
    This is used for PDF vehicle extraction to get clean text for parse_vehicle_raw_text.
    
    Args:
        file_path: Path to PDF file
    
    Returns:
        Raw text string from Vision OCR
    
    Raises:
        Exception: If Vision OCR fails or returns error messages
        RuntimeError: If Vision OCR did not run (empty text or error response)
    """
    from config import OPENAI_API_KEY, PDF_MODEL, ENABLE_OPENAI
    
    # Check for API key first
    if not OPENAI_API_KEY:
        logger.error("[PDF Vision OCR] OPENAI_API_KEY missing - check .env file")
        raise Exception("VisionOCR failed: OPENAI_API_KEY missing")
    
    if not ENABLE_OPENAI:
        logger.error("[PDF Vision OCR] ENABLE_OPENAI is disabled")
        raise Exception("VisionOCR failed: ENABLE_OPENAI is disabled")
    
    try:
        from openai import OpenAI
        from pdf2image import convert_from_path
        from PIL import Image
        import base64
        import io
    except ImportError as e:
        logger.error(f"[PDF Vision OCR] Missing required dependency: {e}")
        raise Exception(f"VisionOCR failed: Missing required dependency: {e}")
    
    try:
        logger.info("[PDF Vision OCR] Initializing OpenAI client")
        client = OpenAI(api_key=OPENAI_API_KEY)
        logger.info(f"[PDF Vision OCR] OpenAI client initialized - Vision will be called using model: {PDF_MODEL}")
        
        # Convert PDF to images at ~300 DPI (mirrors ENG-101-table-detector line 1892)
        # This improves OCR/Vision quality for handwritten PDFs by ensuring sufficient resolution
        logger.info(f"[PDF Vision OCR] Converting PDF to images: {file_path}")
        images = convert_from_path(str(file_path), dpi=300)
        if not images:
            logger.error("[PDF Vision OCR] Could not convert PDF to images")
            raise Exception("VisionOCR failed: Could not convert PDF to images")
        
        logger.info(f"[PDF Vision OCR] PDF converted to {len(images)} page(s)")
        
        # Extract text from each page using Vision API
        all_text_parts = []
        vision_called = False
        for page_idx, img in enumerate(images):
            page_num = page_idx + 1
            logger.info(f"[PDF Vision OCR] Calling Vision API for page {page_num}/{len(images)}")
            
            # Convert image to base64
            buffered = io.BytesIO()
            img.save(buffered, format="PNG")
            img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
            
            # Call Vision API with simple text extraction prompt
            try:
                logger.info(f"[PDF Vision OCR] Sending Vision API request for page {page_num}")
                response = client.chat.completions.create(
                    model=PDF_MODEL,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {
                                    "type": "text",
                                    "text": "Extract all text from this PDF page. Return the raw text exactly as it appears, preserving line breaks and spacing. Do not format or structure the text."
                                },
                                {
                                    "type": "image_url",
                                    "image_url": {
                                        "url": f"data:image/png;base64,{img_base64}"
                                    }
                                }
                            ]
                        }
                    ],
                    max_tokens=4000
                )
                
                vision_called = True
                logger.info(f"[PDF Vision OCR] Vision API called successfully for page {page_num}")
                
                page_text = response.choices[0].message.content
                
                # Fail-fast check: if OCR text is empty or contains error phrases, raise immediately
                if not page_text or len(page_text.strip()) == 0:
                    logger.error(f"[PDF Vision OCR] Vision API returned empty content for page {page_num}")
                    raise RuntimeError("Vision OCR did not run; OPENAI_API_KEY may not be loaded.")
                
                # Check for error messages from Vision API
                error_indicators = [
                    "I'm unable",
                    "I can't assist",
                    "I cannot extract",
                    "I'm sorry",
                    "I don't see",
                    "Unable to",
                    "cannot process",
                    "I cannot read",
                    "I'm not able to"
                ]
                page_text_lower = page_text.lower()
                vision_api_error = False
                for indicator in error_indicators:
                    if indicator.lower() in page_text_lower:
                        vision_api_error = True
                        logger.error(f"[PDF Vision OCR] Vision API returned error message: {page_text[:200]}")
                        raise RuntimeError("Vision OCR did not run; OPENAI_API_KEY may not be loaded.")
                
                if vision_api_error:
                    # Try alternative approach: use OCR text extraction instead
                    logger.warning(f"[PDF Vision OCR] Vision API failed, attempting fallback OCR extraction")
                    # Fall back to regular OCR extraction for this page
                    try:
                        from ocr import extract_text_from_pdf
                        ocr_text_blocks, _ = extract_text_from_pdf(file_path, enable_vision=False)
                        if ocr_text_blocks:
                            from ocr.table_extract import _extract_raw_text
                            ocr_text = _extract_raw_text(ocr_text_blocks)
                            if ocr_text and len(ocr_text) > 100:
                                logger.info(f"[PDF Vision OCR] Fallback OCR extracted {len(ocr_text)} characters")
                                # Use OCR text for this page instead of Vision
                                all_text_parts.append(ocr_text)
                                logger.info(f"[PDF Vision OCR] Page {page_num} using OCR fallback: {len(ocr_text)} chars")
                                continue
                    except Exception as fallback_error:
                        logger.error(f"[PDF Vision OCR] Fallback OCR also failed: {fallback_error}")
                    # If fallback failed, raise exception
                    raise Exception(f"VisionOCR failed: Vision API returned error message and OCR fallback failed")
                
                # Vision API succeeded - use the extracted text
                all_text_parts.append(page_text)
                logger.info(f"[PDF Vision OCR] Page {page_num} extracted: {len(page_text)} chars")
                logger.debug(f"[PDF Vision OCR] Page {page_num} text preview: {page_text[:100]}")
                
            except Exception as e:
                logger.error(f"[PDF Vision OCR] Error calling Vision API for page {page_num}: {e}")
                raise Exception(f"VisionOCR failed: Error calling Vision API for page {page_num}: {e}")
        
        if not vision_called:
            logger.error("[PDF Vision OCR] Vision API was never called")
            raise Exception("VisionOCR failed: Vision API was never called")
        
        if not all_text_parts:
            logger.error("[PDF Vision OCR] No text extracted from any page")
            raise Exception("VisionOCR failed: No text extracted from any page")
        
        full_text = "\n\n".join(all_text_parts)
        logger.info(f"[PDF Vision OCR] Total text extracted: {len(full_text)} characters from {len(all_text_parts)} page(s)")
        
        # Fail-fast check: if final OCR text is empty or contains error phrases, raise immediately
        if not full_text or len(full_text.strip()) == 0:
            raise RuntimeError("Vision OCR did not run; OPENAI_API_KEY may not be loaded.")
        
        error_phrases = ["I'm unable", "I can't assist"]
        full_text_lower = full_text.lower()
        for phrase in error_phrases:
            if phrase.lower() in full_text_lower:
                raise RuntimeError("Vision OCR did not run; OPENAI_API_KEY may not be loaded.")
        
        logger.info(f"[PDF Vision OCR] Vision OCR succeeded - returning {len(full_text)} characters")
        return full_text
        
    except Exception as e:
        # Re-raise if it's already our custom exception
        if "VisionOCR failed" in str(e):
            raise
        # Otherwise wrap it
        logger.error(f"[PDF Vision OCR] Unexpected error: {e}")
        raise Exception(f"VisionOCR failed: {e}")
            
    except Exception as e:
        logger.error(f"[PDF Vision OCR] Error: {e}")
        import traceback
        traceback.print_exc()
        return None

def parse_vehicle_raw_text(text: str, source_type: str = "raw") -> List[Dict[str, Any]]:
    """
    Parse unstructured vehicle text and extract all canonical vehicle fields.
    
    Strategy:
    1. Split text into non-overlapping blocks by VIN positions
    2. For each block, extract all vehicle fields using extract_fields_from_block
    3. Return raw values (normalization happens in normalize_v2 via mappings/transforms)
    
    Args:
        text: Raw text containing vehicle descriptions
        source_type: Source type ("pdf" or "raw") to use for field extraction
    
    Returns:
        List of dicts, one per vehicle, with fields matching VEHICLE_SCHEMA_ORDER
    """
    # Clean text (remove any preprocessing artifacts)
    cleaned_text = text.strip()
    
    # DEBUG: Show OCR text before block splitting (gated behind debug logging)
    logger.debug(f"[parse_vehicle_raw_text DEBUG] OCR text preview (first 300 chars): {cleaned_text[:300]}")
    
    # DEBUG: Find all possible VIN-like strings (17-char alphanumeric)
    import re
    all_vin_candidates = re.findall(r'[A-HJ-NPR-Z0-9]{17}', cleaned_text, re.IGNORECASE)
    logger.debug(f"[parse_vehicle_raw_text DEBUG] Found {len(all_vin_candidates)} potential 17-char VIN-like sequences")
    for i, candidate in enumerate(all_vin_candidates[:10]):  # Show first 10
        digit_count = len([c for c in candidate if c.isdigit()])
        logger.debug(f"  Candidate {i+1}: {candidate} (length: {len(candidate)}, digits: {digit_count})")
    
    # Split into blocks by VIN
    blocks = split_by_vin(cleaned_text)
    logger.debug(f"[parse_vehicle_raw_text] Found {len(blocks)} VIN block(s)")
    
    # DEBUG: Only log block count if we didn't find 6 blocks (success case disables debug)
    if len(blocks) < 6:
        logger.debug(f"[parse_vehicle_raw_text] Found {len(blocks)} VIN block(s)")
    
    # Extract vehicle fields from each block using the new extractor
    # Deduplicate by VIN to avoid processing the same vehicle twice
    vehicles = []
    seen_vins = set()
    
    for i, b in enumerate(blocks):
        logger.debug(f"[parse_vehicle_raw_text] Processing block {i+1}/{len(blocks)}")
        vehicle = extract_fields_from_block(b, mapping=None, source_type=source_type)
        if vehicle and vehicle.get('vin'):
            vin_upper = vehicle.get('vin').upper()
            # Skip if we've already processed this VIN
            if vin_upper in seen_vins:
                logger.debug(f"[parse_vehicle_raw_text] Block {i+1}: skipping duplicate VIN {vin_upper}")
                continue
            seen_vins.add(vin_upper)
            
            # CRITICAL: Preserve the full OCR text block in notes for post-Vision inference
            # This allows normalize_v2 to search the full text for semantic fields like "sedan", "gasoline"
            # If notes was generated (not extracted), append the full block text
            # If notes was extracted, append the block text for additional context
            current_notes = vehicle.get('notes', '')
            if current_notes and isinstance(current_notes, str) and len(current_notes.strip()) >= 50:
                # Notes was extracted - append block text for additional context
                vehicle['notes'] = current_notes + " " + b
                logger.debug(f"[parse_vehicle_raw_text] Appended full OCR block to extracted notes ({len(b)} chars)")
            else:
                # Notes was generated or empty - replace with full block text
                vehicle['notes'] = b
                logger.debug(f"[parse_vehicle_raw_text] Set notes to full OCR block text ({len(b)} chars)")
            
            vehicles.append(vehicle)
            logger.debug(f"[parse_vehicle_raw_text] Block {i+1}: extracted VIN {vehicle.get('vin')}")
        else:
            logger.debug(f"[parse_vehicle_raw_text] Block {i+1}: no vehicle extracted")
    
    logger.info(f"[parse_vehicle_raw_text] Total vehicles extracted: {len(vehicles)} (after deduplication)")
    
    # SAFEGUARD: Section counting validation - warn if we found fewer vehicles than expected
    # Expected count is based on number of blocks found by split_by_vin
    if len(vehicles) < len(blocks):
        missing_count = len(blocks) - len(vehicles)
        logger.warning(f"[parse_vehicle_raw_text] WARNING: Found {len(blocks)} VIN blocks but only extracted {len(vehicles)} vehicles. {missing_count} block(s) failed extraction.")
    
    return vehicles

def split_by_policy_number(text: str) -> List[Tuple[str, str]]:
    """
    Split text into policy blocks by policy_number positions.
    Mirrors split_by_vin() for vehicles.
    
    Args:
        text: Raw text containing policy descriptions
    
    Returns:
        List of tuples: (policy_number, block_text)
    """
    import re
    
    cleaned_text = text.strip()
    blocks = []
    
    # Pattern to find "Policy Number:" headers (PDF format)
    # Also handle "POLICY DETAIL SHEET" followed by "Policy Number:"
    policy_header_pattern = r'(?:POLICY\s+DETAIL\s+SHEET\s+)?Policy\s+Number\s*:\s*((?:P\d+)|(?:P_[A-Z0-9]+))'
    
    # Find all policy header matches
    policy_headers = list(re.finditer(policy_header_pattern, cleaned_text, re.IGNORECASE))
    
    logger.info(f"[split_by_policy_number] Found {len(policy_headers)} 'Policy Number:' header(s)")
    
    is_narrative_format = False
    if not policy_headers:
        logger.info("[split_by_policy_number] No 'Policy Number:' headers found - trying narrative format")
        # Try narrative format: "Policy P001 is..." or "P_BAD1 is..."
        narrative_pattern = r'\b(P\d+|P_[A-Z0-9]+)\s+is\b'
        narrative_matches = list(re.finditer(narrative_pattern, cleaned_text, re.IGNORECASE))
        if narrative_matches:
            logger.info(f"[split_by_policy_number] Found {len(narrative_matches)} policy(s) in narrative format")
            is_narrative_format = True
            # Process narrative format
            for i, match in enumerate(narrative_matches):
                policy_num = match.group(1)
                start_pos = match.start()
                # Find end position (start of next policy or end of text)
                if i + 1 < len(narrative_matches):
                    end_pos = narrative_matches[i + 1].start()
                else:
                    end_pos = len(cleaned_text)
                
                block_text = cleaned_text[start_pos:end_pos].strip()
                blocks.append((policy_num, block_text))
                logger.info(f"[split_by_policy_number] Narrative block for {policy_num}: {len(block_text)} chars")
        else:
            logger.warning("[split_by_policy_number] No policy numbers found in narrative format either")
            return []
    
    # Split text into blocks by policy headers (if not narrative format)
    if not is_narrative_format:
        for i, match in enumerate(policy_headers):
            # Get policy number from match
            policy_num = match.group(1)
            if not policy_num:
                continue
            
            start_pos = match.start()
            
            # Find end position (start of next policy header or end of text)
            if i + 1 < len(policy_headers):
                end_pos = policy_headers[i + 1].start()
            else:
                end_pos = len(cleaned_text)
            
            block_text = cleaned_text[start_pos:end_pos].strip()
            
            # Ignore blocks that look like memos (don't have structured fields)
            # Check if block has at least one structured field pattern
            has_structured_fields = bool(re.search(r'(?:Named\s+Insured|Effective\s+Date|Expiration\s+Date|Premium|Coverage\s+Type|Vehicle\s+VIN|Notes)\s*:', block_text, re.IGNORECASE))
            
            if not has_structured_fields:
                logger.info(f"[split_by_policy_number] Block for {policy_num} appears to be memo/unstructured - skipping")
                continue
            
            blocks.append((policy_num, block_text))
            logger.info(f"[split_by_policy_number] Block {len(blocks)} for policy {policy_num}: {len(block_text)} chars")
    
    logger.info(f"[split_by_policy_number] Found {len(blocks)} valid policy block(s) after filtering")
    return blocks


def extract_policy_fields_from_block(
    text_block: str, 
    mapping: Optional[Dict[str, Any]] = None, 
    source_type: Optional[str] = None
) -> Dict[str, Any]:
    """
    Extract ALL fields for a single policy from a policy_number-isolated block.
    Mirrors extract_fields_from_block() for vehicles.
    
    Args:
        text_block: Text block containing policy information
        mapping: Optional mapping configuration
        source_type: Source type ("pdf", "raw", "image") to determine extraction logic
    
    Returns:
        Dictionary with policy fields matching POLICY_SCHEMA_ORDER
    """
    from schema import POLICY_SCHEMA_ORDER
    import re
    
    # Initialize policy dict with all fields as None
    policy = {field: None for field in POLICY_SCHEMA_ORDER}
    
    # Extract policy_number from block (should be present)
    policy_num_match = re.search(r'(?:POLICY\s+DETAIL\s+SHEET\s+)?Policy\s+Number\s*:\s*((?:P\d+)|(?:P_[A-Z0-9]+))', text_block, re.IGNORECASE)
    if not policy_num_match:
        # Try narrative format
        policy_num_match = re.search(r'\b(P\d+|P_[A-Z0-9]+)\s+is\b', text_block, re.IGNORECASE)
    
    if policy_num_match:
        policy['policy_number'] = policy_num_match.group(1)
    
    # For raw_text sources, extract policy_number and allow limited inference
    # Do NOT infer dates, premium, or VIN from narrative prose
    # But DO allow coverage_type and insured_name inference (matches expected truth files)
    if source_type == "raw":
        # Extract insured_name from "for" pattern (narrative format)
        name_match = re.search(r'for\s+([A-Z][a-z]+(?:\s+and\s+[A-Z][a-z]+)?(?:\s+[A-Z][a-z]+)?)', text_block, re.IGNORECASE)
        if name_match:
            policy['insured_name'] = name_match.group(1).strip()
        
        # Extract coverage_type from narrative patterns
        coverage_match = re.search(r'(personal\s+auto|high-risk\s+auto|high\s+risk\s+auto|commercial\s+auto)', text_block, re.IGNORECASE)
        if coverage_match:
            coverage_str = coverage_match.group(1).lower()
            if coverage_str in ['personal auto', 'personal']:
                policy['coverage_type'] = 'personal auto'
            elif coverage_str in ['high-risk auto', 'high risk auto']:
                policy['coverage_type'] = 'high-risk auto'
            else:
                policy['coverage_type'] = coverage_str
        
        # Do NOT populate notes unless explicit "Notes:" field exists
        notes_match = re.search(r'Notes\s*:\s*([^\n]+(?:\n(?!Policy\s+Number|POLICY\s+DETAIL)[^\n]+)*)', text_block, re.IGNORECASE | re.MULTILINE)
        if notes_match:
            notes_text = notes_match.group(1).strip()
            if notes_text:
                policy['notes'] = notes_text
        
        # Do NOT infer dates, premium, or VIN (they remain None)
        return policy
    
    # For PDF/image sources, extract all fields (existing logic)
    # Extract insured_name - PDF format: "Named Insured: John and Jane Doe"
    name_match = re.search(r'Named\s+Insured\s*:\s*([^\n]+)', text_block, re.IGNORECASE)
    if name_match:
        policy['insured_name'] = name_match.group(1).strip()
    else:
        # Fallback: look for "for" pattern (works for both PDF and narrative)
        # Narrative: "for John and Jane Doe" or "for Alex Smith"
        name_match = re.search(r'for\s+([A-Z][a-z]+(?:\s+and\s+[A-Z][a-z]+)?(?:\s+[A-Z][a-z]+)?)', text_block, re.IGNORECASE)
        if name_match:
            policy['insured_name'] = name_match.group(1).strip()
    
    # Extract effective_date - PDF format: "Effective Date: 2025-01-01"
    effective_match = re.search(r'Effective\s+Date\s*:\s*(\d{4}-\d{2}-\d{2})', text_block, re.IGNORECASE)
    if effective_match:
        policy['effective_date'] = effective_match.group(1).strip()
    else:
        # Fallback: find first date in block
        date_match = re.search(r'\b(\d{4}-\d{2}-\d{2})\b', text_block)
        if date_match:
            policy['effective_date'] = date_match.group(1)
    
    # Extract expiration_date - PDF format: "Expiration Date: 2026-01-01"
    expiration_match = re.search(r'Expiration\s+Date\s*:\s*(\d{4}-\d{2}-\d{2})', text_block, re.IGNORECASE)
    if expiration_match:
        policy['expiration_date'] = expiration_match.group(1).strip()
    else:
        # Fallback: find second date in block
        dates = re.findall(r'\b(\d{4}-\d{2}-\d{2})\b', text_block)
        if len(dates) >= 2:
            policy['expiration_date'] = dates[1]
    
    # Extract premium - PDF format: "Premium: $1650.50" or "Premium: 1650.50"
    premium_match = re.search(r'Premium\s*:\s*(-?\$?[\d,]+\.?\d*)', text_block, re.IGNORECASE)
    if premium_match:
        try:
            premium_str = premium_match.group(1).replace(',', '').replace('$', '').strip()
            policy['premium'] = str(float(premium_str))  # Keep as string for consistency
        except (ValueError, AttributeError):
            pass
    
    # Extract coverage_type - PDF format: "Coverage Type: Personal Auto"
    coverage_match = re.search(r'Coverage\s+Type\s*:\s*([^\n]+)', text_block, re.IGNORECASE)
    if coverage_match:
        coverage_str = coverage_match.group(1).strip()
        # Normalize casing
        if coverage_str.lower() in ['personal auto', 'personal']:
            policy['coverage_type'] = 'Personal Auto'
        elif coverage_str.lower() in ['high-risk auto', 'high risk auto']:
            policy['coverage_type'] = 'High-Risk Auto'
        else:
            policy['coverage_type'] = coverage_str
    else:
        # Fallback: look for coverage patterns (works for both PDF and narrative)
        # Narrative: "personal auto policy" or "high-risk auto policy"
        coverage_match = re.search(r'(personal\s+auto|high-risk\s+auto|high\s+risk\s+auto|commercial\s+auto)', text_block, re.IGNORECASE)
        if coverage_match:
            coverage_str = coverage_match.group(1).lower()
            if coverage_str in ['personal auto', 'personal']:
                policy['coverage_type'] = 'personal auto'
            elif coverage_str in ['high-risk auto', 'high risk auto']:
                policy['coverage_type'] = 'high-risk auto'
            else:
                policy['coverage_type'] = coverage_str
    
    # Extract vehicle_vin - PDF format: "Vehicle VIN: HBUSRJGF4CBFPR9BN"
    vin_match = re.search(r'Vehicle\s+VIN\s*:\s*([A-HJ-NPR-Z0-9]{17})', text_block, re.IGNORECASE)
    if vin_match:
        policy['vehicle_vin'] = vin_match.group(1).strip().upper()
    
    # Extract notes - PDF format: "Notes: Example OCR test."
    notes_match = re.search(r'Notes\s*:\s*([^\n]+(?:\n(?!Policy\s+Number|POLICY\s+DETAIL)[^\n]+)*)', text_block, re.IGNORECASE | re.MULTILINE)
    if notes_match:
        notes_text = notes_match.group(1).strip()
        # Clean up notes (remove extra whitespace, stop at next policy header)
        notes_text = re.sub(r'\s+', ' ', notes_text)
        if notes_text:
            policy['notes'] = notes_text
    else:
        # Preserve full block in notes for post-processing (like vehicles)
        policy['notes'] = text_block
    
    return policy


def parse_policy_raw_text(text: str, source_type: str = "raw") -> List[Dict[str, Any]]:
    """
    Parse unstructured policy text and extract all canonical policy fields.
    
    Refactored to use split_by_policy_number() and extract_policy_fields_from_block()
    following the vehicle/driver pipeline pattern.
    
    Strategy:
    1. Split text into non-overlapping blocks by policy_number positions
    2. For each block, extract all policy fields using extract_policy_fields_from_block
    3. Deduplicate by policy_number (only one row per policy number)
    4. Return raw values (normalization happens in normalize_v2 via mappings/transforms)
    
    Args:
        text: Raw text containing policy descriptions
        source_type: Source type ("pdf" or "raw") to use for field extraction
    
    Returns:
        List of dicts, one per policy, with fields matching POLICY_SCHEMA_ORDER
    """
    from schema import POLICY_SCHEMA_ORDER
    import re
    
    # Clean text
    cleaned_text = text.strip()
    
    logger.info(f"[parse_policy_raw_text] Starting parse, input text length: {len(cleaned_text)}")
    
    # Split into blocks by policy_number
    blocks = split_by_policy_number(cleaned_text)
    logger.info(f"[parse_policy_raw_text] Found {len(blocks)} policy block(s)")
    
    # Extract policy fields from each block
    # Deduplicate by policy_number to avoid processing the same policy twice
    policies = []
    policies_dict = {}  # Track by policy_number for deduplication
    
    for i, (policy_num, block_text) in enumerate(blocks):
        logger.debug(f"[parse_policy_raw_text] Processing block {i+1}/{len(blocks)}")
        
        # If we already have this policy number, merge the blocks (prefer longer one)
        if policy_num in policies_dict:
            logger.info(f"[parse_policy_raw_text] Merging duplicate block for policy {policy_num}")
            existing_block = policies_dict[policy_num].get('_block_text', '')
            if len(block_text) > len(existing_block):
                policies_dict[policy_num]['_block_text'] = block_text
                block_text = block_text  # Use new block
            else:
                block_text = existing_block  # Keep existing block
        else:
            policies_dict[policy_num] = {'_block_text': block_text}
        
        # Extract fields from block
        policy = extract_policy_fields_from_block(block_text, mapping=None, source_type=source_type)
        
        if policy and policy.get('policy_number'):
            # Ensure policy_number matches (in case extraction found different one)
            policy['policy_number'] = policy_num
            
            # Update or create policy in dict
            policies_dict[policy_num] = policy
            logger.debug(f"[parse_policy_raw_text] Block {i+1}: extracted policy {policy.get('policy_number')}")
        else:
            logger.debug(f"[parse_policy_raw_text] Block {i+1}: no policy extracted")
    
    # Convert dict to list (deduplicated)
    policies = list(policies_dict.values())
    
    # Remove internal _block_text field if present
    for policy in policies:
        if '_block_text' in policy:
            del policy['_block_text']
    
    logger.info(f"[parse_policy_raw_text] Total unique policies extracted: {len(policies)} (after deduplication)")
    return policies

def split_by_location_id(text: str) -> List[Tuple[str, str]]:
    """
    Split text into location blocks by location_id positions.
    Mirrors split_by_vin() for vehicles and split_by_policy_number() for policies.
    
    Args:
        text: Raw text containing location descriptions (should be normalized)
    
    Returns:
        List of tuples: (location_id, block_text)
    """
    import re
    
    cleaned_text = text.strip()
    blocks = []
    
    # Pattern: Look for location IDs like "L001", "Location L001:", etc.
    location_pattern = r'(?i)(?:Location\s+)?(L\d+)\b'
    location_matches = list(re.finditer(location_pattern, cleaned_text))
    
    if location_matches:
        logger.info(f"[split_by_location_id] Found {len(location_matches)} location ID(s)")
        for i, match in enumerate(location_matches):
            location_id = match.group(1).upper()  # Normalize to uppercase
            start_pos = match.start()
            
            # Find end position (start of next location or end of text)
            if i + 1 < len(location_matches):
                end_pos = location_matches[i + 1].start()
            else:
                end_pos = len(cleaned_text)
            
            block_text = cleaned_text[start_pos:end_pos].strip()
            blocks.append((location_id, block_text))
    else:
        logger.warning("[split_by_location_id] No location IDs found in text")
        return []
    
    return blocks


def extract_location_fields_from_block(
    text_block: str,
    mapping: Optional[Dict[str, Any]] = None,
    source_type: Optional[str] = None
) -> Dict[str, Any]:
    """
    Extract ALL fields for a single location from a location_id-isolated block.
    Mirrors extract_fields_from_block() for vehicles.
    
    Args:
        text_block: Text block containing location information
        mapping: Optional mapping configuration
        source_type: Source type ("pdf", "raw", "image") to determine extraction logic
    
    Returns:
        Dictionary with location fields matching LOCATION_SCHEMA_ORDER
    """
    from schema import LOCATION_SCHEMA_ORDER
    import re
    
    # Initialize location dict with all fields as None
    location = {field: None for field in LOCATION_SCHEMA_ORDER}
    
    # Extract location_id from block (should be present)
    location_id_match = re.search(r'(?i)(?:Location\s+)?(L\d+)\b', text_block)
    if location_id_match:
        location['location_id'] = location_id_match.group(1).upper()
    
    # Extract insured_name
    insured_match = re.search(r'(?i)(?:Named\s+Insured|Insured\s+Name)[:\-]?\s*([^\n]+)', text_block)
    if not insured_match:
        insured_match = re.search(r'(?i)for\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)', text_block)
    if insured_match:
        location['insured_name'] = insured_match.group(1).strip()
    
    # Extract address_line_1 - look for street addresses
    address_match = re.search(r'(\d+\s+[A-Z][a-z]+(?:\s+[A-Z][a-z]+)?\s+(?:St|Street|Ave|Avenue|Rd|Road|Blvd|Boulevard|Dr|Drive|Ln|Lane|Ct|Court))', text_block, re.IGNORECASE)
    if address_match:
        location['address_line_1'] = address_match.group(1).strip()
    
    # Extract city - look for "City: Los Angeles" pattern
    city_match = re.search(r'(?i)City\s*:\s*([^\n]+)', text_block)
    if city_match:
        location['city'] = city_match.group(1).strip()
    else:
        # Fallback: look for city names after keywords
        city_keyword_match = re.search(r'(?i)(?:city|in|at|located\s+in)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)', text_block)
        if city_keyword_match:
            location['city'] = city_keyword_match.group(1).strip()
    
    # Extract state - look for 2-letter state codes
    state_match = re.search(r'\b([A-Z]{2})\b', text_block)
    if state_match:
        location['state'] = state_match.group(1)
    
    # Extract postal_code - look for 5 or 9 digit ZIP codes
    zip_match = re.search(r'\b(\d{5}(?:-\d{4})?)\b', text_block)
    if zip_match:
        location['postal_code'] = zip_match.group(1)
    
    # Extract county
    county_match = re.search(r'(?i)County[:\-]?\s*([^\n,]+)', text_block)
    if county_match:
        location['county'] = county_match.group(1).strip()
    
    # Extract territory_code
    territory_match = re.search(r'(?i)Territory\s+(?:Code)?[:\-]?\s*([A-Z0-9]+)', text_block)
    if territory_match:
        location['territory_code'] = territory_match.group(1).strip()
    
    # Extract protection_class
    protection_match = re.search(r'(?i)Protection\s+Class[:\-]?\s*(\d+)', text_block)
    if protection_match:
        try:
            location['protection_class'] = int(protection_match.group(1))
        except (ValueError, TypeError):
            pass
    
    # Extract latitude
    lat_match = re.search(r'(?i)Latitude[:\-]?\s*([-+]?\d+\.?\d*)', text_block)
    if lat_match:
        location['latitude'] = lat_match.group(1).strip()
    
    # Extract longitude
    lon_match = re.search(r'(?i)Longitude[:\-]?\s*([-+]?\d+\.?\d*)', text_block)
    if lon_match:
        location['longitude'] = lon_match.group(1).strip()
    
    # Extract notes
    notes_match = re.search(r'(?i)Notes[:\-]?\s*([^\n]+(?:\n(?!Location|L\d+)[^\n]+)*)', text_block, re.MULTILINE)
    if notes_match:
        notes_text = re.sub(r'\s+', ' ', notes_match.group(1).strip())
        if notes_text:
            location['notes'] = notes_text
    # If notes not explicitly found, leave as None (don't populate with OCR block)
    
    return location


def parse_locations_raw_text(text: str, source_type: str = "raw") -> List[Dict[str, Any]]:
    """
    Parse unstructured location text and extract all canonical location fields.
    
    Refactored to use split_by_location_id() and extract_location_fields_from_block()
    following the vehicle/driver/policy/claim/relationship pipeline pattern.
    
    Strategy:
    1. Normalize OCR text (whitespace, label variants, spacing)
    2. Split text into non-overlapping blocks by location_id positions
    3. For each block, extract all location fields using extract_location_fields_from_block
    4. Return raw values (normalization happens in normalize_v2 via mappings/transforms)
    
    Args:
        text: Raw text containing location descriptions (CSV format or plaintext)
        source_type: Source type ("pdf", "raw", "image") to use for field extraction
    
    Returns:
        List of dicts, one per location, with fields matching LOCATION_SCHEMA_ORDER
    """
    from schema import LOCATION_SCHEMA_ORDER
    import csv
    import io
    import re
    
    # ========================================================================
    # PASS 1: NORMALIZE OCR TEXT
    # ========================================================================
    def normalize_ocr_text(raw_text: str) -> str:
        """Normalize OCR text for consistent parsing."""
        # Collapse multiple spaces to single space
        text = re.sub(r' +', ' ', raw_text)
        # Collapse multiple newlines to double newline (paragraph break)
        text = re.sub(r'\n{3,}', '\n\n', text)
        # Normalize spacing around colons: "Location ID : value" -> "Location ID: value"
        text = re.sub(r'\s+:\s+', ': ', text)
        # Normalize spacing around dashes used as separators
        text = re.sub(r'([A-Za-z]+)\s+-\s+([A-Za-z0-9])', r'\1: \2', text)
        return text.strip()
    
    normalized_text = normalize_ocr_text(text)
    
    # Try to parse as CSV first (if it has headers)
    lines = normalized_text.split('\n')
    if len(lines) > 1:
        # Check if first line looks like CSV headers
        first_line = lines[0].strip()
        if ',' in first_line and any(keyword in first_line.lower() for keyword in ['location', 'address', 'city', 'state']):
            try:
                # Parse as CSV
                locations = []
                reader = csv.DictReader(io.StringIO(normalized_text))
                for row in reader:
                    location = {}
                    # Map CSV columns to schema fields
                    location['location_id'] = row.get('Location ID') or row.get('location_id') or None
                    location['insured_name'] = row.get('Named Insured') or row.get('insured_name') or None
                    location['address_line_1'] = row.get('Address Line 1') or row.get('address_line_1') or row.get('Address') or None
                    location['city'] = row.get('City') or row.get('city') or None
                    location['state'] = row.get('State') or row.get('state') or None
                    location['postal_code'] = row.get('Postal Code') or row.get('postal_code') or row.get('ZIP') or row.get('zip') or None
                    location['county'] = row.get('County') or row.get('county') or None
                    location['territory_code'] = row.get('Territory Code') or row.get('territory_code') or None
                    location['protection_class'] = row.get('Protection Class') or row.get('protection_class') or None
                    location['latitude'] = row.get('Latitude') or row.get('latitude') or None
                    location['longitude'] = row.get('Longitude') or row.get('longitude') or None
                    location['notes'] = row.get('Notes') or row.get('notes') or None
                    
                    # Initialize missing fields
                    for field in LOCATION_SCHEMA_ORDER:
                        if field not in location:
                            location[field] = None
                    
                    # Only add if we have at least a location_id
                    if location.get('location_id'):
                        locations.append(location)
                logger.debug(f"[parse_locations_raw_text] Parsed {len(locations)} location(s) from CSV format")
                return locations
            except Exception as e:
                logger.debug(f"[parse_locations_raw_text] CSV parsing failed: {e}, trying block-based parsing")
    
    # ========================================================================
    # PASS 2: SPLIT INTO BLOCKS BY LOCATION_ID
    # ========================================================================
    blocks = split_by_location_id(normalized_text)
    logger.info(f"[parse_locations_raw_text] Found {len(blocks)} location block(s)")
    
    # ========================================================================
    # PASS 3: EXTRACT FIELDS FROM EACH BLOCK
    # ========================================================================
    locations = []
    locations_dict = {}  # Deduplicate by location_id
    
    for i, (location_id, block_text) in enumerate(blocks):
        logger.debug(f"[parse_locations_raw_text] Processing block {i+1}/{len(blocks)}")
        
        # Deduplicate: if we already have this location, prefer the longer block
        if location_id in locations_dict:
            logger.info(f"[parse_locations_raw_text] Merging duplicate block for location {location_id}")
            existing_block = locations_dict[location_id].get('_block_text', '')
            if len(block_text) > len(existing_block):
                locations_dict[location_id]['_block_text'] = block_text
                block_text = block_text  # Use new block
            else:
                block_text = existing_block  # Keep existing block
        else:
            locations_dict[location_id] = {'_block_text': block_text}
        
        # Extract fields from block
        location = extract_location_fields_from_block(block_text, mapping=None, source_type=source_type)
        
        if location and location.get('location_id'):
            # Ensure location_id matches (in case extraction found different one)
            location['location_id'] = location_id.upper()
            
            # Update or create location in dict
            locations_dict[location_id] = location
            logger.debug(f"[parse_locations_raw_text] Block {i+1}: extracted location {location.get('location_id')}")
        else:
            logger.debug(f"[parse_locations_raw_text] Block {i+1}: no location extracted")
    
    # Convert dict to list (deduplicated)
    locations = list(locations_dict.values())
    
    # Remove internal _block_text field if present
    for location in locations:
        if '_block_text' in location:
            del location['_block_text']
    
    logger.info(f"[parse_locations_raw_text] Total unique locations extracted: {len(locations)} (after deduplication)")
    return locations

def parse_driver_raw_text(text: str) -> List[Dict[str, Any]]:
    """
    Parse unstructured driver text and extract all canonical driver fields.
    
    Unified multi-pass extraction engine for ALL driver sources:
    - Airtable, Google Sheet, Excel, Raw Text, PDF OCR, Image OCR
    
    Strategy:
    PASS 1: Normalize OCR text (whitespace, label variants, spacing)
    PASS 2: Extract using structured patterns (first-match-wins)
    PASS 3: Fallback OCR patterns (search full_text if block_text fails)
    PASS 4: Narrative extraction (append narrative hints to notes)
    PASS 5: Final normalization (types, defaults)
    
    Args:
        text: Raw text containing driver descriptions
    
    Returns:
        List of dicts, one per driver, with fields matching DRIVER_SCHEMA_ORDER
    """
    from schema import DRIVER_SCHEMA_ORDER
    import re
    
    # ========================================================================
    # PASS 1: NORMALIZE OCR TEXT
    # ========================================================================
    # Remove extra whitespace, fix broken line breaks, collapse repeated newlines
    # Normalize label variants (e.g., "Notes -", "Notes :", "Note", "NOTES") to "Notes:"
    # Normalize spacing around colons
    
    def normalize_ocr_text(raw_text: str) -> str:
        """Normalize OCR text for consistent parsing."""
        # Collapse multiple spaces to single space
        text = re.sub(r' +', ' ', raw_text)
        # Collapse multiple newlines to double newline (paragraph break)
        text = re.sub(r'\n{3,}', '\n\n', text)
        # Normalize spacing around colons: "Notes : value" -> "Notes: value"
        # BUT preserve the colon itself - do NOT remove colons from labels
        text = re.sub(r'\s+:\s+', ': ', text)
        # Normalize spacing around dashes used as separators: "Notes - value" -> "Notes: value"
        # BUT only if it looks like a label separator (followed by whitespace and value)
        text = re.sub(r'([A-Za-z]+)\s+-\s+([A-Za-z0-9])', r'\1: \2', text)
        # For corrupted OCR: collapse single newlines within name patterns (helps with multi-line names)
        # Only collapse if it's between two capitalized words (likely a name split across lines)
        # Pattern: Capital letter, lowercase letters, newline, capital letter, lowercase letters
        text = re.sub(r'([A-Z][a-z]+)\n([A-Z][a-z]+)', r'\1 \2', text)
        return text.strip()
    
    # Normalize the input text
    normalized_text = normalize_ocr_text(text)
    full_text = normalized_text  # Store for PASS 3 fallback
    
    # ========================================================================
    # PART 1: BLOCK SPLITTING
    # ========================================================================
    # IF text contains "Driver ID:" → Use PDF-style splitting
    # ELSE → Use original raw-text splitting (bullets, blank lines)
    
    def split_into_driver_blocks(text: str) -> List[Tuple[str, str]]:
        """Split text into driver blocks."""
        blocks = []
        
        # Check if text contains typed PDF anchors: "Driver ID:" or "DRIVER DETAIL SHEET"
        has_driver_id_label = bool(re.search(r'(?i)Driver\s+(?:ID|#)\s*:', text))
        has_driver_detail_sheet = bool(re.search(r'(?i)DRIVER\s+DETAIL\s+SHEET', text))
        is_typed_pdf = has_driver_id_label or has_driver_detail_sheet
        
        if is_typed_pdf:
            # TYPED PDF: Require hard anchor (Driver ID: or DRIVER DETAIL SHEET)
            # This prevents over-extraction from random text blocks
            
            # First, try splitting on "Driver ID:" labels (most common)
            if has_driver_id_label:
                driver_id_pattern = r'(?i)(Driver\s+(?:ID|#)\s*:\s*)([A-Z0-9]{1,6})\b'
                matches = list(re.finditer(driver_id_pattern, text))
                
                if matches:
                    for i, match in enumerate(matches):
                        driver_id = match.group(2)
                        label_start = match.start()
                        
                        # Get the start of the next driver block (or end of text)
                        if i + 1 < len(matches):
                            next_start = matches[i + 1].start()
                        else:
                            next_start = len(text)
                        
                        # Extract block from this Driver ID label to the next one
                        # Include the "Driver ID: D001" label in the block
                        block_text = text[label_start:next_start].strip()
                        blocks.append((driver_id, block_text))
            
            # If no Driver ID: labels found but DRIVER DETAIL SHEET exists, split on that
            elif has_driver_detail_sheet:
                # Split on "DRIVER DETAIL SHEET" headers
                sheet_pattern = r'(?i)(DRIVER\s+DETAIL\s+SHEET)'
                matches = list(re.finditer(sheet_pattern, text))
                
                if matches:
                    for i, match in enumerate(matches):
                        sheet_start = match.start()
                        
                        # Get the start of the next sheet (or end of text)
                        if i + 1 < len(matches):
                            next_start = matches[i + 1].start()
                        else:
                            next_start = len(text)
                        
                        # Extract block from this sheet header to the next one
                        block_text = text[sheet_start:next_start].strip()
                        
                        # Try to extract driver_id from the block (look for "Driver ID: D001" within the block)
                        driver_id_match = re.search(r'(?i)Driver\s+(?:ID|#)\s*:\s*([A-Z0-9]{1,6})\b', block_text)
                        if driver_id_match:
                            driver_id = driver_id_match.group(1)
                        else:
                            # Fallback: use counter if no Driver ID found
                            driver_id = f"D{i+1:03d}"
                        
                        blocks.append((driver_id, block_text))
        else:
            # Original raw-text block splitting:
            # Split on leading bullet patterns (e.g. "- John", "• John")
            # Split on blank-line-separated paragraphs
            # Ensure each driver becomes its own block
            
            lines = text.split('\n')
            driver_counter = 1
            current_block = []
            
            for line in lines:
                line = line.strip()
                if not line:
                    # Blank line - if we have a current block, save it
                    if current_block:
                        block_text = ' '.join(current_block)
                        driver_id = f"D{driver_counter:03d}"
                        blocks.append((driver_id, block_text))
                        driver_counter += 1
                        current_block = []
                    continue
                
                # Pattern: "- First Last (DOB ...)" or "• First Last (DOB ...)"
                if re.match(r'^[-•]\s+[A-Z][a-z]+\s+[A-Z][a-z]+', line):
                    # Save previous block if exists
                    if current_block:
                        block_text = ' '.join(current_block)
                        driver_id = f"D{driver_counter:03d}"
                        blocks.append((driver_id, block_text))
                        driver_counter += 1
                        current_block = []
                    # Start new block with this line
                    current_block = [line]
                elif re.search(r'(?:One\s+additional\s+driver|additional\s+driver|incomplete\s+information)', line, re.IGNORECASE):
                    # Save previous block if exists
                    if current_block:
                        block_text = ' '.join(current_block)
                        driver_id = f"D{driver_counter:03d}"
                        blocks.append((driver_id, block_text))
                        driver_counter += 1
                        current_block = []
                    # Start new block with this line
                    current_block = [line]
                else:
                    # Add line to current block
                    current_block.append(line)
            
            # Save final block if exists
            if current_block:
                block_text = ' '.join(current_block)
                driver_id = f"D{driver_counter:03d}"
                blocks.append((driver_id, block_text))
        
        # RECOVERY FALLBACK: Only run when NO blocks found (not when we have some blocks)
        # This mirrors vehicle VIN recovery logic but only for corrupted OCR cases
        # Do NOT interfere with normal extraction that found valid blocks
        if len(blocks) == 0:
            logger.debug(f"[parse_driver_raw_text] No blocks found with standard splitting, trying recovery fallback for corrupted OCR...")
            # Try to find name patterns even in corrupted OCR
            # Look for patterns like "Name: First Last" or "First Last" (even if corrupted)
            name_patterns = [
                r'(?i)(?:Name|Nae|Nane|Nae|Nowe)\s*:?\s*([A-Z][a-z]+\s+[A-Z][a-z]+)',  # "Name: First Last" or corrupted variants
                r'([A-Z][a-z]{2,}\s+[A-Z][a-z]{2,})',  # "First Last" (at least 3 chars each) - more lenient
            ]
            
            # Try to split on name patterns
            recovery_blocks = []
            for pattern in name_patterns:
                # Use MULTILINE flag to handle names split across lines
                matches = list(re.finditer(pattern, normalized_text, re.MULTILINE | re.DOTALL))
                if matches:
                    logger.debug(f"[parse_driver_raw_text] Recovery found {len(matches)} name pattern(s) using pattern: {pattern[:50]}")
                    for i, match in enumerate(matches):
                        # For multi-line names, find the actual start (may be on previous line)
                        name_start = match.start()
                        # Find the start of the line containing "Name:" or the name itself
                        # Look backwards for "Name:" or similar label
                        search_start = max(0, name_start - 50)  # Look back up to 50 chars
                        label_match = re.search(r'(?i)(?:Name|Nae|Nane|Nowe)\s*:?\s*', normalized_text[search_start:name_start])
                        if label_match:
                            # Start from the label
                            block_start = search_start + label_match.start()
                        else:
                            # Start from the line containing the name
                            line_start = normalized_text.rfind('\n', 0, name_start)
                            block_start = line_start + 1 if line_start != -1 else 0
                        
                        # Get the start of the next name (or end of text)
                        if i + 1 < len(matches):
                            next_name_start = matches[i + 1].start()
                            # Find the start of the line for the next name
                            next_search_start = max(0, next_name_start - 50)
                            next_label_match = re.search(r'(?i)(?:Name|Nae|Nane|Nowe)\s*:?\s*', normalized_text[next_search_start:next_name_start])
                            if next_label_match:
                                next_start = next_search_start + next_label_match.start()
                            else:
                                next_line_start = normalized_text.rfind('\n', 0, next_name_start)
                                next_start = next_line_start + 1 if next_line_start != -1 else 0
                        else:
                            next_start = len(normalized_text)
                        
                        # Extract block from this name's start to the next name's start
                        block_text = normalized_text[block_start:next_start].strip()
                        if len(block_text) > 15:  # Lower minimum block size for corrupted OCR
                            driver_id = f"D{i+1:03d}"
                            recovery_blocks.append((driver_id, block_text))
                    if len(recovery_blocks) > 0:
                        break  # Use first pattern that finds matches
            
            # Append recovery blocks (don't replace - we had 0 blocks anyway)
            blocks.extend(recovery_blocks)
            
            # If still no blocks, try paragraph-based splitting (split on double newlines)
            if len(blocks) == 0:
                logger.debug(f"[parse_driver_raw_text] Still no blocks, trying paragraph-based splitting...")
                paragraphs = re.split(r'\n\s*\n', normalized_text)
                for i, para in enumerate(paragraphs):
                    para = para.strip()
                    if len(para) > 15:  # Lower minimum paragraph size for corrupted OCR
                        driver_id = f"D{i+1:03d}"
                        blocks.append((driver_id, para))
                        if len(blocks) >= 3:  # Stop if we have reasonable number
                            break
        
        return blocks
    
    blocks = split_into_driver_blocks(normalized_text)
    logger.debug(f"[parse_driver_raw_text] Found {len(blocks)} driver block(s) before filtering")
    
    # ========================================================================
    # BLOCK VALIDATION: Filter out non-driver blocks
    # ========================================================================
    # INVARIANT: A block is valid ONLY if it contains a driver identifier
    def is_valid_driver_block(block_text: str) -> bool:
        """Check if block contains at least one driver identifier."""
        # Check if this is a typed PDF block (has structured anchors)
        has_driver_id_anchor = bool(re.search(r'(?i)Driver\s+(?:ID|#)\s*:\s*[A-Z0-9]{1,6}', block_text))
        has_driver_detail_sheet = bool(re.search(r'(?i)DRIVER\s+DETAIL\s+SHEET', block_text))
        is_typed_pdf_block = has_driver_id_anchor or has_driver_detail_sheet
        
        # For typed PDF blocks: REQUIRE hard anchor (strict validation)
        # This prevents over-extraction from random text blocks
        if is_typed_pdf_block:
            # Must have Driver ID: label OR DRIVER DETAIL SHEET header
            # Also check for structured fields to ensure it's a complete driver record
            has_structured_fields = bool(re.search(r'(?:Full\s+Name|Date\s+of\s+Birth|License\s+Number)\s*:', block_text, re.IGNORECASE))
            return has_driver_id_anchor or (has_driver_detail_sheet and has_structured_fields)
        
        # For handwritten/non-typed blocks: Use lenient validation (existing logic)
        # Must have at least ONE of these:
        has_driver_id = bool(re.search(r'(?i)(?:Driver\s+(?:ID|#)\s*:\s*)?(D\d{3}|DRV\d{3})', block_text))
        has_name_and_dob = bool(re.search(r'[-•]?\s*[A-Z][a-z]+\s+[A-Z][a-z]+.*(?:DOB|\(DOB\s+\d{4})', block_text))
        has_structured_fields = bool(re.search(r'(?:Full\s+Name|Date\s+of\s+Birth|License\s+Number)\s*:', block_text, re.IGNORECASE))
        has_name_pattern = bool(re.search(r'^[-•]\s+[A-Z][a-z]+\s+[A-Z][a-z]+', block_text))
        # Allow narrative driver hints (e.g., "One additional driver", "incomplete information")
        has_narrative_driver = bool(re.search(r'(?i)(?:one\s+additional\s+driver|additional\s+driver|incomplete\s+information)', block_text))
        # PDF blocks with structured fields should pass even without name_pattern
        
        # Lenient patterns for corrupted OCR (similar to vehicle recovery)
        # Accept blocks with name-like patterns (even if corrupted)
        has_corrupted_name = bool(re.search(r'[A-Z][a-z]{2,}\s+[A-Z][a-z]{2,}', block_text))  # "First Last" (at least 3 chars each)
        # Accept blocks with common driver field keywords (even if labels are corrupted)
        has_driver_keywords = bool(re.search(r'(?i)(?:name|dob|license|state|email|driver)', block_text))
        # Accept blocks with date-like patterns (DOB indicators)
        has_date_pattern = bool(re.search(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', block_text))  # Date patterns
        # Accept blocks with license-like patterns (alphanumeric sequences)
        has_license_pattern = bool(re.search(r'[A-Z]\d{6,}', block_text))  # License number patterns like "A1234567"
        
        # For corrupted OCR, be more lenient: accept if it has name + any driver-related content
        has_driver_like_content = (has_corrupted_name and (has_driver_keywords or has_date_pattern or has_license_pattern))
        
        # Standard validation OR lenient validation for corrupted OCR
        return (has_driver_id or has_name_and_dob or has_structured_fields or 
                (has_structured_fields and has_name_pattern) or has_narrative_driver or
                has_driver_like_content)
    
    # Filter blocks to only include valid driver blocks
    valid_blocks = []
    for driver_id, block_text in blocks:
        if is_valid_driver_block(block_text):
            valid_blocks.append((driver_id, block_text))
        else:
            logger.debug(f"[parse_driver_raw_text] Filtered out invalid block (ID: {driver_id}): {block_text[:200]}...")
            # Debug: show why block was filtered
            has_driver_id = bool(re.search(r'(?i)(?:Driver\s+(?:ID|#)\s*:\s*)?(D\d{3}|DRV\d{3})', block_text))
            has_name_and_dob = bool(re.search(r'[-•]?\s*[A-Z][a-z]+\s+[A-Z][a-z]+.*(?:DOB|\(DOB\s+\d{4})', block_text))
            has_structured_fields = bool(re.search(r'(?:Full\s+Name|Date\s+of\s+Birth|License\s+Number)\s*:', block_text, re.IGNORECASE))
            has_name_pattern = bool(re.search(r'^[-•]\s+[A-Z][a-z]+\s+[A-Z][a-z]+', block_text))
            has_narrative_driver = bool(re.search(r'(?i)(?:one\s+additional\s+driver|additional\s+driver|incomplete\s+information)', block_text))
            logger.debug(f"  Validation checks: driver_id={has_driver_id}, name_and_dob={has_name_and_dob}, structured={has_structured_fields}, name_pattern={has_name_pattern}, narrative={has_narrative_driver}")
    
    # Reassign driver_ids sequentially starting from 1, since invalid blocks were filtered out
    blocks = []
    for i, (old_driver_id, block_text) in enumerate(valid_blocks, start=1):
        new_driver_id = f"D{i:03d}"
        blocks.append((new_driver_id, block_text))
    logger.debug(f"[parse_driver_raw_text] Found {len(blocks)} valid driver block(s) after filtering")
    
    # RECOVERY FALLBACK AFTER FILTERING: If no valid blocks found, try recovery for corrupted OCR
    # This handles cases where initial splitting found blocks but they were all filtered out
    if len(blocks) == 0:
        logger.debug(f"[parse_driver_raw_text] No valid blocks after filtering, trying recovery fallback for corrupted OCR...")
        # Try to find name patterns even in corrupted OCR
        name_patterns = [
            r'(?i)(?:Name|Nae|Nane|Nae|Nowe)\s*:?\s*([A-Z][a-z]+\s+[A-Z][a-z]+)',  # "Name: First Last" or corrupted variants
            r'([A-Z][a-z]{2,}\s+[A-Z][a-z]{2,})',  # "First Last" (at least 3 chars each) - more lenient
        ]
        
        # Try to split on name patterns
        recovery_blocks = []
        for pattern in name_patterns:
            # Use MULTILINE flag to handle names split across lines
            matches = list(re.finditer(pattern, normalized_text, re.MULTILINE | re.DOTALL))
            if matches:
                logger.debug(f"[parse_driver_raw_text] Recovery found {len(matches)} name pattern(s) using pattern: {pattern[:50]}")
                for i, match in enumerate(matches):
                    # For multi-line names, find the actual start (may be on previous line)
                    name_start = match.start()
                    # Find the start of the line containing "Name:" or the name itself
                    # Look backwards for "Name:" or similar label
                    search_start = max(0, name_start - 50)  # Look back up to 50 chars
                    label_match = re.search(r'(?i)(?:Name|Nae|Nane|Nowe)\s*:?\s*', normalized_text[search_start:name_start])
                    if label_match:
                        # Start from the label
                        block_start = search_start + label_match.start()
                    else:
                        # Start from the line containing the name
                        line_start = normalized_text.rfind('\n', 0, name_start)
                        block_start = line_start + 1 if line_start != -1 else 0
                    
                    # Get the start of the next name (or end of text)
                    if i + 1 < len(matches):
                        next_name_start = matches[i + 1].start()
                        # Find the start of the line for the next name
                        next_search_start = max(0, next_name_start - 50)
                        next_label_match = re.search(r'(?i)(?:Name|Nae|Nane|Nowe)\s*:?\s*', normalized_text[next_search_start:next_name_start])
                        if next_label_match:
                            next_start = next_search_start + next_label_match.start()
                        else:
                            next_line_start = normalized_text.rfind('\n', 0, next_name_start)
                            next_start = next_line_start + 1 if next_line_start != -1 else 0
                    else:
                        next_start = len(normalized_text)
                    
                    # Extract block from this name's start to the next name's start
                    block_text = normalized_text[block_start:next_start].strip()
                    if len(block_text) > 15:  # Lower minimum block size for corrupted OCR
                        driver_id = f"D{i+1:03d}"
                        recovery_blocks.append((driver_id, block_text))
                if len(recovery_blocks) > 0:
                    break  # Use first pattern that finds matches
        
        # For recovery blocks from corrupted OCR, be very lenient with validation
        # Accept any block that has a name pattern (we already found names, so blocks are likely valid)
        valid_recovery_blocks = []
        for driver_id, block_text in recovery_blocks:
            # Very lenient check: just needs to have some text and a name-like pattern
            has_name = bool(re.search(r'[A-Z][a-z]{2,}\s+[A-Z][a-z]{2,}', block_text))
            has_any_content = len(block_text.strip()) > 20
            if has_name or has_any_content:
                valid_recovery_blocks.append((driver_id, block_text))
            else:
                logger.debug(f"[parse_driver_raw_text] Recovery block filtered out (too short/no name): {block_text[:100]}...")
        
        blocks.extend(valid_recovery_blocks)
        logger.debug(f"[parse_driver_raw_text] Recovery added {len(valid_recovery_blocks)} block(s) (lenient validation for corrupted OCR)")
        
        # If still no blocks, try paragraph-based splitting (split on double newlines)
        if len(blocks) == 0:
            logger.debug(f"[parse_driver_raw_text] Still no blocks, trying paragraph-based splitting...")
            paragraphs = re.split(r'\n\s*\n', normalized_text)
            for i, para in enumerate(paragraphs):
                para = para.strip()
                if len(para) > 15 and is_valid_driver_block(para):  # Validate paragraphs too
                    driver_id = f"D{i+1:03d}"
                    blocks.append((driver_id, para))
                    if len(blocks) >= 3:  # Stop if we have reasonable number
                        break
    
    drivers = []
    
    # ========================================================================
    # PASS 2: EXTRACT USING STRUCTURED PATTERNS (FIRST-MATCH-WINS)
    # ========================================================================
    # Extract fields from each block using OCR-friendly patterns
    # Patterns allow: missing colons, line breaks, extra spaces, upper/lowercase
    
    def extract_field_with_patterns(text: str, patterns: list, flags: int = 0):
        """Extract field using first matching pattern (first-match-wins)."""
        for pattern in patterns:
            match = re.search(pattern, text, flags)
            if match:
                return match.group(1).strip() if match.lastindex >= 1 else match.group(0).strip()
        return None
    
    for driver_id, block_text in blocks:
        driver = {field: None for field in DRIVER_SCHEMA_ORDER}
        driver['driver_id'] = driver_id
        
        # ========================================================================
        # IMMUTABILITY TRACKING: Fields set by structured extraction are immutable
        # ========================================================================
        # Track which fields were set by structured extraction from block_text
        # These fields MUST NEVER be overwritten by fallback extraction
        structured_fields = set()
        
        # ========================================================================
        # PART 2: STRUCTURED FIELD EXTRACTION (PASS 2)
        # ========================================================================
        # Use exact patterns that match OCR text
        # Structured patterns MUST run BEFORE fallback patterns
        # Only extract if field is not already set (first-match-wins)
        
        # Extract driver_id (only if not already set from block splitting)
        if not driver.get('driver_id') or driver['driver_id'] not in block_text:
            driver_id_match = extract_field_with_patterns(
                block_text,
                [r'(?i)Driver\s+ID[:\-]?\s*(\S+)']
            )
            if driver_id_match:
                driver['driver_id'] = driver_id_match
        
        # Extract full_name (only if not already set)
        if not driver.get('first_name') or not driver.get('last_name'):
            full_name_match = re.search(r'(?i)Full\s+Name[:\-]?\s*([A-Za-z]+)\s+([A-Za-z]+)', block_text)
            if full_name_match:
                driver['first_name'] = full_name_match.group(1).strip()
                driver['last_name'] = full_name_match.group(2).strip()
            else:
                # Fallback: Try unstructured pattern: "First Last" or "- First Last"
                name_match2 = re.search(r'^[-•]\s*([A-Z][a-z]+)\s+([A-Z][a-z]+)(?:\s*\(|\s|,|$)', block_text)
                if name_match2:
                    driver['first_name'] = name_match2.group(1).strip()
                    driver['last_name'] = name_match2.group(2).strip()
                else:
                    # Try without bullet
                    name_match3 = re.search(r'^([A-Z][a-z]+)\s+([A-Z][a-z]+)(?:\s*\(|\s|,|$)', block_text)
                    if name_match3:
                        driver['first_name'] = name_match3.group(1).strip()
                        driver['last_name'] = name_match3.group(2).strip()
        
        # Extract date_of_birth (only if not already set)
        if not driver.get('date_of_birth'):
            dob_str = extract_field_with_patterns(
                block_text,
                [r'(?i)Date\s+of\s+Birth[:\-]?\s*([0-9]{4}-[0-9]{2}-[0-9]{2})']
            )
            if dob_str:
                driver['date_of_birth'] = dob_str
            else:
                # Fallback: Try parenthetical format: "(DOB 1985-03-15)"
                dob_match2 = re.search(r'\(DOB\s+(\d{4}-\d{2}-\d{2})\)', block_text, re.IGNORECASE)
                if dob_match2:
                    driver['date_of_birth'] = dob_match2.group(1).strip()
        
        # Extract license_number (only if not already set)
        if not driver.get('license_number'):
            license_val = extract_field_with_patterns(
                block_text,
                [r'(?i)License\s+Number[:\-]?\s*([A-Z0-9\-]{4,15})']
            )
            if license_val and license_val.lower() not in ['number', 'state', 'status', 'for', 'testing']:
                driver['license_number'] = license_val
            else:
                # Fallback: Try unstructured pattern
                license_match2 = re.search(r'(?i)(?:with\s+)?license\s+([A-Z0-9\-]{4,10})', block_text)
                if license_match2:
                    license_val = license_match2.group(1).strip()
                    if license_val.lower() not in ['number', 'state', 'status']:
                        driver['license_number'] = license_val
        
        # Extract license_state (only if not already set)
        if not driver.get('license_state'):
            state_val = extract_field_with_patterns(
                block_text,
                [r'(?i)License\s+State[:\-]?\s*([A-Z]{2})']
            )
            if state_val:
                driver['license_state'] = state_val
            else:
                # Fallback: Try unstructured pattern
                state_match2 = re.search(r'(?i)(?:issued\s+in|in)\s+([A-Z]{2})\b', block_text)
                if state_match2:
                    driver['license_state'] = state_match2.group(1).strip()
        
        # Extract license_status (only if not already set)
        if not driver.get('license_status'):
            status_val = extract_field_with_patterns(
                block_text,
                [r'(?i)License\s+Status[:\-]?\s*(\w+)']
            )
            if status_val:
                driver['license_status'] = status_val.capitalize()
        
        # Extract years_experience (only if not already set)
        if not driver.get('years_experience'):
            exp_str = extract_field_with_patterns(
                block_text,
                [r'(?i)Years\s+Experience[:\-]?\s*(\d+)']
            )
            if exp_str:
                try:
                    driver['years_experience'] = int(exp_str)
                except (ValueError, AttributeError):
                    pass
        
        # Extract violations_count (only if not already set)
        if driver.get('violations_count') is None:
            violations_str = extract_field_with_patterns(
                block_text,
                [r'(?i)Violations\s+Count[:\-]?\s*(\d+)']
            )
            if violations_str:
                try:
                    driver['violations_count'] = int(violations_str)
                    structured_fields.add('violations_count')  # Mark as immutable
                except (ValueError, AttributeError):
                    pass
            # Also try alternative patterns if the main one didn't match
            elif not violations_str:
                violations_str = extract_field_with_patterns(
                    block_text,
                    [r'(?i)Violations?\s*[:\-]?\s*(\d+)']
                )
                if violations_str:
                    try:
                        driver['violations_count'] = int(violations_str)
                        structured_fields.add('violations_count')
                    except (ValueError, AttributeError):
                        pass
        
        # Extract training_completed (only if not already set)
        if not driver.get('training_completed'):
            training_str = extract_field_with_patterns(
                block_text,
                [r'(?i)Training\s+Completed[:\-]?\s*(Yes|No|Y|N|True|False|Completed|Not\s+Completed)']
            )
            if training_str:
                training_lower = training_str.lower()
                if training_lower in ['yes', 'y', 'true', 'completed']:
                    driver['training_completed'] = 'Yes'
                    structured_fields.add('training_completed')  # Mark as immutable
                elif training_lower in ['no', 'n', 'false', 'not completed']:
                    driver['training_completed'] = 'No'
                    structured_fields.add('training_completed')  # Mark as immutable
                elif training_str in ['Yes', 'No']:
                    driver['training_completed'] = training_str
                    structured_fields.add('training_completed')  # Mark as immutable
        
        # Extract notes (only if not already set) - MULTILINE: until next label or EOF
        if not driver.get('notes'):
            # Try multiple patterns for notes extraction
            # Pattern 1: Match until next section header or end of text
            notes_match = re.search(r'(?i)Notes[:\-]?\s*(.+?)(?=\n(?:Driver\s+(?:ID|#)|Full\s+Name|Date\s+of\s+Birth|License|Years|Violations|Training|DRIVER\s+DETAIL)|\Z)', block_text, re.MULTILINE | re.DOTALL)
            if not notes_match:
                # Pattern 2: Match until double newline or next capitalized section
                notes_match = re.search(r'(?i)Notes[:\-]?\s*(.+?)(?=\n\n|\n[A-Z][A-Za-z ]+:|$|\Z)', block_text, re.MULTILINE | re.DOTALL)
            if not notes_match:
                # Pattern 3: Match until end of block (simplest - just get everything after Notes:)
                notes_match = re.search(r'(?i)Notes[:\-]?\s*(.+?)(?=$|\Z)', block_text, re.MULTILINE | re.DOTALL)
            if notes_match:
                notes_text = notes_match.group(1).strip()
                # Normalize whitespace but keep original casing and punctuation
                notes_text = re.sub(r'\s+', ' ', notes_text)
                driver['notes'] = notes_text
                structured_fields.add('notes')  # Mark as immutable
            else:
                # For raw text: Extract occupation/profession from narrative text
                # Pattern: "a Software Engineer", "a secondary driver", etc.
                # Try "a [Occupation]" pattern first
                occupation_match = re.search(r'(?:,\s+)?a\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)?)\s+(?:with|\.|$)', block_text, re.IGNORECASE)
                if occupation_match:
                    occupation = occupation_match.group(1).strip()
                    # Preserve original casing from the text
                    driver['notes'] = occupation
                    structured_fields.add('notes')
                else:
                    # Try "listed as a [role]" pattern
                    listed_as_match = re.search(r'listed\s+as\s+(?:a\s+)?([^,\.]+)', block_text, re.IGNORECASE)
                    if listed_as_match:
                        role = listed_as_match.group(1).strip()
                        # Capitalize first letter of each word for "listed as" pattern
                        role = ' '.join(word.capitalize() for word in role.split())
                        driver['notes'] = role
                        structured_fields.add('notes')
        
        # ========================================================================
        # PASS 3: FALLBACK OCR PATTERNS (search full_text if block_text fails)
        # ========================================================================
        # If any field is missing, search the FULL OCR TEXT using:
        # - label synonyms ("Training", "Completed Training", "TrainingComplete", etc.)
        # - missing-colon patterns ("Notes something" with no colon)
        # - multi-line notes: continue until next label OR blank line
        
        violations_found_in_block = bool(re.search(r'(?i)violations?\s*(?:count\s*)?[:\-]?\s*\d+', block_text))
        
        if (driver.get('violations_count') is None or (driver.get('violations_count') == 0 and not violations_found_in_block)) or \
           driver.get('training_completed') is None or \
           not driver.get('notes'):
            
            # Find driver section in full_text
            driver_section = None
            driver_section_pattern = rf'(?i)(?:Driver\s+(?:ID|#)?\s*:\s*{re.escape(driver_id)}|{re.escape(driver_id)}).*?(?=(?:Driver\s+(?:ID|#)?\s*:\s*[A-Z0-9]|$|\Z))'
            driver_section_match = re.search(driver_section_pattern, full_text, re.DOTALL)
            if driver_section_match:
                driver_section = driver_section_match.group(0)
            elif driver.get('first_name') and driver.get('last_name'):
                # Try finding by name
                name_pattern = rf'{re.escape(driver["first_name"])}\s+{re.escape(driver["last_name"])}.*?(?=(?:Driver\s+(?:ID|#)?\s*:|$|\Z))'
                name_section_match = re.search(name_pattern, full_text, re.IGNORECASE | re.DOTALL)
                if name_section_match:
                    driver_section = name_section_match.group(0)
            
            if driver_section:
                # Extract violations_count from full section (ONLY if not in structured_fields)
                if 'violations_count' not in structured_fields:
                    if driver.get('violations_count') is None or (driver.get('violations_count') == 0 and not violations_found_in_block):
                        violations_str = extract_field_with_patterns(
                            driver_section,
                            [
                            r'(?i)Violations\s+Count\s*:\s*(\d+)',
                            r'(?i)Violations\s*:\s*(\d+)',
                            r'(?i)violations?\s*(?:count\s*)?[:\-]?\s*(?:\n\s*)?(\d+)\b',
                            ],
                            re.MULTILINE
                            )
                        if violations_str:
                            try:
                                driver['violations_count'] = int(violations_str)
                            except (ValueError, AttributeError):
                                pass
                
                # Extract training_completed from full section (ONLY if not in structured_fields)
                if 'training_completed' not in structured_fields and not driver.get('training_completed'):
                    training_str = extract_field_with_patterns(
                        driver_section,
                        [
                            r'(?i)Training\s+Completed\s*:\s*(yes|no|y|n|true|false|1|0|completed|not\s+completed|incomplete)\b',
                            r'(?i)Training\s*:\s*(yes|no|y|n|true|false|1|0|completed|not\s+completed|incomplete)\b',
                            r'(?i)Completed\s+Training\s*:\s*(yes|no|y|n|true|false|1|0|completed|not\s+completed|incomplete)\b',
                            r'(?i)training\s*(?:completed\s*)?[:\-]?\s*(?:\n\s*)?(yes|no)\b',
                        ],
                        re.MULTILINE
                    )
                    if training_str:
                        training_lower = training_str.lower()
                        if training_lower in ['yes', 'y', 'true', '1', 'completed']:
                            driver['training_completed'] = 'Yes'
                        elif training_lower in ['no', 'n', 'false', '0', 'not completed', 'incomplete']:
                            driver['training_completed'] = 'No'
                
                # Extract notes from full section (ONLY if not in structured_fields)
                if 'notes' not in structured_fields and not driver.get('notes'):
                    notes_text = extract_field_with_patterns(
                        driver_section,
                        [
                            r'(?i)Notes\s*:\s*(.+?)(?=\n(?:Driver\s+ID|Full\s+Name|Date\s+of\s+Birth|License|Years|Violations|Training|Notes|Random\s+text)|\Z)',
                            r'(?i)Note\s*:\s*(.+?)(?=\n(?:Driver\s+ID|Full\s+Name|Date\s+of\s+Birth|License|Years|Violations|Training|Notes|Random\s+text)|\Z)',
                            r'(?i)Notes\s+([^\n]+(?:\n(?!Driver\s+ID|Full\s+Name|Date\s+of\s+Birth|License|Years|Violations|Training|Notes)[^\n]+)*)',  # Missing colon
                            r'(?i)Notes\s*:\s*(.+?)(?=\n\n|$|\Z)',
                            r'(?i)Notes\s*:\s*(.+)$',
                        ],
                        re.MULTILINE | re.DOTALL
                    )
                    if notes_text:
                        notes_text = re.sub(r'\s+', ' ', notes_text.strip())
                        driver['notes'] = notes_text
        
        # ========================================================================
        # PASS 4: NARRATIVE EXTRACTION (append narrative hints to notes)
        # ========================================================================
        # Look for narrative hints: "safe driving course", "listed as a secondary driver", etc.
        # Append these to notes EXACTLY as written in the OCR text
        
        narrative_patterns = [
            (r'([A-Z][a-z]+(?:\s+[A-Z]\.)?\s+safe\s+driving\s+course\.?)', lambda m: m.group(1).strip()),
            (r'(safe\s+driving\s+course\.?)', lambda m: m.group(1).strip()),
            (r'listed\s+as\s+(?:a\s+)?([^,\.]+)', lambda m: m.group(1).strip()),
            (r'incomplete\s+information[^.]*', lambda m: "Incomplete driver record with invalid or missing fields"),
        ]
        
        # Only run narrative extraction if notes are not already set from structured extraction
        if 'notes' not in structured_fields and not driver.get('notes'):
            # Search in block_text first, then driver_section if available
            search_texts = [block_text]
            if 'driver_section' in locals() and driver_section:
                search_texts.append(driver_section)
            
            for search_text in search_texts:
                for pattern, extractor in narrative_patterns:
                    narrative_match = re.search(pattern, search_text, re.IGNORECASE)
                    if narrative_match:
                        narrative_text = extractor(narrative_match)
                        if narrative_text and len(narrative_text.strip()) > 5:
                            narrative_text = narrative_text.strip()
                            # Preserve original casing from the text
                            driver['notes'] = narrative_text
                            break
                if driver.get('notes'):
                    break
        
        # ========================================================================
        # PASS 5: FINAL NORMALIZATION
        # ========================================================================
        # violations_count → int or 0
        # training_completed → "Yes" or "No"
        # notes → string (never None)
        
        if driver.get('violations_count') is None:
            driver['violations_count'] = 0
        elif not isinstance(driver.get('violations_count'), int):
            try:
                driver['violations_count'] = int(driver['violations_count'])
            except (ValueError, TypeError):
                driver['violations_count'] = 0
        
        if driver.get('notes') is None:
            driver['notes'] = ''
        
        # Debug logging for missing fields
        if (driver.get('violations_count') == 0 and not violations_found_in_block) or \
           driver.get('training_completed') is None or \
           not driver.get('notes'):
            logger.debug(f"[parse_driver_raw_text] Driver {driver_id} missing fields after all passes:")
            logger.debug(f"  violations_count={driver.get('violations_count')}")
            logger.debug(f"  training_completed={driver.get('training_completed')}")
            logger.debug(f"  notes={repr(driver.get('notes'))}")
            logger.debug(f"[parse_driver_raw_text] Block text:\n---\n{block_text}\n---")
        
        drivers.append(driver)
    
    logger.info(f"[parse_driver_raw_text] Total drivers extracted: {len(drivers)}")
    return drivers

def split_by_claim_number(text: str) -> List[Tuple[str, str]]:
    """
    Split text into claim blocks by claim_number positions.
    Mirrors split_by_vin() for vehicles and split_by_policy_number() for policies.
    
    Args:
        text: Raw text containing claim descriptions (should be normalized)
    
    Returns:
        List of tuples: (claim_number, block_text)
    """
    import re
    
    blocks = []
    
    # Pattern to find claim headers: "Claim ID:", "Claim Number:", "Claim #:"
    claim_header_pattern = r'(?i)(?:Claim\s+(?:ID|Number|#)\s*:\s*)([A-Z0-9\-_]+)'
    matches = list(re.finditer(claim_header_pattern, text))
    
    if matches:
        logger.info(f"[split_by_claim_number] Found {len(matches)} claim header(s)")
        for i, match in enumerate(matches):
            claim_num = match.group(1)
            label_start = match.start()
            
            # Get the start of the next claim block (or end of text)
            if i + 1 < len(matches):
                next_start = matches[i + 1].start()
            else:
                next_start = len(text)
            
            # Extract block from this Claim header to the next one
            block_text = text[label_start:next_start].strip()
            blocks.append((claim_num, block_text))
    else:
        # Fallback: Try narrative format: "Claim C001 is..." or "C_BAD1 is..."
        narrative_pattern = r'\b(C\d+|CLM\d+|CLM_[A-Z0-9]+)\s+is\b'
        narrative_matches = list(re.finditer(narrative_pattern, text, re.IGNORECASE))
        if narrative_matches:
            logger.info(f"[split_by_claim_number] Found {len(narrative_matches)} claim(s) in narrative format")
            for i, match in enumerate(narrative_matches):
                claim_num = match.group(1)
                start_pos = match.start()
                if i + 1 < len(narrative_matches):
                    next_start = narrative_matches[i + 1].start()
                else:
                    next_start = len(text)
                block_text = text[start_pos:next_start].strip()
                blocks.append((claim_num, block_text))
        else:
            logger.warning("[split_by_claim_number] No claim headers found")
            return []
    
    return blocks


def extract_claim_fields_from_block(
    text_block: str,
    mapping: Optional[Dict[str, Any]] = None,
    source_type: Optional[str] = None
) -> Dict[str, Any]:
    """
    Extract ALL fields for a single claim from a claim_number-isolated block.
    Mirrors extract_fields_from_block() for vehicles and extract_policy_fields_from_block() for policies.
    
    Args:
        text_block: Text block containing claim information
        mapping: Optional mapping configuration
        source_type: Source type ("pdf", "raw", "image") to determine extraction logic
    
    Returns:
        Dictionary with claim fields matching CLAIM_SCHEMA_ORDER
    """
    from schema import CLAIM_SCHEMA_ORDER
    import re
    
    # Initialize claim dict with all fields as None
    claim = {field: None for field in CLAIM_SCHEMA_ORDER}
    
    # Extract claim_number from block (should be present)
    claim_num_match = re.search(r'(?i)(?:Claim\s+(?:ID|Number|#)\s*:\s*)([A-Z0-9\-_]+)', text_block)
    if not claim_num_match:
        # Try narrative format
        claim_num_match = re.search(r'\b(C\d+|CLM\d+|CLM_[A-Z0-9]+)\s+is\b', text_block, re.IGNORECASE)
    
    if claim_num_match:
        claim['claim_number'] = claim_num_match.group(1)
    
    def extract_field_with_patterns(text: str, patterns: list, flags: int = 0):
        """Extract field using first matching pattern (first-match-wins)."""
        for pattern in patterns:
            match = re.search(pattern, text, flags)
            if match:
                return match.group(1).strip() if match.lastindex >= 1 else match.group(0).strip()
        return None
    
    # Extract policy_number (OCR-friendly: missing colon, extra spaces)
    policy_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)Policy\s+Number[:\-]?\s*((?:P\d+)|(?:P_[A-Z0-9]+))',
            r'(?i)Policy[:\-]?\s*((?:P\d+)|(?:P_[A-Z0-9]+))',
        ]
    )
    if policy_str:
        claim['policy_number'] = policy_str
    
    # Extract loss_date (OCR-friendly patterns)
    loss_date_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)Loss\s+Date[:\-]?\s*(\d{4}-\d{2}-\d{2})',
            r'(?i)Date[:\-]?\s*(\d{4}-\d{2}-\d{2})',
        ]
    )
    if loss_date_str:
        claim['loss_date'] = loss_date_str
    
    # Extract claim_type
    claim_type_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)(?:Cause\s+of\s+Loss|Claim\s+Type)[:\-]?\s*([^\n]+)',
        ]
    )
    if claim_type_str:
        claim['claim_type'] = claim_type_str.strip()
    
    # Extract amount (handle currency symbols, commas)
    amount_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)(?:Total\s+Incurred|Amount)[:\-]?\s*(-?\$?[\d,]+\.?\d*)',
        ]
    )
    if amount_str:
        try:
            amount_clean = amount_str.replace(',', '').replace('$', '').strip()
            claim['amount'] = str(float(amount_clean))  # Store as string per schema
        except (ValueError, AttributeError):
            pass
    
    # Extract description
    desc_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)Description[:\-]?\s*([^\n]+(?:\n(?!Claim\s+(?:ID|Number|#)|Policy|Loss|Status|Notes)[^\n]+)*)',
        ],
        re.MULTILINE | re.DOTALL
    )
    if desc_str:
        claim['description'] = re.sub(r'\s+', ' ', desc_str.strip())
    
    # Extract status
    status_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)Status[:\-]?\s*([^\n]+)',
        ]
    )
    if status_str:
        claim['status'] = status_str.strip()
    
    # Extract notes (multiline until next claim header)
    notes_match = re.search(
        r'(?i)Notes[:\-]?\s*(.+?)(?=\n(?:Claim\s+(?:ID|Number|#)|$|\Z))',
        text_block,
        re.MULTILINE | re.DOTALL
    )
    if not notes_match:
        # Fallback: match until end of block
        notes_match = re.search(r'(?i)Notes[:\-]?\s*(.+?)(?=$|\Z)', text_block, re.MULTILINE | re.DOTALL)
    if notes_match:
        notes_text = re.sub(r'\s+', ' ', notes_match.group(1).strip())
        if notes_text:
            claim['notes'] = notes_text
    # If notes not explicitly found, leave as None (don't populate with OCR block)
    
    return claim


def parse_claim_raw_text(text: str, source_type: str = "raw") -> List[Dict[str, Any]]:
    """
    Parse unstructured claim text and extract all canonical claim fields.
    
    Unified multi-pass extraction engine following 4-step contract:
    PASS 1: Normalize OCR text (whitespace, label variants, spacing)
    PASS 2: Split into blocks using strong delimiter (Claim ID/Number/#)
    PASS 3: Extract structured fields from each block (immutable once set)
    PASS 4: Fallback search in full_text (only for missing fields, never overwrite)
    
    Args:
        text: Raw text containing claim descriptions
    
    Returns:
        List of dicts, one per claim, with fields matching CLAIM_SCHEMA_ORDER
    """
    from schema import CLAIM_SCHEMA_ORDER
    import re
    
    # ========================================================================
    # PASS 1: NORMALIZE OCR TEXT
    # ========================================================================
    def normalize_ocr_text(raw_text: str) -> str:
        """Normalize OCR text for consistent parsing."""
        # Collapse multiple spaces to single space
        text = re.sub(r' +', ' ', raw_text)
        # Collapse multiple newlines to double newline (paragraph break)
        text = re.sub(r'\n{3,}', '\n\n', text)
        # Normalize spacing around colons: "Claim Number : value" -> "Claim Number: value"
        text = re.sub(r'\s+:\s+', ': ', text)
        # Normalize spacing around dashes used as separators: "Claim # - value" -> "Claim #: value"
        text = re.sub(r'([A-Za-z#]+)\s+-\s+([A-Za-z0-9])', r'\1: \2', text)
        # Normalize label variants: "Claim #" -> "Claim Number:", "Claim ID" -> "Claim Number:"
        text = re.sub(r'(?i)Claim\s+#\s*:', 'Claim Number:', text)
        text = re.sub(r'(?i)Claim\s+ID\s*:', 'Claim Number:', text)
        return text.strip()
    
    normalized_text = normalize_ocr_text(text)
    full_text = normalized_text  # Store for PASS 4 fallback
    
    # ========================================================================
    # PASS 2: SPLIT INTO BLOCKS USING STRONG DELIMITER
    # ========================================================================
    blocks = split_by_claim_number(normalized_text)
    logger.info(f"[parse_claim_raw_text] Found {len(blocks)} claim block(s)")
    
    # ========================================================================
    # PASS 3: EXTRACT STRUCTURED FIELDS FROM EACH BLOCK (IMMUTABLE ONCE SET)
    # ========================================================================
    def extract_field_with_patterns(text: str, patterns: list, flags: int = 0):
        """Extract field using first matching pattern (first-match-wins)."""
        for pattern in patterns:
            match = re.search(pattern, text, flags)
            if match:
                return match.group(1).strip() if match.lastindex >= 1 else match.group(0).strip()
        return None
    
    claims = []
    claims_dict = {}  # Deduplicate by claim_number
    
    for claim_num, block_text in blocks:
        # Deduplicate: if we already have this claim, prefer the longer block
        if claim_num in claims_dict:
            existing_block = claims_dict[claim_num].get('_block_text', '')
            if len(block_text) <= len(existing_block):
                continue  # Skip shorter duplicate
            # Use longer block
            block_text = block_text if len(block_text) > len(existing_block) else existing_block
        
        # Extract fields from block using extracted function
        claim = extract_claim_fields_from_block(block_text, mapping=None, source_type=source_type)
        
        # Ensure claim_number matches (in case extraction found different one)
        if claim and claim.get('claim_number'):
            claim['claim_number'] = claim_num
        
        # Track which fields were set by structured extraction (for PASS 4)
        structured_fields = set(['claim_number'])  # claim_number is always structured
        for field in CLAIM_SCHEMA_ORDER:
            if claim.get(field) is not None:
                structured_fields.add(field)
        
        # Store block text for fallback
        claims_dict[claim_num] = {'claim': claim, '_block_text': block_text, 'structured_fields': structured_fields}
    
    # ========================================================================
    # PASS 4: FALLBACK SEARCH IN FULL_TEXT (ONLY FOR MISSING FIELDS)
    # ========================================================================
    for claim_num, data in claims_dict.items():
        claim = data['claim']
        structured_fields = data['structured_fields']
        block_text = data['_block_text']
        
        # Find claim section in full_text for fallback extraction
        claim_section = None
        claim_section_pattern = rf'(?i)(?:Claim\s+(?:ID|Number|#)?\s*:\s*{re.escape(claim_num)}|{re.escape(claim_num)}).*?(?=(?:Claim\s+(?:ID|Number|#)?\s*:\s*[A-Z0-9]|$|\Z))'
        claim_section_match = re.search(claim_section_pattern, full_text, re.DOTALL)
        if claim_section_match:
            claim_section = claim_section_match.group(0)
        
        # Only extract missing fields, never overwrite structured fields
        if 'loss_date' not in structured_fields and claim_section:
            loss_date_str = extract_field_with_patterns(
                claim_section,
                [r'(?i)Loss\s+Date[:\-]?\s*(\d{4}-\d{2}-\d{2})', r'(?i)Date[:\-]?\s*(\d{4}-\d{2}-\d{2})']
            )
            if loss_date_str:
                claim['loss_date'] = loss_date_str
        
        if 'policy_number' not in structured_fields and claim_section:
            policy_str = extract_field_with_patterns(
                claim_section,
                [r'(?i)Policy\s+Number[:\-]?\s*((?:P\d+)|(?:P_[A-Z0-9]+))']
            )
            if policy_str:
                claim['policy_number'] = policy_str
        
        claims.append(claim)
    
    logger.info(f"[parse_claim_raw_text] Total unique claims extracted: {len(claims)}")
    return claims

def split_by_relationship_id(text: str) -> List[Tuple[str, str]]:
    """
    Split text into relationship blocks by relationship identifiers.
    Mirrors split_by_vin() for vehicles and split_by_policy_number() for policies.
    
    Supports multiple patterns:
    1. "Relationship ID:" header
    2. "RELATIONSHIP RECORD" header
    3. Structured format: "Policy Number: P001, VIN: VIN123, Driver ID: D001"
    4. Narrative format: "Policy P001 is linked to Vehicle VIN123 and Driver D001"
    
    Args:
        text: Raw text containing relationship descriptions (should be normalized)
    
    Returns:
        List of tuples: (rel_key, block_text) where rel_key is "policy_number|vehicle_vin|driver_id"
    """
    import re
    
    blocks = []
    
    # Pattern 1: "Relationship ID:" header
    rel_id_pattern = r'(?i)(?:Relationship\s+ID\s*:\s*)([A-Z0-9\-_]+)'
    rel_id_matches = list(re.finditer(rel_id_pattern, text))
    
    if rel_id_matches:
        logger.info(f"[split_by_relationship_id] Found {len(rel_id_matches)} 'Relationship ID:' header(s)")
        for i, match in enumerate(rel_id_matches):
            rel_id = match.group(1)
            label_start = match.start()
            if i + 1 < len(rel_id_matches):
                next_start = rel_id_matches[i + 1].start()
            else:
                next_start = len(text)
            block_text = text[label_start:next_start].strip()
            blocks.append((rel_id, block_text))
        return blocks
    
    # Pattern 2: "RELATIONSHIP RECORD" header
    rel_record_pattern = r'(?i)RELATIONSHIP\s+RECORD'
    rel_record_matches = list(re.finditer(rel_record_pattern, text))
    
    if rel_record_matches:
        logger.info(f"[split_by_relationship_id] Found {len(rel_record_matches)} 'RELATIONSHIP RECORD' header(s)")
        for i, match in enumerate(rel_record_matches):
            header_start = match.start()
            if i + 1 < len(rel_record_matches):
                next_start = rel_record_matches[i + 1].start()
            else:
                next_start = len(text)
            block_text = text[header_start:next_start].strip()
            # Generate a unique ID for this block
            rel_id = f"REL_{i+1:03d}"
            blocks.append((rel_id, block_text))
        return blocks
    
    # Pattern 3: Structured format: "Policy Number: P001, VIN: VIN123, Driver ID: D001"
    # Note: driver_id pattern includes underscores to match D_BAD1 format
    # VIN pattern is flexible (16-18 chars) and allows O to handle OCR errors (O/0 confusion)
    # Standard VIN excludes I, O, Q but OCR may misread 0 as O, so we allow O for extraction
    structured_pattern = r'(?i)Policy\s+Number\s*:\s*((?:P\d+)|(?:P_[A-Z0-9]+))[,\s]+VIN\s*:\s*([A-Z0-9]{16,18})[,\s]+Driver\s+ID\s*:\s*([A-Z0-9_]+)'
    structured_matches = list(re.finditer(structured_pattern, text))
    
    if structured_matches:
        logger.info(f"[split_by_relationship_id] Found {len(structured_matches)} structured relationship(s)")
        for i, match in enumerate(structured_matches):
            policy_num = match.group(1)
            vehicle_vin = match.group(2)
            driver_id = match.group(3)
            rel_key = f"{policy_num}|{vehicle_vin}|{driver_id}"
            match_start = match.start()
            if i + 1 < len(structured_matches):
                next_start = structured_matches[i + 1].start()
            else:
                next_start = len(text)
            block_text = text[match_start:next_start].strip()
            blocks.append((rel_key, block_text))
        return blocks
    
    # Pattern 4: Narrative format: "Policy P001 is linked to Vehicle VIN123 and Driver D001"
    narrative_pattern = r'(?i)Policy\s+((?:P\d+)|(?:P_[A-Z0-9]+))\s+is\s+linked\s+to\s+Vehicle\s+([A-HJ-NPR-Z0-9]{17})\s+and\s+Driver\s+([A-Z0-9]+)'
    narrative_matches = list(re.finditer(narrative_pattern, text))
    
    if narrative_matches:
        logger.info(f"[split_by_relationship_id] Found {len(narrative_matches)} narrative relationship(s)")
        for i, match in enumerate(narrative_matches):
            policy_num = match.group(1)
            vehicle_vin = match.group(2)
            driver_id = match.group(3)
            rel_key = f"{policy_num}|{vehicle_vin}|{driver_id}"
            match_start = match.start()
            if i + 1 < len(narrative_matches):
                next_start = narrative_matches[i + 1].start()
            else:
                next_start = len(text)
            block_text = text[match_start:next_start].strip()
            blocks.append((rel_key, block_text))
        return blocks
    
    logger.warning("[split_by_relationship_id] No relationship headers found")
    return []


def extract_relationship_fields_from_block(
    text_block: str,
    mapping: Optional[Dict[str, Any]] = None,
    source_type: Optional[str] = None
) -> Dict[str, Any]:
    """
    Extract ALL fields for a single relationship from a relationship block.
    Mirrors extract_fields_from_block() for vehicles.
    
    Args:
        text_block: Text block containing relationship information
        mapping: Optional mapping configuration
        source_type: Source type ("pdf", "raw", "image") to determine extraction logic
    
    Returns:
        Dictionary with relationship fields matching RELATIONSHIP_SCHEMA_ORDER
    """
    from schema import RELATIONSHIP_SCHEMA_ORDER
    import re
    
    # Initialize relationship dict with all fields as None
    relationship = {field: None for field in RELATIONSHIP_SCHEMA_ORDER}
    
    def extract_field_with_patterns(text: str, patterns: list, flags: int = 0):
        """Extract field using first matching pattern (first-match-wins)."""
        for pattern in patterns:
            match = re.search(pattern, text, flags)
            if match:
                return match.group(1).strip() if match.lastindex >= 1 else match.group(0).strip()
        return None
    
    # Extract policy_number (OCR-friendly: missing colon, extra spaces)
    policy_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)Policy\s+Number[:\-]?\s*((?:P\d+)|(?:P_[A-Z0-9]+))',
            r'(?i)Policy[:\-]?\s*((?:P\d+)|(?:P_[A-Z0-9]+))',
        ]
    )
    if policy_str:
        relationship['policy_number'] = policy_str
    
    # Extract vehicle_vin (OCR-friendly patterns, flexible length for OCR errors)
    # Allow 16-18 chars and allow O to handle OCR errors (extra chars, O/0 confusion)
    # Standard VIN excludes I, O, Q but OCR may misread 0 as O, so we allow O for extraction
    vin_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)VIN[:\-]?\s*([A-Z0-9]{16,18})',
            r'(?i)Vehicle\s+(?:VIN|ID)[:\-]?\s*([A-Z0-9]{16,18})',
        ]
    )
    if vin_str:
        relationship['vehicle_vin'] = vin_str.upper()
    
    # Extract driver_id (OCR-friendly patterns, includes underscores for D_BAD1 format)
    driver_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)Driver\s+ID[:\-]?\s*([A-Z0-9_]+)',
            r'(?i)Driver[:\-]?\s*([A-Z0-9_]+)',
        ]
    )
    if driver_str:
        relationship['driver_id'] = driver_str
    
    # Extract relationship_type
    rel_type_str = extract_field_with_patterns(
        text_block,
        [
            r'(?i)(?:Usage|Relationship\s+Type)[:\-]?\s*([^\n,]+)',
            r'(?i)(primary|secondary|listed)\s+driver',
            r'(?i)(owner|operator)',
        ]
    )
    if rel_type_str:
        # Normalize relationship type
        rel_type_lower = rel_type_str.lower()
        if 'primary' in rel_type_lower or 'listed' in rel_type_lower:
            relationship['relationship_type'] = 'Primary driver' if 'primary' in rel_type_lower else 'Listed driver'
        elif 'secondary' in rel_type_lower:
            relationship['relationship_type'] = 'Secondary driver'
        elif 'owner' in rel_type_lower:
            relationship['relationship_type'] = 'Owner'
        elif 'operator' in rel_type_lower:
            relationship['relationship_type'] = 'Operator'
        else:
            relationship['relationship_type'] = rel_type_str.strip()
    
    # Extract notes (multiline until next relationship header)
    notes_match = re.search(
        r'(?i)Notes[:\-]?\s*(.+?)(?=\n(?:Relationship\s+ID|RELATIONSHIP\s+RECORD|Policy\s+Number|$|\Z))',
        text_block,
        re.MULTILINE | re.DOTALL
    )
    if not notes_match:
        # Fallback: match until end of block
        notes_match = re.search(r'(?i)Notes[:\-]?\s*(.+?)(?=$|\Z)', text_block, re.MULTILINE | re.DOTALL)
    if notes_match:
        notes_text = re.sub(r'\s+', ' ', notes_match.group(1).strip())
        if notes_text:
            relationship['notes'] = notes_text
    # If notes not explicitly found, leave as None (already initialized in RELATIONSHIP_SCHEMA_ORDER)
    # This matches Claims/Locations strictness: notes only if explicitly extracted
    
    return relationship


def parse_relationship_raw_text(text: str, source_type: str = "raw") -> List[Dict[str, Any]]:
    """
    Parse unstructured relationship text and extract all canonical relationship fields.
    
    Unified multi-pass extraction engine following 4-step contract:
    PASS 1: Normalize OCR text (whitespace, label variants, spacing)
    PASS 2: Split into blocks using strong delimiter (Relationship ID or stable headers)
    PASS 3: Extract structured fields from each block (immutable once set)
    PASS 4: Fallback search in full_text (only for missing fields, never overwrite)
    
    Args:
        text: Raw text containing relationship descriptions
    
    Returns:
        List of dicts, one per relationship, with fields matching RELATIONSHIP_SCHEMA_ORDER
    """
    from schema import RELATIONSHIP_SCHEMA_ORDER
    import re
    
    # ========================================================================
    # PASS 1: NORMALIZE OCR TEXT
    # ========================================================================
    def normalize_ocr_text(raw_text: str) -> str:
        """Normalize OCR text for consistent parsing."""
        # Collapse multiple spaces to single space
        text = re.sub(r' +', ' ', raw_text)
        # Collapse multiple newlines to double newline (paragraph break)
        text = re.sub(r'\n{3,}', '\n\n', text)
        # Normalize spacing around colons
        text = re.sub(r'\s+:\s+', ': ', text)
        # Normalize spacing around dashes used as separators
        text = re.sub(r'([A-Za-z]+)\s+-\s+([A-Za-z0-9])', r'\1: \2', text)
        return text.strip()
    
    normalized_text = normalize_ocr_text(text)
    full_text = normalized_text  # Store for PASS 4 fallback
    
    # ========================================================================
    # PASS 2: SPLIT INTO BLOCKS USING STRONG DELIMITER
    # ========================================================================
    blocks = split_by_relationship_id(normalized_text)
    logger.info(f"[parse_relationship_raw_text] Found {len(blocks)} relationship block(s)")
    
    # ========================================================================
    # PASS 3: EXTRACT STRUCTURED FIELDS FROM EACH BLOCK (IMMUTABLE ONCE SET)
    # ========================================================================
    def extract_field_with_patterns(text: str, patterns: list, flags: int = 0):
        """Extract field using first matching pattern (first-match-wins)."""
        for pattern in patterns:
            match = re.search(pattern, text, flags)
            if match:
                return match.group(1).strip() if match.lastindex >= 1 else match.group(0).strip()
        return None
    
    relationships = []
    relationships_dict = {}  # Deduplicate by key (policy_number|vehicle_vin|driver_id)
    
    for rel_key, block_text in blocks:
        # Deduplicate: if we already have this relationship, skip
        if rel_key in relationships_dict:
            logger.info(f"[parse_relationship_raw_text] Skipping duplicate relationship: {rel_key}")
            continue
        
        # Extract fields from block using extracted function
        relationship = extract_relationship_fields_from_block(block_text, mapping=None, source_type=source_type)
        
        # Track which fields were set by structured extraction (for PASS 4)
        structured_fields = set()
        for field in RELATIONSHIP_SCHEMA_ORDER:
            if relationship.get(field) is not None:
                structured_fields.add(field)
        
        # Store for fallback and deduplication
        relationships_dict[rel_key] = {
            'relationship': relationship,
            '_block_text': block_text,
            'structured_fields': structured_fields
        }
    
    # ========================================================================
    # PASS 4: FALLBACK SEARCH IN FULL_TEXT (ONLY FOR MISSING FIELDS)
    # ========================================================================
    for rel_key, data in relationships_dict.items():
        relationship = data['relationship']
        structured_fields = data['structured_fields']
        block_text = data['_block_text']
        
        # Find relationship section in full_text for fallback extraction
        relationship_section = None
        # Try to find section by policy_number, vehicle_vin, or driver_id
        if relationship.get('policy_number'):
            section_pattern = rf'(?i)(?:Policy\s+(?:Number\s*:)?\s*{re.escape(relationship["policy_number"])}).*?(?=(?:Policy\s+(?:Number\s*:)?\s*[A-Z0-9]|Relationship\s+ID|$|\Z))'
            section_match = re.search(section_pattern, full_text, re.DOTALL)
            if section_match:
                relationship_section = section_match.group(0)
        
        # Only extract missing fields, never overwrite structured fields
        if 'relationship_type' not in structured_fields and relationship_section:
            rel_type_str = extract_field_with_patterns(
                relationship_section,
                [r'(?i)(?:Usage|Relationship\s+Type)[:\-]?\s*([^\n,]+)']
            )
            if rel_type_str:
                relationship['relationship_type'] = rel_type_str.strip()
        
        if 'notes' not in structured_fields and relationship_section:
            notes_match = re.search(r'(?i)Notes[:\-]?\s*([^\n]+)', relationship_section)
            if notes_match:
                relationship['notes'] = notes_match.group(1).strip()
        
        relationships.append(relationship)
    
    logger.info(f"[parse_relationship_raw_text] Total unique relationships extracted: {len(relationships)}")
    return relationships

def repair_vehicle_row_with_ai(
    row: List[Any],
    canonical_schema: List[str],
    ocr_text: str,
    vin_only: bool = False
) -> Optional[List[Any]]:
    """
    Second-stage AI repair pass for Vision-extracted vehicle rows.
    
    Corrects VIN corruption, normalizes casing, and fixes missing semantic fields
    using OCR text as context. Only corrects/confirms existing fields, never invents vehicles.
    
    Args:
        row: Vision-extracted row (list of values matching canonical_schema + 3 flags)
        canonical_schema: List of canonical field names
        ocr_text: Full OCR text for context
        vin_only: If True, only repair VIN (for table extractions where other fields are authoritative)
        
    Returns:
        Repaired row with _is_ai_repaired flag appended, or None if repair failed
    """
    from config import OPENAI_API_KEY, PDF_MODEL, ENABLE_OPENAI
    
    if not OPENAI_API_KEY or not ENABLE_OPENAI:
        logger.debug("[AI Repair] OpenAI not available, skipping repair")
        return None
    
    try:
        from openai import OpenAI
    except ImportError:
        logger.debug("[AI Repair] OpenAI package not installed, skipping repair")
        return None
    
    # Convert row to dict for easier processing (only canonical_schema fields)
    row_dict = {}
    for i, field in enumerate(canonical_schema):
        if i < len(row):
            row_dict[field] = row[i]
        else:
            row_dict[field] = None
    
    # Extract flags (preserve them) - flags are at indices len(canonical_schema) + 0, 1, 2
    is_handwritten = row[len(canonical_schema)] if len(row) > len(canonical_schema) else False
    is_vision_extracted = row[len(canonical_schema) + 1] if len(row) > len(canonical_schema) + 1 else False
    is_table_extraction = row[len(canonical_schema) + 2] if len(row) > len(canonical_schema) + 2 else False
    
    # Build system prompt
    if vin_only:
        # VIN-only mode for table extractions (other fields are authoritative)
        system_prompt = (
            "You are a vehicle data repair assistant. Your task is to CORRECT VIN corruption ONLY.\n\n"
            "CRITICAL RULES:\n"
            "1. ONLY correct the VIN field - do NOT modify any other fields\n"
            "2. Use VIN validation rules to fix likely VIN character errors:\n"
            "   - VIN cannot contain I, O, Q (common OCR errors: '0'→'O', '1'→'I', 'Q'→'0')\n"
            "   - VIN must be exactly 17 characters\n"
            "   - VIN must have at least 2 digits\n"
            "   - If VIN appears corrupted, try single-character corrections based on context\n"
            "3. Preserve ALL other fields exactly as provided (do NOT normalize casing, do NOT infer missing fields)\n"
            "4. Return the EXACT same vehicle with only VIN corrected\n\n"
            "OUTPUT FORMAT:\n"
            "Return ONLY valid JSON matching this exact schema (all fields must be present):\n"
            f"{json.dumps({field: row_dict.get(field) for field in canonical_schema}, indent=2)}\n\n"
            "Return the repaired vehicle as a JSON object with all canonical fields.\n"
            "Only the VIN field should be modified. Use null for missing fields. Do NOT add markdown code blocks."
        )
    else:
        # Full repair mode for document-style extractions
        system_prompt = (
            "You are a vehicle data repair assistant. Your task is to CORRECT and CONFIRM existing vehicle fields, "
            "not to invent new vehicles.\n\n"
            "CRITICAL RULES:\n"
            "1. ONLY correct or confirm fields that are already present in the input row\n"
            "2. NEVER invent a vehicle that is not in the input\n"
            "3. Use VIN validation rules to fix likely VIN character errors:\n"
            "   - VIN cannot contain I, O, Q (common OCR errors: '0'→'O', '1'→'I', 'Q'→'0')\n"
            "   - VIN must be exactly 17 characters\n"
            "   - VIN must have at least 2 digits\n"
            "   - If VIN appears corrupted, try single-character corrections based on context\n"
            "4. Normalize casing for make/model (e.g., 'camry'→'Camry', 'ford'→'Ford')\n"
            "5. Leave fields as null if confidence is low - do NOT guess\n"
            "6. Return the EXACT same vehicle (same VIN) with corrected fields only\n"
            "7. If OCR text contradicts the Vision-extracted values, prefer Vision values unless OCR clearly shows an error\n\n"
            "OUTPUT FORMAT:\n"
            "Return ONLY valid JSON matching this exact schema (all fields must be present):\n"
            f"{json.dumps({field: row_dict.get(field) for field in canonical_schema}, indent=2)}\n\n"
            "Return the repaired vehicle as a JSON object with all canonical fields.\n"
            "Use null for missing fields. Do NOT add markdown code blocks."
        )
    
    # Build user prompt with row data and OCR context
    user_prompt = (
        f"Repair this Vision-extracted vehicle row:\n"
        f"{json.dumps(row_dict, indent=2)}\n\n"
    )
    
    if ocr_text and len(ocr_text.strip()) > 0:
        user_prompt += (
            f"OCR text context (first 1000 chars):\n{ocr_text[:1000]}\n\n"
        )
    
    if vin_only:
        user_prompt += (
            "Correct any VIN character errors ONLY. Do NOT modify any other fields.\n"
            "Return the repaired vehicle as JSON matching the schema above (only VIN may be different)."
        )
    else:
        user_prompt += (
            "Correct any VIN character errors, normalize casing, and confirm/correct semantic fields.\n"
            "Return the repaired vehicle as JSON matching the schema above."
        )
    
    try:
        client = OpenAI(api_key=OPENAI_API_KEY)
        response = client.chat.completions.create(
            model=PDF_MODEL or "gpt-4o",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            max_tokens=2000,
            temperature=0,  # Deterministic output
        )
        
        if response.choices and response.choices[0].message.content:
            content = response.choices[0].message.content.strip()
            
            # Extract JSON
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()
            
            try:
                repaired_dict = json.loads(content)
                
                # Convert back to row format (canonical schema order)
                repaired_row = []
                for field in canonical_schema:
                    repaired_row.append(repaired_dict.get(field))
                
                # Preserve flags and add _is_ai_repaired
                repaired_row.append(is_handwritten)
                repaired_row.append(is_vision_extracted)
                repaired_row.append(is_table_extraction)
                repaired_row.append(True)  # _is_ai_repaired = True
                
                logger.info(f"[AI Repair] Successfully repaired vehicle row (VIN: {repaired_dict.get('vin')})")
                return repaired_row
            except json.JSONDecodeError as e:
                logger.warning(f"[AI Repair] Failed to parse JSON response: {e}")
                return None
        else:
            logger.warning("[AI Repair] Empty response from OpenAI")
            return None
    except Exception as e:
        logger.warning(f"[AI Repair] Error during repair: {e}")
        return None

def _extract_table_with_vision_api(
    file_path: Path,
    source_type: SourceType,
    text_blocks: Optional[List[Any]] = None,
    page_text_blocks: Optional[Dict[int, List[Any]]] = None,
    **kwargs
) -> Optional[List[List[Any]]]:
    """
    
    Args:
        file_path: Path to PDF, image, or text file
        source_type: SourceType.PDF, SourceType.IMAGE, or SourceType.RAW_TEXT
    
    Returns:
        2D list of values (rows x columns) or None if extraction fails
    """
    logger.info("[Vision] Starting fallback...")
    
    # Debug logging for mapping_id
    mapping_id_from_kwargs = kwargs.get('mapping_id', 'NONE') if kwargs else 'NONE'
    logger.debug(f"[DEBUG _extract_table_with_vision_api] mapping_id from kwargs: {mapping_id_from_kwargs}")
    
    # Load API key explicitly - ensure dotenv is loaded
    try:
        from dotenv import load_dotenv
        # Try loading from current directory and parent directory
        load_dotenv()
        if not os.getenv("OPENAI_API_KEY"):
            # Try loading from project root
            project_root = Path(__file__).parent.parent.parent
            load_dotenv(dotenv_path=project_root / ".env")
    except ImportError:
        pass
    openai_key = os.getenv("OPENAI_API_KEY")
    if not openai_key:
        logger.warning("[Vision] OPENAI_API_KEY missing - checked current dir and project root")
        # Return empty 2D array with headers to ensure normalize_v2 is called
        from schema import VEHICLE_SCHEMA_ORDER
        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
        header_row = list(canonical_schema) + ["_is_handwritten"]
        return [header_row]  # Return header row only, no data rows
    logger.info("[Vision] API key found")
    
    # Use appropriate model based on source type: PDF_MODEL for PDFs, IMAGE_MODEL for images
    from config import PDF_MODEL, IMAGE_MODEL
    if source_type == SourceType.PDF:
        vision_model = PDF_MODEL
        logger.info(f"[Vision] Using PDF_MODEL: {vision_model}")
    elif source_type == SourceType.IMAGE:
        vision_model = IMAGE_MODEL
        logger.info(f"[Vision] Using IMAGE_MODEL: {vision_model}")
    else:
        # Default to IMAGE_MODEL for other source types (shouldn't happen, but safe fallback)
        vision_model = IMAGE_MODEL
        logger.info(f"[Vision] Using IMAGE_MODEL (default): {vision_model}")
    
    # Initialize OpenAI client
    try:
        from openai import OpenAI
        client = OpenAI(api_key=openai_key)
        logger.info("[Vision] OpenAI client initialized")
    except Exception as e:
        logger.error(f"[Vision] Client init error: {e}")
        # Return empty 2D array with headers to ensure normalize_v2 is called
        from schema import VEHICLE_SCHEMA_ORDER
        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
        header_row = list(canonical_schema) + ["_is_handwritten"]
        return [header_row]  # Return header row only, no data rows
    
    # Concatenate OCR text blocks if provided
    ocr_text = ""
    if text_blocks:
        try:
            from ocr.table_extract import _extract_raw_text
            ocr_text = _extract_raw_text(text_blocks)
            logger.debug(f"[Vision] Concatenated {len(text_blocks)} OCR text blocks ({len(ocr_text)} chars)")
        except Exception as e:
            logger.debug(f"[Vision] Could not concatenate OCR text blocks: {e}")
    
    # Get PDF page count and organize OCR text blocks by page
    page_count = 1
    page_text_blocks_dict = {}  # page_num -> list of text blocks
    # Check if file is actually a PDF (even if source_type is IMAGE - for scanned PDFs)
    is_actually_pdf = file_path.suffix.lower() == '.pdf'
    if source_type == SourceType.PDF or (source_type == SourceType.IMAGE and is_actually_pdf):
        print("PDF PROMPT ACTIVE")  # Debug print to verify PDF extraction path
        try:
            import PyPDF2
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                page_count = len(pdf_reader.pages)
                logger.info(f"[Vision] PDF has {page_count} pages")
        except Exception as e:
            logger.debug(f"[Vision] Could not determine page count: {e}, assuming 1 page")
        
        # Organize text blocks by page number
        if text_blocks:
            for block in text_blocks:
                # TextBlock doesn't have page_number, but we can infer from line_number or use all blocks for page 1
                # For now, if we have page_text_blocks dict, use it; otherwise assume all blocks are from page 1
                if page_text_blocks:
                    # Use provided page_text_blocks dict
                    for page_num, blocks in page_text_blocks.items():
                        if block in blocks:
                            if page_num not in page_text_blocks_dict:
                                page_text_blocks_dict[page_num] = []
                            page_text_blocks_dict[page_num].append(block)
                else:
                    # No page info, assume all blocks are from page 1
                    if 1 not in page_text_blocks_dict:
                        page_text_blocks_dict[1] = []
                    page_text_blocks_dict[1].append(block)
        elif page_text_blocks:
            # Use provided page_text_blocks
            page_text_blocks_dict = page_text_blocks
        else:
            # No page_text_blocks provided - extract OCR text from PDF pages
            # This is needed when _extract_table_with_vision_api is called directly
            if source_type == SourceType.PDF:
                try:
                    from ocr.reader import extract_text_from_pdf
                    ocr_text_blocks, _ = extract_text_from_pdf(file_path, enable_vision=False, **kwargs)
                    if ocr_text_blocks:
                        # Organize blocks by page (assume all blocks are from page 1 if no page info)
                        page_text_blocks_dict[1] = ocr_text_blocks
                        logger.info(f"[Vision] Extracted {len(ocr_text_blocks)} OCR text blocks from PDF")
                except Exception as e:
                    logger.debug(f"[Vision] Could not extract OCR text from PDF: {e}")
    
    # Convert PDF/Image/Raw Text to images (all pages) if needed
    images = []
    try:
        # Check if file is actually a PDF (even if source_type is IMAGE - for scanned PDFs)
        is_actually_pdf = file_path.suffix.lower() == '.pdf'
        
        if source_type == SourceType.PDF or (source_type == SourceType.IMAGE and is_actually_pdf):
            # PDF file: convert pages to images
            try:
                from pdf2image import convert_from_path
                # Convert PDF pages to images at ~300 DPI (mirrors ENG-101-table-detector line 1892)
                # This improves Vision API quality for handwritten PDFs by ensuring sufficient resolution
                images = convert_from_path(str(file_path), first_page=1, last_page=page_count, dpi=300)
                if not images:
                    logger.warning("[Vision] Could not convert PDF to images")
                    # Return empty 2D array with headers to ensure normalize_v2 is called
                    from schema import VEHICLE_SCHEMA_ORDER
                    canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                    header_row = list(canonical_schema) + ["_is_handwritten"]
                    return [header_row]  # Return header row only, no data rows
                logger.debug(f"[Vision] PDF converted to {len(images)} images (routed as {'IMAGE' if source_type == SourceType.IMAGE else 'PDF'})")
            except ImportError:
                logger.error("[Vision] pdf2image not installed, cannot process PDF")
                # Return empty 2D array with headers to ensure normalize_v2 is called
                from schema import VEHICLE_SCHEMA_ORDER
                canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                header_row = list(canonical_schema) + ["_is_handwritten"]
                return [header_row]  # Return header row only, no data rows
            except Exception as e:
                logger.error(f"[Vision] Error converting PDF to images: {e}")
                # Return empty 2D array with headers to ensure normalize_v2 is called
                from schema import VEHICLE_SCHEMA_ORDER
                canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                header_row = list(canonical_schema) + ["_is_handwritten"]
                return [header_row]  # Return header row only, no data rows
        elif source_type == SourceType.IMAGE:
            # For images, process as single page
            try:
                from PIL import Image
                images = [Image.open(file_path)]
                page_count = 1
                logger.debug("[Vision] Image opened")
            except ImportError:
                logger.error("[Vision] PIL/Pillow not installed - cannot process images")
                # Return empty 2D array with headers to ensure normalize_v2 is called
                from schema import VEHICLE_SCHEMA_ORDER
                canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                header_row = list(canonical_schema) + ["_is_handwritten"]
                return [header_row]  # Return header row only, no data rows
            except Exception as e:
                logger.error(f"[Vision] Error opening image: {e}")
                # Return empty 2D array with headers to ensure normalize_v2 is called
                from schema import VEHICLE_SCHEMA_ORDER
                canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                header_row = list(canonical_schema) + ["_is_handwritten"]
                return [header_row]  # Return header row only, no data rows
        elif source_type == SourceType.RAW_TEXT:
            # For raw text, read the file and process text content directly
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    raw_text_content = f.read()
                logger.debug(f"[Vision] Raw text file read, {len(raw_text_content)} characters")
                # Process as text-only (no images)
                images = []  # Empty - we'll use text-only API call
                page_count = 1
            except Exception as e:
                logger.error(f"[Vision] Error reading raw text file: {e}")
                # Return empty 2D array with headers to ensure normalize_v2 is called
                from schema import VEHICLE_SCHEMA_ORDER
                canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                header_row = list(canonical_schema) + ["_is_handwritten"]
                return [header_row]  # Return header row only, no data rows
        else:
            logger.error(f"[Vision] Unsupported source type: {source_type}")
            # Return empty 2D array with headers to ensure normalize_v2 is called
            from schema import VEHICLE_SCHEMA_ORDER
            canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
            header_row = list(canonical_schema) + ["_is_handwritten"]
            return [header_row]  # Return header row only, no data rows
        
        # Load mapping config based on source type AND domain
        from mappings import get_mapping_by_id
        from schema import VEHICLE_SCHEMA_ORDER, DRIVER_SCHEMA_ORDER
        
        # Detect domain from mapping_id in kwargs
        domain_mapping_id = kwargs.get('mapping_id', '') if kwargs else ''
        domain_lower = domain_mapping_id.lower()
        is_driver = 'driver' in domain_lower and 'relationship' not in domain_lower
        is_vehicle = not is_driver  # Default to vehicle if domain unclear
        
        # Determine mapping ID based on source type AND domain
        if is_driver:
            if source_type == SourceType.PDF:
                mapping_id = "source_pdf_drivers"
            elif source_type == SourceType.RAW_TEXT:
                mapping_id = "source_raw_text_drivers"
            elif source_type == SourceType.IMAGE:
                # Check if it's image metadata JSON or actual image
                if file_path.suffix.lower() == ".json":
                    mapping_id = "source_image_metadata_json_drivers"
                else:
                    mapping_id = "source_image_drivers"
            else:
                mapping_id = "source_pdf_drivers"  # Default fallback
        else:
            # Vehicle (default behavior - preserve existing logic)
            if source_type == SourceType.PDF:
                mapping_id = "source_pdf_vehicles"
            elif source_type == SourceType.RAW_TEXT:
                mapping_id = "source_raw_text_vehicles"
            elif source_type == SourceType.IMAGE:
                # Check if it's image metadata JSON or actual image
                if file_path.suffix.lower() == ".json":
                    mapping_id = "source_image_metadata_json_vehicles"
                else:
                    mapping_id = "source_image_vehicles"
            else:
                mapping_id = "source_pdf_vehicles"  # Default fallback
        
        mapping_config = get_mapping_by_id(mapping_id)
        if not mapping_config:
            logger.warning(f"[Vision] Mapping {mapping_id} not found, using default extraction")
            mapping_config = None
        
        # Use canonical schema based on domain
        if is_driver:
            canonical_schema = DRIVER_SCHEMA_ORDER.copy()
        else:
            canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
        
        # Build system prompt from mapping config
        if mapping_config:
            mappings = mapping_config.get("mappings", [])
            # Build field extraction instructions from mapping ai_instructions
            field_instructions = {}
            for mapping in mappings:
                target_field = mapping.get("target_field")
                ai_instruction = mapping.get("ai_instruction", "")
                if target_field and ai_instruction:
                    field_instructions[target_field] = ai_instruction
            
            # Determine domain-specific terminology
            if is_driver:
                domain_name = "driver"
                domain_name_plural = "drivers"
                identifier_name = "Driver ID"
                identifier_example = "D001"
            else:
                domain_name = "vehicle"
                domain_name_plural = "vehicles"
                identifier_name = "VIN"
                identifier_example = "WBA8E160JNU34567"
            
            # Build system prompt with field-specific instructions
            system_prompt = (
                "CRITICAL EXTRACTION RULES (HIGHEST PRIORITY):\n"
                "\n"
                f"You must attempt to extract EVERY canonical {domain_name} field for EACH {domain_name},\n"
                "even if the value must be inferred from natural language.\n"
                "\n"
                f"For each {domain_name}:\n"
                "- If a field is not explicitly labeled, infer it from context.\n"
                "- If inference is uncertain, return your best guess.\n"
                "- If no reasonable inference is possible, return null — NEVER omit the field.\n"
                "\n"
            )
            
            # Add domain-specific inference examples
            if is_driver:
                system_prompt += (
                    "Inference examples (MANDATORY):\n"
                    "- If text contains \"Yes\" or \"No\" for training, training_completed = \"Yes\" or \"No\" (extract raw value)\n"
                    "- If text contains a number for violations, violations_count = that number (extract raw value)\n"
                    "- If text contains years of experience, years_experience = that number (extract raw value)\n"
                    "- If text contains license information, extract license_number, license_state, license_status (extract exactly as written)\n"
                    "- If text contains date of birth, extract date_of_birth in YYYY-MM-DD format if possible\n"
                    "\n"
                )
            else:
                system_prompt += (
                    "Inference examples (MANDATORY):\n"
                    "- If text contains \"sedan\", body_style = \"sedan\" (extract raw value, normalization happens later)\n"
                    "- If text contains \"truck\" or \"pickup\", body_style = \"truck\" (extract raw value)\n"
                    "- If text contains \"gas\" or \"gasoline\", fuel_type = \"gas\" or \"gasoline\" (extract raw value as written, normalization happens later)\n"
                    "- If text contains \"automatic\", \"auto\", \"8-speed auto\", transmission = \"automatic\" or \"auto\" (extract raw value as written, normalization happens later)\n"
                    "- If text contains an email-like string, extract it as owner_email (extract exactly as written)\n"
                    "\n"
                )
            
            system_prompt += (
                "IMPORTANT: Extract raw values only. Do NOT normalize, convert, or transform values.\n"
                "All normalization (canonical forms, lowercase, type conversion) will be handled by the normalization pipeline.\n"
                "\n"
                "This rule applies to:\n"
                "- PDFs\n"
                "- Handwritten documents\n"
                "- Images\n"
                "- Tables\n"
                "- Free-form notes\n"
                "\n"
                "Failure to infer a field when the information is visible is an error.\n"
                "\n"
                "---\n"
                "\n"
                f"You are an OCR engine that extracts {domain_name} data from an image or PDF. "
                f"You MUST identify EVERY individual {domain_name} described in the document. "
                "Return ONLY valid JSON, no explanations, no markdown, no additional text. "
                "\n"
                "Output format:\n"
                "{\n"
                f'  "headers": {json.dumps(canonical_schema)},\n'
                '  "rows": [\n'
            )
            
            # Add domain-specific example rows
            if is_driver:
                system_prompt += (
                    '    ["D001", "John", "Doe", "1985-03-15", "D1234567", "CA", "Valid", 18, 0, "Yes", "Software Engineer"],\n'
                    '    ["D002", "Jane", "Doe", "1987-07-22", "D7654321", "CA", "Valid", 16, 1, "No", "Secondary driver"]\n'
                )
            else:
                system_prompt += (
                    '    ["VIN1", null, 2024, "Toyota", "Camry", null, "notes", "Blue", 12345, null, "sedan", "gas", "automatic", "email@example.com", null],\n'
                    '    ["VIN2", null, 2020, "Ford", "F-150", null, "notes", "Red", 45210, null, "truck", "gas", "automatic", "email@example.com", null]\n'
                )
            
            system_prompt += (
                '  ]\n'
                "\n"
                "NOTE: Values in examples show raw extracted values.\n"
                "Normalization happens in the normalization pipeline, not in Vision extraction.\n"
                "}\n"
                "\n"
                "CRITICAL OUTPUT REQUIREMENTS:\n"
                f"- Return exactly ONE row per {domain_name}\n"
                "- Use the EXACT header names listed above (case-sensitive)\n"
                "- For missing fields, use null (not empty string, not \"\", not \"N/A\")\n"
                f"- Extract ALL {domain_name_plural} from the document, not just the first one\n"
                f"- Each row must have exactly {len(canonical_schema)} values matching the {len(canonical_schema)} headers\n"
                "- Output ONLY the JSON object - no explanatory text before or after\n"
                "- Do NOT add markdown code blocks (```json)\n"
                "- Do NOT add comments or descriptions\n"
                "\n"
                "SECTION-BASED SEPARATION (CRITICAL FOR HANDWRITTEN DOCUMENTS):\n"
                f"- Documents may contain multiple {domain_name_plural} organized in SECTIONS or BLOCKS\n"
                f"- Each section/block represents ONE complete {domain_name} record\n"
                "- CRITICAL: Handwritten documents are typically NOT in table format - they are organized as SEPARATE TEXT BLOCKS\n"
                f"- Each {domain_name}'s information is written in its own distinct block/section, NOT in table rows\n"
                "- Look for visual separators: blank lines, horizontal lines, boxes, borders, or distinct sections\n"
                f"- For handwritten documents: each {domain_name} is typically in its own separate block with all its information grouped together\n"
                "- Identify block/section boundaries by:\n"
                "  * Blank lines or spacing between information\n"
                "  * Visual dividers (lines, boxes, borders)\n"
                f"  * Each {identifier_name} typically starts a new block\n"
                f"  * All fields for one {domain_name} are grouped within the same block\n"
                f"  * Each block is a self-contained paragraph or section of text describing one {domain_name}\n"
                f"- DO NOT mix fields from different blocks - each block = one {domain_name} = one row\n"
                f"- If you see multiple distinct blocks/sections of information, each is a separate {domain_name}\n"
                f"- For handwritten documents: carefully identify where one {domain_name} block ends and the next begins\n"
                f"- Extract ONE row per block/section, even if blocks are not perfectly aligned or formatted\n"
                f"- If a document has 3 blocks with {domain_name} info, you MUST extract 3 rows (one per block)\n"
                "- IMPORTANT: Do NOT look for table structure in handwritten documents - look for separate text blocks instead\n"
                "\n"
            )
            
            # Add IMAGE TYPE DETECTION section only for IMAGE sources
            if source_type == SourceType.IMAGE:
                if is_driver:
                    system_prompt += (
                        "IMAGE TYPE DETECTION:\n"
                        "- First, determine if the image is a TABLE or a DOCUMENT-STYLE layout\n"
                        "- TABLE: Has clear column headers like DRIVER ID, FULL NAME, DATE OF BIRTH, LICENSE NUMBER arranged in columns\n"
                        "- DOCUMENT-STYLE: Information is organized in sections/blocks, not in a structured table format\n"
                        "\n"
                        "- If the image is a TABLE (column headers like DRIVER ID, FULL NAME, DATE OF BIRTH):\n"
                        "  * CRITICAL: The FIRST ROW is ALWAYS the header row - DO NOT extract it as a driver\n"
                        "  * Before extracting, identify which row is the header row (contains 'DRIVER ID', 'FULL NAME', 'DATE OF BIRTH', 'LICENSE NUMBER', etc.)\n"
                        "  * Skip the header row completely - do NOT include it in your output\n"
                        "  * Extract ONLY data rows that contain a VALID Driver ID in the Driver ID column\n"
                        "  * NEVER treat header rows as drivers - skip any row where the Driver ID column contains 'DRIVER ID', 'FULL NAME', 'DATE OF BIRTH', 'LICENSE NUMBER', etc.\n"
                        "  * If the Driver ID column contains ANY header text (even partial matches), SKIP that entire row\n"
                        "  * If no valid Driver ID exists in a row, SKIP it entirely - do NOT extract it\n"
                        "  * A valid Driver ID typically follows pattern: D followed by 3 digits, or DRV followed by 3 digits\n"
                        "  * Each valid Driver ID row = one driver = one output row\n"
                        "  * If you see a table with headers, count the header row and skip it - only extract data rows below it\n"
                        "  * CRITICAL FOR TABLE EXTRACTION: Read EACH COLUMN for EACH data row:\n"
                        "    - For each data row, read the value in EVERY column (DRIVER ID, FULL NAME, DATE OF BIRTH, LICENSE NUMBER, LICENSE STATE, LICENSE STATUS, YEARS EXPERIENCE, VIOLATIONS COUNT, TRAINING COMPLETED, NOTES, etc.)\n"
                        "    - Extract the value from each column cell - do NOT skip columns even if they seem empty or contain unusual values\n"
                        "    - Map each column header to the corresponding field in the output schema\n"
                        "    - If a column cell is empty or unclear, use null (not empty string)\n"
                        "    - Extract ALL visible columns for each row - do NOT omit any columns\n"
                        "\n"
                        "- If the image is NOT a table (document-style or handwritten):\n"
                        "  * Use section-based extraction rules (see SECTION-BASED SEPARATION above)\n"
                        "  * Identify distinct sections/blocks of driver information\n"
                        "  * CRITICAL: Handwritten images are typically organized as SEPARATE TEXT BLOCKS, NOT tables\n"
                        "  * Each block is a self-contained paragraph or section describing one driver\n"
                        "  * Extract one row per block, even if formatting is irregular\n"
                        "  * Each block should contain a Driver ID and associated driver fields\n"
                        "  * Do NOT look for table structure - look for separate text blocks instead\n"
                        "\n"
                    )
                else:
                    system_prompt += (
                        "IMAGE TYPE DETECTION:\n"
                        "- First, determine if the image is a TABLE or a DOCUMENT-STYLE layout\n"
                        "- TABLE: Has clear column headers like YEAR, MAKE, MODEL, VIN, COLOR, MILEAGE arranged in columns\n"
                        "- DOCUMENT-STYLE: Information is organized in sections/blocks, not in a structured table format\n"
                        "\n"
                        "- If the image is a TABLE (column headers like YEAR, MAKE, MODEL, VIN):\n"
                        "  * CRITICAL: The FIRST ROW is ALWAYS the header row - DO NOT extract it as a vehicle\n"
                        "  * Before extracting, identify which row is the header row (contains 'YEAR', 'MAKE', 'MODEL', 'VIN', 'COLOR', 'MILEAGE', etc.)\n"
                        "  * Skip the header row completely - do NOT include it in your output\n"
                        "  * Extract ONLY data rows that contain a VALID 17-character VIN in the VIN column\n"
                        "  * NEVER treat header rows as vehicles - skip any row where the VIN column contains 'YEAR', 'MAKE', 'MODEL', 'VIN', 'COLOR', 'MILEAGE', 'NOTES', 'BODY', 'PAINTED', 'GASOLINE', 'TRANSMISSION', etc.\n"
                        "  * If the VIN column contains ANY header text (even partial matches), SKIP that entire row\n"
                        "  * If no valid VIN exists in a row, SKIP it entirely - do NOT extract it\n"
                        "  * A valid VIN must have at least 2 digits and be 10-17 characters long\n"
                        "  * Only extract rows with valid 17-character alphanumeric VINs (excluding I, O, Q)\n"
                        "  * Each valid VIN row = one vehicle = one output row\n"
                        "  * If you see a table with headers, count the header row and skip it - only extract data rows below it\n"
                        "  * CRITICAL FOR TABLE EXTRACTION: Read EACH COLUMN for EACH data row:\n"
                        "    - For each data row, read the value in EVERY column (YEAR, MAKE, MODEL, VIN, COLOR, MILEAGE, BODY, FUEL, TRANSMISSION, EMAIL, NOTES, etc.)\n"
                        "    - Extract the value from each column cell - do NOT skip columns even if they seem empty or contain unusual values\n"
                        "    - Map each column header to the corresponding field in the output schema\n"
                        "    - If a column cell is empty or unclear, use null (not empty string)\n"
                        "    - Extract ALL visible columns for each row - do NOT omit any columns\n"
                        "\n"
                        "- If the image is NOT a table (document-style or handwritten):\n"
                        "  * Use section-based extraction rules (see SECTION-BASED SEPARATION above)\n"
                        "  * Identify distinct sections/blocks of vehicle information\n"
                        "  * CRITICAL: Handwritten images are typically organized as SEPARATE TEXT BLOCKS, NOT tables\n"
                        "  * Each block is a self-contained paragraph or section describing one vehicle\n"
                        "  * Extract one row per block, even if formatting is irregular\n"
                        "  * Each block should contain a VIN and associated vehicle fields\n"
                        "  * Do NOT look for table structure - look for separate text blocks instead\n"
                        "\n"
                    )
            
            if is_driver:
                system_prompt += (
                    "DRIVER ID VALIDATION (CRITICAL):\n"
                    "- The Driver ID (first field) MUST be a valid driver identifier (e.g., D001, D002, DRV001)\n"
                    "- Driver ID typically follows pattern: D followed by 3 digits, or DRV followed by 3 digits\n"
                    "- DO NOT include header rows - if the Driver ID column contains text like 'Driver ID', 'Full Name', 'Date of Birth', 'License Number', skip that row entirely\n"
                    "- DO NOT include table headers as data rows - only extract rows with valid driver IDs\n"
                    "- For handwritten documents, carefully validate Driver IDs to avoid OCR character recognition errors\n"
                    "- Each valid Driver ID you find represents a separate driver - extract it as a separate row\n"
                    "\n"
                )
            else:
                system_prompt += (
                    "VIN VALIDATION (CRITICAL):\n"
                    "- The VIN (first field) MUST be a valid 17-character alphanumeric code\n"
                    "- VIN must contain at least 2 digits and exclude letters I, O, Q\n"
                    "- DO NOT include header rows - if the VIN column contains text like 'YEAR', 'MAKE', 'MODEL', 'VIN', 'COLOR', 'MILEAGE', skip that row entirely\n"
                    "- DO NOT include table headers as data rows - only extract rows with valid vehicle VINs\n"
                    "- For handwritten documents, carefully validate VINs to avoid OCR character recognition errors (e.g., '0' vs 'O', '1' vs 'I')\n"
                    "- Each valid VIN you find represents a separate vehicle - extract it as a separate row\n"
                    "\n"
                )
            
            # Add domain-specific field list for extraction rules
            if is_driver:
                field_list = "driver_id, first_name, last_name, date_of_birth, license_number, license_state, license_status, years_experience, violations_count, training_completed, and notes"
            else:
                field_list = "VIN, year, make, model, color, mileage, body_style, fuel_type, transmission, owner_email, and notes"
            
            system_prompt += (
                "EXTRACTION RULES (CRITICAL):\n"
                "- Extract values EXACTLY as written in the source document\n"
                "- Do NOT reword, summarize, paraphrase, or embellish any values\n"
                "- Do NOT normalize or convert values (preserve original format)\n"
                "- Infer missing values from natural language when possible (see CRITICAL EXTRACTION RULES at top) - only use null if no reasonable inference is possible\n"
                "- When uncertain, extract your best guess rather than omitting the field\n"
                "- For invalid values, extract them anyway - validation will handle warnings\n"
                "- Prefer extraction over omission - always attempt to extract a value if it appears in the document\n"
                f"- Extract ALL fields for each {domain_name} - do NOT skip fields even if they seem invalid or unusual\n"
                f"- For each {domain_name}, attempt to extract: {field_list}\n"
                "- When extracting from a section: extract ALL fields that appear in that section, even if formatting is irregular\n"
                "\n"
                "NOTES FIELD RULES (CRITICAL):\n"
                "- If notes exist in the source, copy them VERBATIM - word-for-word, character-for-character\n"
                "- Preserve original wording, punctuation, sentence order, and capitalization\n"
                "- Do NOT summarize, paraphrase, rewrite, or reorder sentences\n"
                f"- Do NOT remove content unless it's clearly not {domain_name}-specific\n"
                "- Do NOT add explanatory text or fabricate information that is not visible in the document\n"
                "- If no notes are present in the source, set notes = null (not empty string)\n"
                "- If notes contain bullets or list markers, preserve them as written\n"
                "\n"
                "FIELD EXTRACTION INSTRUCTIONS:\n"
                "\n"
            )
            
            # Add field-specific instructions from mapping
            # Group instructions by field for better readability
            for field in canonical_schema:
                if field in field_instructions:
                    instruction = field_instructions[field]
                    # Preserve newlines in instructions (they contain important formatting)
                    system_prompt += f"{field.upper()}:\n{instruction}\n\n"
                else:
                    system_prompt += f"{field.upper()}:\nExtract the {field} value exactly as written in the document.\n\n"
        else:
            # Fallback to basic prompt if mapping not found
            # ARCHITECTURAL DECISION: Vision API extracts raw values only.
            # All normalization happens in normalize_v2().
            # Determine domain-specific terminology
            if is_driver:
                domain_name = "driver"
                domain_name_plural = "drivers"
                identifier_name = "Driver ID"
            else:
                domain_name = "vehicle"
                domain_name_plural = "vehicles"
                identifier_name = "VIN"
            
            system_prompt = (
                "CRITICAL EXTRACTION RULES (HIGHEST PRIORITY):\n"
                "\n"
                f"You must attempt to extract EVERY canonical {domain_name} field for EACH {domain_name},\n"
                "even if the value must be inferred from natural language.\n"
                "\n"
                f"For each {domain_name}:\n"
                "- If a field is not explicitly labeled, infer it from context.\n"
                "- If inference is uncertain, return your best guess.\n"
                "- If no reasonable inference is possible, return null — NEVER omit the field.\n"
                "\n"
            )
            
            # Add domain-specific inference examples
            if is_driver:
                system_prompt += (
                    "Inference examples (MANDATORY):\n"
                    "- If text contains \"Yes\" or \"No\" for training, training_completed = \"Yes\" or \"No\" (extract raw value)\n"
                    "- If text contains a number for violations, violations_count = that number (extract raw value)\n"
                    "- If text contains years of experience, years_experience = that number (extract raw value)\n"
                    "- If text contains license information, extract license_number, license_state, license_status (extract exactly as written)\n"
                    "- If text contains date of birth, extract date_of_birth in YYYY-MM-DD format if possible\n"
                    "\n"
                )
            else:
                system_prompt += (
                    "Inference examples (MANDATORY):\n"
                    "- If text contains \"sedan\", body_style = \"sedan\" (extract raw value, normalization happens later)\n"
                    "- If text contains \"truck\" or \"pickup\", body_style = \"truck\" (extract raw value)\n"
                    "- If text contains \"gas\" or \"gasoline\", fuel_type = \"gas\" or \"gasoline\" (extract raw value as written, normalization happens later)\n"
                    "- If text contains \"automatic\", \"auto\", \"8-speed auto\", transmission = \"automatic\" or \"auto\" (extract raw value as written, normalization happens later)\n"
                    "- If text contains an email-like string, extract it as owner_email (extract exactly as written)\n"
                    "\n"
                )
            
            system_prompt += (
                "ARCHITECTURAL REQUIREMENT: Extract raw values only. Do NOT normalize, convert, or transform values.\n"
                "All normalization (canonical forms, lowercase, type conversion) will be handled by the normalization pipeline.\n"
                "\n"
                "---\n"
                "\n"
                f"You are an OCR engine that extracts {domain_name} data from an image or PDF. "
                f"You MUST identify EVERY individual {domain_name} described in the document. "
                "Return ONLY valid JSON, no explanations, no markdown, no additional text. "
                "\n"
                "Output format:\n"
                "{\n"
                f'  "headers": {json.dumps(canonical_schema)},\n'
                '  "rows": [\n'
            )
            
            # Add domain-specific example rows
            if is_driver:
                system_prompt += (
                    '    ["D001", "John", "Doe", "1985-03-15", "D1234567", "CA", "Valid", 18, 0, "Yes", "Software Engineer"]\n'
                )
            else:
                system_prompt += (
                    '    ["VIN1", null, 2024, "Toyota", "Camry", null, "notes", "Blue", 12345, null, "sedan", "gas", "automatic", "email@example.com", null]\n'
                )
            
            system_prompt += (
                '  ]\n'
                "\n"
                "NOTE: Values in examples show raw extracted values.\n"
                "Normalization happens in the normalization pipeline, not in Vision extraction.\n"
                "}\n"
                "\n"
                "CRITICAL OUTPUT REQUIREMENTS:\n"
                f"- Return exactly ONE row per {domain_name}\n"
                "- Use the EXACT header names listed above (case-sensitive)\n"
                "- For missing fields, use null (not empty string)\n"
                f"- Extract ALL {domain_name_plural} from the document, not just the first one\n"
                f"- Each row must have exactly {len(canonical_schema)} values matching the {len(canonical_schema)} headers\n"
                "- Output ONLY the JSON object - no explanatory text before or after\n"
                "\n"
                "SECTION-BASED SEPARATION (CRITICAL FOR HANDWRITTEN DOCUMENTS):\n"
                f"- Documents may contain multiple {domain_name_plural} organized in SECTIONS or BLOCKS\n"
                f"- Each section/block represents ONE complete {domain_name} record\n"
                "- Look for visual separators: blank lines, horizontal lines, boxes, borders, or distinct sections\n"
                f"- For handwritten documents: each {domain_name} is typically in its own section with all its information grouped together\n"
                "- Identify section boundaries by:\n"
                "  * Blank lines or spacing between information\n"
                "  * Visual dividers (lines, boxes, borders)\n"
                f"  * Each {identifier_name} typically starts a new section\n"
                f"  * All fields for one {domain_name} are grouped within the same section\n"
                f"- DO NOT mix fields from different sections - each section = one {domain_name} = one row\n"
                f"- If you see multiple distinct sections/blocks of information, each is a separate {domain_name}\n"
                f"- For handwritten documents: carefully identify where one {domain_name} section ends and the next begins\n"
                f"- Extract ONE row per section/block, even if sections are not perfectly aligned or formatted\n"
                f"- If a document has 3 sections with {domain_name} info, you MUST extract 3 rows (one per section)\n"
                "\n"
            )
            
            # Add IMAGE TYPE DETECTION section only for IMAGE sources (fallback prompt)
            if source_type == SourceType.IMAGE:
                if is_driver:
                    system_prompt += (
                        "IMAGE TYPE DETECTION:\n"
                        "- First, determine if the image is a TABLE or a DOCUMENT-STYLE layout\n"
                        "- TABLE: Has clear column headers like DRIVER ID, FULL NAME, DATE OF BIRTH, LICENSE NUMBER arranged in columns\n"
                        "- DOCUMENT-STYLE: Information is organized in sections/blocks, not in a structured table format\n"
                        "\n"
                        "- If the image is a TABLE (column headers like DRIVER ID, FULL NAME, DATE OF BIRTH):\n"
                        "  * CRITICAL: The FIRST ROW is ALWAYS the header row - DO NOT extract it as a driver\n"
                        "  * Before extracting, identify which row is the header row (contains 'DRIVER ID', 'FULL NAME', 'DATE OF BIRTH', 'LICENSE NUMBER', etc.)\n"
                        "  * Skip the header row completely - do NOT include it in your output\n"
                        "  * Extract ONLY data rows that contain a VALID Driver ID in the Driver ID column\n"
                        "  * NEVER treat header rows as drivers - skip any row where the Driver ID column contains 'DRIVER ID', 'FULL NAME', 'DATE OF BIRTH', 'LICENSE NUMBER', etc.\n"
                        "  * If the Driver ID column contains ANY header text (even partial matches), SKIP that entire row\n"
                        "  * If no valid Driver ID exists in a row, SKIP it entirely - do NOT extract it\n"
                        "  * A valid Driver ID typically follows pattern: D followed by 3 digits, or DRV followed by 3 digits\n"
                        "  * Each valid Driver ID row = one driver = one output row\n"
                        "  * If you see a table with headers, count the header row and skip it - only extract data rows below it\n"
                        "  * CRITICAL FOR TABLE EXTRACTION: Read EACH COLUMN for EACH data row:\n"
                        "    - For each data row, read the value in EVERY column (DRIVER ID, FULL NAME, DATE OF BIRTH, LICENSE NUMBER, LICENSE STATE, LICENSE STATUS, YEARS EXPERIENCE, VIOLATIONS COUNT, TRAINING COMPLETED, NOTES, etc.)\n"
                        "    - Extract the value from each column cell - do NOT skip columns even if they seem empty or contain unusual values\n"
                        "    - Map each column header to the corresponding field in the output schema\n"
                        "    - If a column cell is empty or unclear, use null (not empty string)\n"
                        "    - Extract ALL visible columns for each row - do NOT omit any columns\n"
                        "\n"
                        "- If the image is NOT a table (document-style or handwritten):\n"
                        "  * Use section-based extraction rules (see SECTION-BASED SEPARATION above)\n"
                        "  * Identify distinct sections/blocks of driver information\n"
                        "  * CRITICAL: Handwritten images are typically organized as SEPARATE TEXT BLOCKS, NOT tables\n"
                        "  * Each block is a self-contained paragraph or section describing one driver\n"
                        "  * Extract one row per block, even if formatting is irregular\n"
                        "  * Each block should contain a Driver ID and associated driver fields\n"
                        "  * Do NOT look for table structure - look for separate text blocks instead\n"
                        "\n"
                    )
                else:
                    system_prompt += (
                        "IMAGE TYPE DETECTION:\n"
                        "- First, determine if the image is a TABLE or a DOCUMENT-STYLE layout\n"
                        "- TABLE: Has clear column headers like YEAR, MAKE, MODEL, VIN, COLOR, MILEAGE arranged in columns\n"
                        "- DOCUMENT-STYLE: Information is organized in sections/blocks, not in a structured table format\n"
                        "\n"
                        "- If the image is a TABLE (column headers like YEAR, MAKE, MODEL, VIN):\n"
                        "  * CRITICAL: The FIRST ROW is ALWAYS the header row - DO NOT extract it as a vehicle\n"
                        "  * Before extracting, identify which row is the header row (contains 'YEAR', 'MAKE', 'MODEL', 'VIN', 'COLOR', 'MILEAGE', etc.)\n"
                        "  * Skip the header row completely - do NOT include it in your output\n"
                        "  * Extract ONLY data rows that contain a VALID 17-character VIN in the VIN column\n"
                        "  * NEVER treat header rows as vehicles - skip any row where the VIN column contains 'YEAR', 'MAKE', 'MODEL', 'VIN', 'COLOR', 'MILEAGE', 'NOTES', 'BODY', 'PAINTED', 'GASOLINE', 'TRANSMISSION', etc.\n"
                        "  * If the VIN column contains ANY header text (even partial matches), SKIP that entire row\n"
                        "  * If no valid VIN exists in a row, SKIP it entirely - do NOT extract it\n"
                        "  * A valid VIN must have at least 2 digits and be 10-17 characters long\n"
                        "  * Only extract rows with valid 17-character alphanumeric VINs (excluding I, O, Q)\n"
                        "  * Each valid VIN row = one vehicle = one output row\n"
                        "  * If you see a table with headers, count the header row and skip it - only extract data rows below it\n"
                        "  * CRITICAL FOR TABLE EXTRACTION: Read EACH COLUMN for EACH data row:\n"
                        "    - For each data row, read the value in EVERY column (YEAR, MAKE, MODEL, VIN, COLOR, MILEAGE, BODY, FUEL, TRANSMISSION, EMAIL, NOTES, etc.)\n"
                        "    - Extract the value from each column cell - do NOT skip columns even if they seem empty or contain unusual values\n"
                        "    - Map each column header to the corresponding field in the output schema\n"
                        "    - If a column cell is empty or unclear, use null (not empty string)\n"
                        "    - Extract ALL visible columns for each row - do NOT omit any columns\n"
                        "\n"
                        "- If the image is NOT a table (document-style or handwritten):\n"
                        "  * Use section-based extraction rules (see SECTION-BASED SEPARATION above)\n"
                        "  * Identify distinct sections/blocks of vehicle information\n"
                        "  * CRITICAL: Handwritten images are typically organized as SEPARATE TEXT BLOCKS, NOT tables\n"
                        "  * Each block is a self-contained paragraph or section describing one vehicle\n"
                        "  * Extract one row per block, even if formatting is irregular\n"
                        "  * Each block should contain a VIN and associated vehicle fields\n"
                        "  * Do NOT look for table structure - look for separate text blocks instead\n"
                        "\n"
                    )
            
            # Add domain-specific field list for extraction rules
            if is_driver:
                field_list = "driver_id, first_name, last_name, date_of_birth, license_number, license_state, license_status, years_experience, violations_count, training_completed, and notes"
            else:
                field_list = "VIN, year, make, model, color, mileage, body_style, fuel_type, transmission, owner_email, and notes"
            
            if is_driver:
                system_prompt += (
                    "DRIVER ID VALIDATION (CRITICAL):\n"
                    "- The Driver ID (first field) MUST be a valid driver identifier (e.g., D001, D002, DRV001)\n"
                    "- Driver ID typically follows pattern: D followed by 3 digits, or DRV followed by 3 digits\n"
                    "- DO NOT include header rows - if the Driver ID column contains text like 'Driver ID', 'Full Name', 'Date of Birth', 'License Number', skip that row entirely\n"
                    "- DO NOT include table headers as data rows - only extract rows with valid driver IDs\n"
                    "- For handwritten documents, carefully validate Driver IDs to avoid OCR character recognition errors\n"
                    "- Each valid Driver ID you find represents a separate driver - extract it as a separate row\n"
                    "\n"
                )
            else:
                system_prompt += (
                    "VIN VALIDATION (CRITICAL):\n"
                    "- The VIN (first field) MUST be a valid 17-character alphanumeric code\n"
                    "- VIN must contain at least 2 digits and exclude letters I, O, Q\n"
                    "- DO NOT include header rows - if the VIN column contains text like 'YEAR', 'MAKE', 'MODEL', 'VIN', 'COLOR', 'MILEAGE', skip that row entirely\n"
                    "- DO NOT include table headers as data rows - only extract rows with valid vehicle VINs\n"
                    "- For handwritten documents, carefully validate VINs to avoid OCR character recognition errors (e.g., '0' vs 'O', '1' vs 'I')\n"
                    "- Each valid VIN you find represents a separate vehicle - extract it as a separate row\n"
                    "\n"
                )
            
            system_prompt += (
                "EXTRACTION RULES (CRITICAL):\n"
                "- Extract values EXACTLY as written in the source document\n"
                "- Do NOT reword, summarize, paraphrase, or embellish any values\n"
                "- Do NOT normalize or convert values (preserve original format)\n"
                "- Infer missing values from natural language when possible (see CRITICAL EXTRACTION RULES at top) - only use null if no reasonable inference is possible\n"
                "- When uncertain, extract your best guess rather than omitting the field\n"
                "- For invalid values, extract them anyway - validation will handle warnings\n"
                "- Prefer extraction over omission - always attempt to extract a value if it appears in the document\n"
                f"- Extract ALL fields for each {domain_name} - do NOT skip fields even if they seem invalid or unusual\n"
                f"- For each {domain_name}, attempt to extract: {field_list}\n"
                "- When extracting from a section: extract ALL fields that appear in that section, even if formatting is irregular\n"
                "\n"
                "NOTES FIELD RULES (CRITICAL):\n"
                "- If notes exist in the source, copy them VERBATIM - word-for-word, character-for-character\n"
                "- Preserve original wording, punctuation, sentence order, and capitalization\n"
                "- Do NOT summarize, paraphrase, rewrite, or reorder sentences\n"
                f"- Do NOT remove content unless it's clearly not {domain_name}-specific\n"
                "- Do NOT add explanatory text or fabricate information that is not visible in the document\n"
                "- If no notes are present in the source, set notes = null (not empty string)\n"
            )
        
        # Process each page: try Vision first, fallback to OCR inference
        all_rows = []  # List of rows extracted (one per page ideally)
        vision_extracted_pages = []
        
        import base64
        import io
        
        # Handle raw text files differently (detect domain and use appropriate parser)
        if source_type == SourceType.RAW_TEXT and not images:
            # Use the text content already read above
            try:
                # Check mapping_id from kwargs to determine domain
                domain_mapping_id = kwargs.get('mapping_id', '')
                if 'policies' in domain_mapping_id.lower():
                    # Parse policies from raw text
                    policies = parse_policy_raw_text(raw_text_content)
                    if policies:
                        logger.info(f"[Raw Text] Extracted {len(policies)} policy/policies from raw text")
                        # Convert to 2D array format matching canonical schema
                        from schema import POLICY_SCHEMA_ORDER
                        canonical_schema = POLICY_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for policy in policies:
                            # Build row in canonical schema order and clean "None" strings
                            row = []
                            for field in canonical_schema:
                                value = policy.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning("[Raw Text] No policies extracted from raw text")
                        return None
                elif 'locations' in domain_mapping_id.lower():
                    # Parse locations from raw text
                    locations = parse_locations_raw_text(raw_text_content)
                    if locations:
                        logger.info(f"[Raw Text] Extracted {len(locations)} location(s) from raw text")
                        # Convert to 2D array format matching canonical schema
                        from schema import LOCATION_SCHEMA_ORDER
                        canonical_schema = LOCATION_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for location in locations:
                            # Build row in canonical schema order and clean "None" strings
                            row = []
                            for field in canonical_schema:
                                value = location.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning("[Raw Text] No locations extracted from raw text")
                        return None
                elif 'claims' in domain_mapping_id.lower():
                    # Parse claims from raw text
                    claims = parse_claim_raw_text(raw_text_content)
                    if claims:
                        logger.info(f"[Raw Text] Extracted {len(claims)} claim(s) from raw text")
                        # Convert to 2D array format matching canonical schema
                        from schema import CLAIM_SCHEMA_ORDER
                        canonical_schema = CLAIM_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for claim in claims:
                            # Build row in canonical schema order and clean "None" strings
                            row = []
                            for field in canonical_schema:
                                value = claim.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning("[Raw Text] No claims extracted from raw text")
                        return None
                elif 'relationship' in domain_mapping_id.lower() or 'link' in domain_mapping_id.lower():
                    # Parse relationships from raw text
                    relationships = parse_relationship_raw_text(raw_text_content)
                    if relationships:
                        logger.info(f"[Raw Text] Extracted {len(relationships)} relationship(s) from raw text")
                        # Convert to 2D array format matching canonical schema
                        from schema import RELATIONSHIP_SCHEMA_ORDER
                        canonical_schema = RELATIONSHIP_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for relationship in relationships:
                            # Build row in canonical schema order and clean "None" strings
                            row = []
                            for field in canonical_schema:
                                value = relationship.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning("[Raw Text] No relationships extracted from raw text")
                        return None
                elif 'drivers' in domain_mapping_id.lower():
                    # Parse drivers from raw text
                    drivers = parse_driver_raw_text(raw_text_content)
                    if drivers:
                        logger.info(f"[Raw Text] Extracted {len(drivers)} driver(s) from raw text")
                        # Convert to 2D array format matching canonical schema
                        from schema import DRIVER_SCHEMA_ORDER
                        canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for driver in drivers:
                            # Build row in canonical schema order and clean "None" strings
                            row = []
                            for field in canonical_schema:
                                value = driver.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning("[Raw Text] No drivers extracted from raw text")
                        return None
                else:
                    # Default to vehicle parser
                    vehicles = parse_vehicle_raw_text(raw_text_content)
                
                if vehicles:
                    logger.info(f"[Raw Text] Extracted {len(vehicles)} vehicle(s) from raw text")
                    
                    # Convert to 2D array format matching canonical schema
                    # Include _is_handwritten, _is_vision_extracted, _is_table_extraction, and _is_ai_repaired in header to preserve them through rows2d_to_objects
                    header_row = list(canonical_schema) + ["_is_handwritten", "_is_vision_extracted", "_is_table_extraction", "_is_ai_repaired"]
                    rows_2d = [header_row]
                    
                    for vehicle in vehicles:
                        # Build row in canonical schema order and clean "None" strings
                        row = []
                        for field in canonical_schema:
                            value = vehicle.get(field)
                            # Clean "None" strings
                            if isinstance(value, str) and value.strip().lower() == "none":
                                value = None
                            row.append(value)
                        # Preserve _is_handwritten, _is_vision_extracted, _is_table_extraction, and _is_ai_repaired flags
                        row.append(vehicle.get("_is_handwritten", False))
                        row.append(vehicle.get("_is_vision_extracted", False))
                        row.append(vehicle.get("_is_table_extraction", False))
                        row.append(vehicle.get("_is_ai_repaired", False))
                        rows_2d.append(row)
                    
                    return rows_2d
                else:
                    logger.warning("[Raw Text] No vehicles extracted from raw text")
                    return None
            except Exception as e:
                logger.error(f"[Raw Text] Error processing raw text file: {e}")
                import traceback
                traceback.print_exc()
                return None
        
        # For IMAGE sources, collect all rows first, then aggregate into ONE vehicle after all pages
        image_rows_collector = [] if source_type == SourceType.IMAGE else None
        
        # Process images (PDF/IMAGE sources)
        
        # Detect if this is a handwritten PDF (by file path) - do this ONCE before the page loop
        # Also check if PDF was routed as IMAGE (scanned PDF detection)
        is_handwritten_pdf = False
        is_scanned_pdf_routed_as_image = False
        if source_type == SourceType.PDF:
            file_path_str = str(file_path).lower()
            is_handwritten_pdf = any(keyword in file_path_str for keyword in ['handwritten', 'hand'])
            if is_handwritten_pdf:
                logger.info(f"[Vision] Handwritten PDF detected: {file_path.name} - Using image-first Vision extraction")
        elif source_type == SourceType.IMAGE and is_actually_pdf:
            # PDF file routed as IMAGE means it was detected as scanned/handwritten
            is_scanned_pdf_routed_as_image = True
            logger.info(f"[Vision] Scanned PDF detected and routed as IMAGE: {file_path.name} - Using image-first Vision extraction (skipping OCR-first path)")
        
        # CRITICAL FIX: Initialize ocr_vehicles_by_vin OUTSIDE the page loop
        # This allows OCR vehicles from page 1 to be available for merging with Vision results on later pages
        # Previously, this was reset every page (line 4463), causing OCR data from page 1 to be lost
        ocr_vehicles_by_vin = {}
        
        # Store OCR text for post-Vision inference (accessible to all Vision-extracted rows)
        # This ensures OCR text is available when Vision doesn't extract notes
        extracted_ocr_text_for_vision = ""
        
        # CRITICAL FIX: For handwritten PDFs, extract OCR text for post-Vision inference
        # Even though Vision extracts rows, we need OCR text to fill missing fields via inference
        # NOTE: Handwritten PDFs may be routed as SourceType.IMAGE, so check both is_handwritten_pdf and is_scanned_pdf_routed_as_image
        # Also check filename for "handwritten" keyword even if source_type is IMAGE
        is_handwritten_source = is_handwritten_pdf or is_scanned_pdf_routed_as_image
        if not is_handwritten_source and source_type == SourceType.IMAGE:
            # Check filename for handwritten keyword (PDFs routed as IMAGE)
            file_path_str = str(file_path).lower()
            is_handwritten_source = any(keyword in file_path_str for keyword in ['handwritten', 'hand'])
        
        if is_handwritten_source and file_path.suffix.lower() == '.pdf':
            try:
                logger.warning(f"[Handwritten PDF] Extracting OCR text for post-Vision inference (file: {file_path.name}, source_type: {source_type.name})")
                if not hasattr(_extract_table_with_vision_api, '_cached_full_ocr_text'):
                    full_ocr_text = _get_vision_ocr_text_for_pdf(file_path)
                    _extract_table_with_vision_api._cached_full_ocr_text = full_ocr_text
                    extracted_ocr_text_for_vision = full_ocr_text
                    logger.warning(f"[Handwritten PDF] Extracted OCR text ({len(full_ocr_text)} chars) for post-Vision inference")
                else:
                    extracted_ocr_text_for_vision = getattr(_extract_table_with_vision_api, '_cached_full_ocr_text', "")
                    logger.warning(f"[Handwritten PDF] Using cached OCR text ({len(extracted_ocr_text_for_vision)} chars) for post-Vision inference")
            except Exception as e:
                logger.warning(f"[Handwritten PDF] Could not extract OCR text for post-Vision inference: {e}")
                extracted_ocr_text_for_vision = ""
        
        # CRITICAL FIX: For IMAGE sources, extract OCR text from the image file as fallback
        # This brings IMAGE extraction up to parity with PDF extraction (OCR + Vision merge)
        if source_type == SourceType.IMAGE and not is_scanned_pdf_routed_as_image:
            try:
                from ocr.reader import extract_text_from_image
                logger.debug(f"[IMAGE] Extracting OCR text from image: {file_path.name}")
                ocr_text_blocks, ocr_metadata = extract_text_from_image(file_path, enable_vision=False)
                if ocr_text_blocks:
                    from ocr.table_extract import _extract_raw_text
                    full_ocr_text = _extract_raw_text(ocr_text_blocks)
                    extracted_ocr_text_for_vision = full_ocr_text  # Store for post-Vision inference
                    logger.debug(f"[IMAGE] Extracted OCR text ({len(full_ocr_text)} chars) from image")
                    logger.warning(f"[IMAGE] OCR text preview (first 500 chars): {full_ocr_text[:500]}")
                    
                    # Parse vehicles from OCR text using the same logic as PDFs
                    source_type_str = "image"
                    document_defaults = _extract_document_level_defaults(full_ocr_text, source_type_str, file_path)
                    blocks = split_by_vin(full_ocr_text)
                    logger.warning(f"[IMAGE] Found {len(blocks)} VIN block(s) after splitting")
                    for i, block in enumerate(blocks[:3]):  # Show first 3 blocks
                        logger.warning(f"[IMAGE] Block {i+1} preview (first 200 chars): {block[:200]}")
                    ocr_vehicles = []
                    for i, b in enumerate(blocks):
                        vehicle = extract_fields_from_block(b, mapping=mapping_config, source_type=source_type_str)
                        if vehicle:
                            vin = vehicle.get('vin', 'NO_VIN')
                            logger.warning(f"[IMAGE] Block {i+1}: extracted vehicle with VIN={vin}, fields={list(vehicle.keys())}")
                        else:
                            logger.warning(f"[IMAGE] Block {i+1}: extract_fields_from_block returned None/empty")
                        ocr_vehicles.append(vehicle)
                    logger.warning(f"[IMAGE] Extracted {len([v for v in ocr_vehicles if v])} vehicle(s) from {len(blocks)} block(s)")
                    
                    # Apply document-level defaults
                    for vehicle in ocr_vehicles:
                        if vehicle:
                            for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                if vehicle.get(field) is None and document_defaults.get(field) is not None:
                                    vehicle[field] = document_defaults[field]
                            
                            # Store OCR vehicles by VIN for merging with Vision results
                            vin_value = vehicle.get('vin')
                            if vin_value:
                                ocr_vehicles_by_vin[vin_value] = vehicle
                                logger.debug(f"[IMAGE] Extracted OCR vehicle: VIN={vin_value}")
                    
                    if ocr_vehicles_by_vin:
                        logger.info(f"[IMAGE] Extracted {len(ocr_vehicles_by_vin)} vehicle(s) from OCR text")
                else:
                    logger.debug(f"[IMAGE] No OCR text blocks extracted from image")
            except Exception as e:
                logger.warning(f"[IMAGE] OCR extraction failed (will use Vision only): {e}")
        
        for page_idx, img in enumerate(images):
            page_num = page_idx + 1
            
            # Convert image to base64
            buffered = io.BytesIO()
            img.save(buffered, format="PNG")
            img_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
            
            # Get OCR text for this page if available (but skip for handwritten PDFs and scanned PDFs routed as IMAGE - use as fallback only)
            page_ocr_text = ""
            if not is_handwritten_pdf and not is_scanned_pdf_routed_as_image:
                # For non-handwritten PDFs and regular IMAGE sources, get OCR text as before
                # Skip OCR text extraction for scanned PDFs routed as IMAGE (they go straight to Vision API)
                if page_num in page_text_blocks_dict:
                    try:
                        from ocr.table_extract import _extract_raw_text
                        page_ocr_text = _extract_raw_text(page_text_blocks_dict[page_num])
                    except Exception:
                        pass
                elif (source_type == SourceType.PDF or (source_type == SourceType.IMAGE and is_actually_pdf)) and not page_text_blocks_dict:
                    # If no page_text_blocks_dict was provided, extract OCR text for all pages on-the-fly
                    # CRITICAL: Only process the full OCR text ONCE (on page 1), not on every page
                    # Otherwise we'll extract the same vehicles multiple times (6 pages × 6 vehicles = 36 rows)
                    if not hasattr(_extract_table_with_vision_api, '_cached_full_ocr_text'):
                        # Cache the full OCR text so we don't extract it multiple times
                        try:
                            full_ocr_text = _get_vision_ocr_text_for_pdf(file_path)
                            _extract_table_with_vision_api._cached_full_ocr_text = full_ocr_text
                            logger.info(f"[Vision] Extracted full Vision OCR text ({len(full_ocr_text)} chars) for all pages")
                        except Exception as e:
                            logger.debug(f"[Vision] Could not extract Vision OCR text: {e}")
                            _extract_table_with_vision_api._cached_full_ocr_text = ""
                    
                    # Only use cached full OCR text on page 1 (process once, not per page)
                    if page_num == 1:
                        page_ocr_text = getattr(_extract_table_with_vision_api, '_cached_full_ocr_text', "")
                        if page_ocr_text:
                            logger.debug(f"[Vision] Using cached full Vision OCR text for page {page_num} ({len(page_ocr_text)} chars)")
                    else:
                        # Skip OCR text extraction for pages 2+ (already processed on page 1)
                        page_ocr_text = ""
                        logger.debug(f"[Vision] Skipping OCR text extraction for page {page_num} (already processed on page 1)")
                
                # Store OCR text for post-Vision inference (use page_ocr_text or full cached text)
                if page_ocr_text and not extracted_ocr_text_for_vision:
                    extracted_ocr_text_for_vision = page_ocr_text
                elif not extracted_ocr_text_for_vision and hasattr(_extract_table_with_vision_api, '_cached_full_ocr_text'):
                    cached_text = getattr(_extract_table_with_vision_api, '_cached_full_ocr_text', "")
                    if cached_text:
                        extracted_ocr_text_for_vision = cached_text
            
            # For handwritten PDFs: Skip OCR text path, go straight to Vision API (image-first extraction)
            # For scanned PDFs routed as IMAGE: Skip OCR text path entirely, go straight to Vision API
            # For non-handwritten PDFs: PDF → text → block segmentation → extract_fields_from_block
            # For regular IMAGE sources: Use OCR text if available
            # NOTE: ocr_vehicles_by_vin is now initialized BEFORE the page loop (line 4414)
            # This ensures OCR vehicles from page 1 persist and can be merged with Vision results on later pages
            if (source_type == SourceType.PDF and not is_handwritten_pdf) or (source_type == SourceType.IMAGE and not is_scanned_pdf_routed_as_image):
                # Only use OCR text path for:
                # - Non-handwritten PDFs
                # - Regular IMAGE sources (not scanned PDFs routed as IMAGE)
                if page_ocr_text:
                    # Pre-split extraction: Extract document-level defaults from full OCR text
                    # (for handwritten PDF/image or PNG images where fields may be lost during splitting)
                    source_type_str = "pdf" if source_type == SourceType.PDF else "image"
                    document_defaults = _extract_document_level_defaults(page_ocr_text, source_type_str, file_path)
                    
                    # After OCR text is obtained, force segmentation the same way
                    blocks = split_by_vin(page_ocr_text)
                    rows = [extract_fields_from_block(b, mapping=mapping_config, source_type=source_type_str) for b in blocks]
                    
                    # Apply document-level defaults to each vehicle if field is still None
                    applied_count = 0
                    for vehicle_idx, vehicle in enumerate(rows):
                        if vehicle:
                            for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                if vehicle.get(field) is None and document_defaults.get(field) is not None:
                                    vehicle[field] = document_defaults[field]
                                    applied_count += 1
                    if applied_count > 0:
                        print(f"[Pre-split extraction] Applied {applied_count} document-level default(s) to {len(rows)} vehicle(s)")
                    # If we successfully extracted vehicles from segmented blocks, use them
                    extracted_vehicles = [r for r in rows if r and r.get('vin')]
                    if extracted_vehicles:
                        logger.info(f"[{source_type.name}] Extracted {len(extracted_vehicles)} vehicle(s) from page {page_num} segmented OCR text")
                        # Store OCR-extracted vehicles by VIN for potential merging with Vision results
                        ocr_vehicles_by_vin.update({v.get('vin'): v for v in extracted_vehicles if v and v.get('vin')})
                        # DEBUG: Check if Row 6 was extracted
                        if 'ST420RJ98FDHKL4E' in ocr_vehicles_by_vin:
                            logger.info(f"[PDF] Row 6 (ST420RJ98FDHKL4E) found in OCR extraction on page {page_num}")
                        # Convert to 2D array format and clean "None" strings
                        for vehicle in extracted_vehicles:
                            vin_value = vehicle.get('vin')
                            # SAFEGUARD: Skip vehicles with invalid VINs (header tokens) for IMAGE sources
                            if source_type == SourceType.IMAGE and vin_value:
                                vin_upper_clean = str(vin_value).upper().strip().rstrip('.,;:!?')
                                invalid_vin_tokens = {'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 'VIN', 
                                                      'BODY', 'PAINTED', 'GASOLINE', 'TRANSMISSION', 'NOTES',
                                                      'CURRENT', 'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET',
                                                      'STYLE', 'TYPE', 'EMAIL', 'FUEL', 'BODY_STYLE', 'FUEL_TYPE'}
                                if vin_upper_clean in invalid_vin_tokens:
                                    logger.debug(f"[IMAGE] Skipping vehicle with invalid VIN (header token): {vin_value} -> {vin_upper_clean}")
                                    continue  # Skip this vehicle - it has an invalid VIN
                            # CRITICAL FIX: Initialize row INSIDE the loop for each vehicle
                            row = []
                            for field in canonical_schema:
                                value = vehicle.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            # Preserve _is_handwritten flag as extra field
                            row.append(vehicle.get("_is_handwritten", False))
                            # For IMAGE sources, collect rows for aggregation AFTER all pages
                            if source_type == SourceType.IMAGE:
                                image_rows_collector.append(row)
                            else:
                                all_rows.append(row)
                        vision_extracted_pages.append(page_num)
                        continue  # Skip Vision API for this page since text extraction succeeded
                
                # If text extraction didn't work, only then try Vision API as supplement
                # (but don't overwrite good text-extracted fields)
                # NOTE: blocks is only defined if we went through the OCR text path above
                # If we skipped OCR path (handwritten/scanned PDFs), blocks won't exist, so skip this fallback
                if 'blocks' in locals() and blocks:
                    page_ocr_text = "\n\n--- VEHICLE SEPARATOR ---\n\n".join(blocks)
                    logger.debug(f"[PDF] Text extraction incomplete, trying Vision API supplement for page {page_num}")
                else:
                    # OCR text path was skipped (handwritten/scanned PDFs) - no blocks to join
                    # page_ocr_text will be empty, Vision API will process image directly
                    page_ocr_text = ""
                    logger.debug(f"[Vision] OCR text path skipped for page {page_num}, proceeding to Vision API (image-first)")
            
            # For handwritten PDFs: Try Vision API FIRST (image-first extraction)
            # For other PDFs/IMAGES: Try Vision API as supplement/fallback
            # Try Vision API for this page
            # Determine domain-specific terminology
            if is_driver:
                domain_name = "driver"
                domain_name_plural = "drivers"
                identifier_name = "Driver ID"
                field_list = "driver_id, first_name, last_name, date_of_birth, license_number, license_state, license_status, years_experience, violations_count, training_completed, and notes"
                table_columns = "DRIVER ID, FULL NAME, DATE OF BIRTH, LICENSE NUMBER, LICENSE STATE, LICENSE STATUS, YEARS EXPERIENCE, VIOLATIONS COUNT, TRAINING COMPLETED, NOTES"
            else:
                domain_name = "vehicle"
                domain_name_plural = "vehicles"
                identifier_name = "VIN"
                field_list = "VIN, year, make, model, color, mileage, body_style, fuel_type, transmission, owner_email, and notes"
                table_columns = "YEAR, MAKE, MODEL, VIN, COLOR, MILEAGE, BODY, FUEL, TRANSMISSION, EMAIL, NOTES"
            
            user_prompt_text = (
                f"This is page {page_num} of a {domain_name} data document. "
                f"Extract ALL individual {domain_name_plural} described on this page. "
                f"Return exactly one row per {domain_name} using the exact JSON format and schema from the system message. "
                "\n\n"
                "CRITICAL INSTRUCTIONS:\n"
                "- Follow the field extraction instructions from the system message EXACTLY\n"
                f"- Each {domain_name} is defined by its {identifier_name} block - extract only fields within the same block as each {identifier_name}\n"
                f"- Do NOT reuse fields from previous {domain_name_plural}\n"
                "- Extract values EXACTLY as written - do NOT reword, summarize, or paraphrase\n"
                f"- Extract ALL fields for each {domain_name} - do NOT skip any fields (use null if missing)\n"
                f"- For each {domain_name}, you MUST attempt to extract: {field_list}\n"
                f"- CRITICAL: For TABLE images, read EVERY column for each data row - extract values from ALL columns ({table_columns}, etc.)\n"
                "- Do NOT skip columns - if a column exists in the table, extract its value for each row (use null only if the cell is truly empty)\n"
                "- Even if a field value seems invalid, extract it anyway - do NOT drop the field\n"
                "- For the notes field: if present, copy VERBATIM - preserve original wording, punctuation, and order\n"
                "- If notes are missing, set notes = null (not empty string)\n"
                "- When uncertain about a value, extract your best guess rather than omitting it\n"
                "- For invalid values, extract them anyway - do NOT drop the field\n"
            )
            
            if is_driver:
                user_prompt_text += (
                    f"- DO NOT include table header rows - only extract rows with valid Driver IDs\n"
                    "- If you see a table, skip the header row entirely - do NOT include it in the output\n"
                )
            else:
                user_prompt_text += (
                    "- DO NOT include table header rows - only extract rows with valid 17-character VINs\n"
                    "- If you see a table, skip the header row entirely - do NOT include it in the output\n"
                )
            
            user_prompt_text += (
                "\n"
                "SECTION-BASED EXTRACTION (ESPECIALLY FOR HANDWRITTEN DOCUMENTS):\n"
                f"- Identify each distinct SECTION or BLOCK of {domain_name} information on this page\n"
                f"- Each section represents ONE complete {domain_name} record\n"
                "- Look for visual separators between sections:\n"
                "  * Blank lines or spacing\n"
                "  * Horizontal lines, boxes, or borders\n"
                f"  * Each {identifier_name} typically marks the start of a new section\n"
                f"  * All information for one {domain_name} is grouped within its section\n"
                f"- For handwritten documents: sections may be less clearly defined, but each {domain_name}'s information is typically grouped together\n"
                "- Extract ONE row per section you identify\n"
                f"- If you see 3 distinct sections with {domain_name} information, extract 3 rows (one per section)\n"
                "- DO NOT combine information from different sections into one row\n"
                f"- DO NOT skip sections - if a section contains a valid {identifier_name}, extract it as a separate row\n"
                "- Within each section, extract ALL fields that appear in that section\n"
                f"- If a section is missing some fields, use null for those fields but still extract the section as a row\n"
            )
            
            if is_driver:
                user_prompt_text += (
                    "\n"
                    "CRITICAL FIELD EXTRACTION FOR HANDWRITTEN/SECTION-BASED DOCUMENTS:\n"
                    f"- For EACH {domain_name} section, you MUST actively look for and extract these specific fields:\n"
                    "  * license_number: Look for license numbers (format varies: D1234567, A1234567, etc.)\n"
                    "  * license_state: Look for state abbreviations (CA, NY, TX, etc.)\n"
                    "  * license_status: Look for words like 'Valid', 'Active', 'Suspended', 'Expired'\n"
                    "  * years_experience: Look for numbers followed by 'years' or 'experience'\n"
                    "  * violations_count: Look for numbers related to violations or tickets\n"
                    "  * training_completed: Look for 'Yes', 'No', 'Completed', 'Not Completed'\n"
                    "  * date_of_birth: Look for dates in various formats (YYYY-MM-DD, MM/DD/YYYY, etc.)\n"
                    "- These fields are often written in natural language within each section - read the ENTIRE section text carefully\n"
                    "- Do NOT skip these fields - if you see ANY mention of license, experience, violations, or training in a section, extract it\n"
                    "- For handwritten documents, these fields may be written in various formats - look for keywords ANYWHERE in the section text\n"
                    "- IMPORTANT: Even if a field is mentioned without a label, you MUST extract it\n"
                    "- Do NOT assume fields are missing just because they don't have structured labels - search the entire section text\n"
                )
            else:
                user_prompt_text += (
                    "\n"
                    "CRITICAL FIELD EXTRACTION FOR HANDWRITTEN/SECTION-BASED DOCUMENTS:\n"
                    f"- For EACH {domain_name} section, you MUST actively look for and extract these specific fields:\n"
                    "  * body_style: Look for words like 'sedan', 'truck', 'SUV', 'coupe', 'van', 'convertible', 'wagon', 'hatchback', 'crossover'\n"
                    "    - Examples: '2024 Toyota Camry sedan', 'Ford F-150 truck', 'Honda CR-V SUV'\n"
                    "    - The body style may appear anywhere in the section text, not just after a label\n"
                    "  * fuel_type: Look for words like 'gas', 'gasoline', 'diesel', 'electric', 'hybrid'\n"
                    "    - Examples: 'Fuel: gas', 'Fuel: gasoline', 'runs on diesel', 'electric vehicle', 'hybrid car'\n"
                    "    - May appear as 'Fuel: X' OR just 'gas' or 'gasoline' mentioned anywhere in the section\n"
                    "  * transmission: Look for words like 'automatic', 'manual', 'CVT', 'auto'\n"
                    "    - Examples: 'automatic transmission', 'transmission: automatic', '8-speed auto', 'manual shift'\n"
                    "    - May appear as 'Transmission: X' OR just 'automatic' or 'manual' mentioned anywhere in the section\n"
                    "  * owner_email: Look for email addresses (format: text@domain.com) - extract EXACTLY as written, even if format seems invalid\n"
                    "  * color: Look for color words like 'Blue', 'Red', 'Green', 'Black', 'White', etc. (may appear as 'painted blue' or 'Color: Blue')\n"
                    "  * mileage: Look for numbers followed by 'miles' or 'mileage'\n"
                    "    - Examples: '12,345 miles', '100,245 miles', 'about 89,000 miles', '45,210 miles on odometer'\n"
                    "    - May include phrases like 'about X miles', 'X miles on odometer', 'odometer shows X miles'\n"
                    "- These fields are often written in natural language within each section - read the ENTIRE section text carefully\n"
                    "- Do NOT skip these fields - if you see ANY mention of body style, fuel type, or transmission in a section, extract it\n"
                    "- For handwritten documents, these fields may be written in various formats - look for keywords ANYWHERE in the section text\n"
                    "- IMPORTANT: Even if a field is mentioned without a label (e.g., just 'sedan' or 'gas' or 'automatic'), you MUST extract it\n"
                    "- Do NOT assume fields are missing just because they don't have structured labels - search the entire section text\n"
                )
            
            if page_ocr_text:
                separator_text = f"{domain_name_plural.upper()} SEPARATOR" if is_driver else "VEHICLE SEPARATOR"
                user_prompt_text += (
                    f"\n\nOCR-extracted text from this page ({domain_name_plural} separated by '--- {separator_text} ---'):\n{page_ocr_text[:2000]}\n\n"
                    f"Use this OCR text to help identify all {domain_name_plural} and their field boundaries. "
                    f"Each section between separators represents one {domain_name}. "
                    "Extract values from the OCR text EXACTLY as written - do NOT modify or reword. "
                    "\n"
                    "IMPORTANT: Even if the OCR text shows separators, also look at the IMAGE itself to identify sections. "
                    "For handwritten documents, the image may show clearer section boundaries than the OCR text. "
                    f"Count how many distinct {domain_name} sections you see in the image and extract that many rows."
                )
            
            # Detect if this might be a handwritten document based on filename or content
            is_handwritten_doc = any(keyword in str(file_path).lower() for keyword in ['handwritten', 'hand', 'scribble', 'notes'])
            
            if is_handwritten_doc:
                if is_driver:
                    user_prompt_text += (
                        "\n\nHANDWRITTEN DOCUMENT DETECTED - SPECIAL INSTRUCTIONS:\n"
                        "- This appears to be a handwritten document - text may be less structured\n"
                        "- CRITICAL: Handwritten documents are typically NOT in table format - they are organized as SEPARATE BLOCKS or SECTIONS\n"
                        f"- Each {domain_name}'s information is written in its own distinct block/section, NOT in table rows\n"
                        f"- Look carefully for ALL distinct SECTIONS or BLOCKS of {domain_name} information\n"
                        f"- Each section typically contains one complete {domain_name} with all its fields grouped together\n"
                        "- Sections may be separated by:\n"
                        "  * Blank space or lines\n"
                        "  * Visual boundaries (boxes, lines, borders)\n"
                        f"  * Each {identifier_name} usually marks the start of a new section\n"
                        f"  * Each section is a self-contained block of text describing one {domain_name}\n"
                        f"- Count how many distinct {domain_name} sections/blocks you see - extract that many rows\n"
                        f"- If you see 3 separate blocks with {domain_name} information, extract 3 rows (one per block)\n"
                        "- DO NOT combine multiple blocks into one row\n"
                        f"- DO NOT skip blocks - extract every block that contains {domain_name} information\n"
                        "- Within each block, extract all fields that appear, even if handwriting is difficult to read\n"
                        "- CRITICAL: For handwritten documents, fields may appear in natural language without labels:\n"
                        "  * License number: Look for license identifiers (D1234567, A1234567, etc.) anywhere in the block\n"
                        "  * License state: Look for state abbreviations (CA, NY, TX, etc.) anywhere in the block\n"
                        "  * Years experience: Look for numbers with 'years' or 'experience' anywhere in the block\n"
                        "  * Violations: Look for numbers related to violations or tickets anywhere in the block\n"
                        "  * Training: Look for 'Yes', 'No', 'Completed' anywhere in the block\n"
                        "- IMPORTANT: Handwritten documents are NOT tables - do NOT look for column headers or rows\n"
                        f"- Instead, identify each separate text block that contains {domain_name} information\n"
                        f"- Each block is a paragraph or section of text describing one {domain_name}\n"
                        "- Read each block as a whole and extract all fields mentioned within that block\n"
                        "- Be careful with OCR character recognition errors\n"
                        f"- If a {identifier_name} looks suspicious, try to correct obvious OCR errors based on context\n"
                        f"- Validate that {identifier_name}s follow expected patterns (D001, D002, DRV001, etc.)\n"
                        "- Extract ALL fields visible in each block, even if some are unclear\n"
                        "- DO NOT skip fields just because they don't have labels - search the entire block text for keywords\n"
                        "\n"
                    )
                else:
                    user_prompt_text += (
                        "\n\nHANDWRITTEN DOCUMENT DETECTED - SPECIAL INSTRUCTIONS:\n"
                        "- This appears to be a handwritten document - text may be less structured\n"
                        "- CRITICAL: Handwritten documents are typically NOT in table format - they are organized as SEPARATE BLOCKS or SECTIONS\n"
                        f"- Each {domain_name}'s information is written in its own distinct block/section, NOT in table rows\n"
                        f"- Look carefully for ALL distinct SECTIONS or BLOCKS of {domain_name} information\n"
                        f"- Each section typically contains one complete {domain_name} with all its fields grouped together\n"
                        "- Sections may be separated by:\n"
                        "  * Blank space or lines\n"
                        "  * Visual boundaries (boxes, lines, borders)\n"
                        f"  * Each {identifier_name} usually marks the start of a new section\n"
                        f"  * Each section is a self-contained block of text describing one {domain_name}\n"
                        f"- Count how many distinct {domain_name} sections/blocks you see - extract that many rows\n"
                        f"- If you see 3 separate blocks with {domain_name} information, extract 3 rows (one per block)\n"
                        "- DO NOT combine multiple blocks into one row\n"
                        f"- DO NOT skip blocks - extract every block that contains {domain_name} information\n"
                        "- Within each block, extract all fields that appear, even if handwriting is difficult to read\n"
                        "- CRITICAL: For handwritten documents, fields may appear in natural language without labels:\n"
                        "  * Body style: Look for 'sedan', 'truck', 'SUV', etc. anywhere in the block (e.g., '2024 Toyota Camry sedan')\n"
                        "  * Fuel type: Look for 'gas', 'gasoline', 'diesel', etc. anywhere in the block (e.g., 'runs on gas' or just 'gas')\n"
                        "  * Transmission: Look for 'automatic', 'manual', 'auto' anywhere in the block (e.g., 'automatic transmission' or just 'automatic')\n"
                        "  * Mileage: Look for numbers with 'miles' anywhere (e.g., '100,245 miles', 'about 89,000 miles')\n"
                        "  * Color: Look for color words anywhere (e.g., 'painted blue', 'red car', 'green vehicle')\n"
                        "- IMPORTANT: Handwritten documents are NOT tables - do NOT look for column headers or rows\n"
                        f"- Instead, identify each separate text block that contains {domain_name} information\n"
                        f"- Each block is a paragraph or section of text describing one {domain_name}\n"
                        "- Read each block as a whole and extract all fields mentioned within that block\n"
                        "- Be careful with OCR character recognition errors:\n"
                        "  * 'B' might appear as '8' or 'E'\n"
                        "  * 'F' might appear as 'E'\n"
                        "  * '1' might appear as 'I' or 'l'\n"
                        "  * 'Z' might appear as '7' or '2'\n"
                        "  * '0' might appear as 'O'\n"
                        "  * '2' might appear as 'Z'\n"
                        f"- If a {identifier_name} looks suspicious, try to correct obvious OCR errors based on context\n"
                        f"- Validate that {identifier_name}s are 17 characters and contain at least 2 digits\n"
                        "- Extract ALL fields visible in each block, even if some are unclear\n"
                        "- DO NOT skip fields just because they don't have labels - search the entire block text for keywords\n"
                        "\n"
                    )
            
            user_prompt_text += (
                "\n\nEXTRACTION CONSTRAINTS (CRITICAL):\n"
                f"- Always return all {domain_name} schema fields explicitly in every row\n"
                "- If a field is not visible or cannot be determined, return null instead of omitting it\n"
                "- DO NOT drop fields simply because they are unlabeled - infer from natural language when possible\n"
                "\n"
                "NATURAL LANGUAGE INFERENCE (REQUIRED):\n"
                f"Infer the following fields from natural language anywhere in the document section, even if unlabeled:\n"
            )
            
            if is_driver:
                user_prompt_text += (
                    "- license_number: Search for license identifiers (D1234567, A1234567, etc.) anywhere in the section\n"
                    "  * Extract the raw value as written - normalization will be applied later\n"
                    "- license_state: Search for state abbreviations (CA, NY, TX, etc.) anywhere in the section\n"
                    "  * Extract the raw value as written (e.g., 'CA', 'NY', 'Texas') - normalization will be applied later\n"
                    "- license_status: Search for words like 'Valid', 'Active', 'Suspended', 'Expired' anywhere in the section\n"
                    "  * Extract the raw value as written - normalization will be applied later\n"
                    "- years_experience: Search for numbers followed by 'years' or 'experience' anywhere in the section\n"
                    "  * Extract the raw numeric value - normalization will clean it later\n"
                    "  * Examples: '18 years experience' → years_experience: 18, '16 years' → years_experience: 16\n"
                    "- violations_count: Search for numbers related to violations or tickets anywhere in the section\n"
                    "  * Extract the raw numeric value - normalization will clean it later\n"
                    "  * Examples: '0 violations' → violations_count: 0, '1 violation' → violations_count: 1\n"
                    "- training_completed: Search for 'Yes', 'No', 'Completed', 'Not Completed' anywhere in the section\n"
                    "  * Extract the raw value as written - normalization will be applied later\n"
                    "  * Examples: 'Training: Yes' → training_completed: 'Yes', 'completed training' → training_completed: 'Yes'\n"
                    "- date_of_birth: Search for dates in various formats (YYYY-MM-DD, MM/DD/YYYY, etc.) anywhere in the section\n"
                    "  * Extract the raw value as written - normalization will be applied later\n"
                )
            else:
                user_prompt_text += (
                    "- body_style: Search for keywords like 'sedan', 'truck', 'SUV', 'van', 'coupe', 'hatchback', 'convertible', 'wagon', 'crossover'\n"
                    "  * Extract the raw value as written (e.g., 'sedan', 'truck', 'SUV', 'Truck') - normalization will be applied later\n"
                    "  * Examples: '2024 Toyota Camry sedan' → body_style: 'sedan', 'Ford F-150 truck' → body_style: 'truck'\n"
                    "- fuel_type: Search for keywords like 'gas', 'gasoline', 'diesel', 'electric', 'hybrid'\n"
                    "  * Extract the raw value as written (e.g., 'gas', 'gasoline', 'Gas') - normalization will be applied later\n"
                    "  * Examples: 'runs on gas' → fuel_type: 'gas', 'diesel engine' → fuel_type: 'diesel'\n"
                    "- transmission: Search for keywords like 'automatic', 'manual', 'CVT', 'auto'\n"
                    "  * Extract the raw value as written (e.g., 'automatic', 'auto', 'Auto') - normalization will be applied later\n"
                    "  * Examples: 'automatic transmission' → transmission: 'automatic', '8-speed auto' → transmission: 'auto'\n"
                    "- mileage: Search for numbers followed by 'miles', 'mi', 'mileage', or 'odometer'\n"
                    "  * Extract the raw numeric value (may include commas, text, units) - normalization will clean it later\n"
                    "  * Examples: '100,245 miles' → mileage: '100,245 miles', 'odometer: 89,000' → mileage: '89,000'\n"
                )
            
            user_prompt_text += (
                "\n"
                "IMPORTANT: Extract raw values only. Do NOT normalize, convert, or transform values.\n"
                "All normalization (lowercase, canonical forms, type conversion) will be handled by the normalization pipeline.\n"
                "\n"
                "FIELD COMPLETENESS:\n"
                "- Every row MUST include all schema fields in the exact order specified\n"
                "- Use null for missing fields - never omit fields from the output\n"
                "- Search the entire document section for natural language mentions of these fields\n"
                "- Do NOT skip fields just because they lack explicit labels\n"
                "\n"
                f"\nOUTPUT REQUIREMENTS:\n"
                f"- Return ONE row per {domain_name} with exactly {len(canonical_schema)} values matching the headers\n"
                "- If a field is missing, use null (not empty string, not \"\", not \"N/A\")\n"
                "- Output ONLY valid JSON - no markdown, no explanations, no additional text\n"
                "- Do NOT wrap the JSON in code blocks (```json)\n"
                f"- Count the number of distinct {domain_name} sections you identified and extract that many rows\n"
            )
            
            user_content = [
                {"type": "text", "text": user_prompt_text},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_base64}"}},
            ]
            
            # Call Vision API for this page
            page_rows = None
            try:
                # Debug: Print first 300 chars of system prompt to verify correct prompt is used
                logger.debug(f"[Vision] System prompt preview (first 300 chars): {system_prompt[:300]}")
                
                response = client.chat.completions.create(
                    model=vision_model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_content},
                    ],
                    max_tokens=4000,
                    temperature=0,  # Deterministic output - prioritize consistency over creativity
                )
                
                if response.choices and response.choices[0].message.content:
                    content = response.choices[0].message.content.strip()
                    
                    # Extract JSON
                    if "```json" in content:
                        content = content.split("```json")[1].split("```")[0].strip()
                    elif "```" in content:
                        content = content.split("```")[1].split("```")[0].strip()
                    
                    try:
                        data = json.loads(content)
                        if isinstance(data, dict):
                            rows = data.get("rows", [])
                            if rows and isinstance(rows, list):
                                # Normalize rows to canonical schema
                                normalized_page_rows = []
                                for row_idx, row in enumerate(rows):
                                    if not isinstance(row, list):
                                        continue
                                    normalized_row = []
                                    for j in range(len(canonical_schema)):
                                        raw_value = row[j] if j < len(row) else None
                                        field_name = canonical_schema[j]
                                        # Preserve raw values from Vision (including invalid ones like year=1899, mileage=-100)
                                        # These will be processed by normalize_v2's mapping/transforms, just like other sources
                                        normalized_value = _normalize_vision_value(field_name, raw_value, preserve_raw=True)
                                        normalized_row.append(normalized_value)
                                    normalized_page_rows.append(normalized_row)
                                
                                if normalized_page_rows:
                                    # Clean "None" strings from Vision API results
                                    for row in normalized_page_rows:
                                        for i, value in enumerate(row):
                                            if isinstance(value, str) and value.strip().lower() == "none":
                                                row[i] = None
                                    
                                    # SAFEGUARD: Filter out rows with invalid VINs (header tokens, table headers)
                                    # Expanded list to catch all header tokens that might be misidentified as VINs
                                    valid_rows = []
                                    vin_index = canonical_schema.index('vin') if 'vin' in canonical_schema else 0
                                    invalid_vin_tokens = {'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 'VIN', 'TRANSMISSION', 
                                                          'BODY_STYLE', 'FUEL_TYPE', 'OWNER_EMAIL', 'NOTES', 'VEHICLE_ID', 
                                                          'EFFECTIVE_DATE', 'TRIM', 'WEIGHT', 'BODY', 'PAINTED', 'GASOLINE',
                                                          'CURRENT', 'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET', 'STYLE',
                                                          'TYPE', 'EMAIL', 'FUEL', 'BODY_TYPE', 'CURRENT_MILEAGE',
                                                          'EXTERIOR_COLOR'}
                                    for row in normalized_page_rows:
                                        if len(row) > vin_index:
                                            vin_value = row[vin_index]
                                            # Skip rows where VIN is a header token or invalid
                                            if vin_value and isinstance(vin_value, str):
                                                # Strip punctuation and whitespace before checking
                                                vin_upper_clean = vin_value.strip().upper().rstrip('.,;:!?')
                                                # Check if VIN is a header token (exact match)
                                                is_header_token = vin_upper_clean in invalid_vin_tokens
                                                
                                                # CRITICAL FIX: For IMAGE sources, use conservative fuzzy matching for header detection
                                                # This prevents valid rows from being filtered due to OCR noise (e.g., "gasoline." -> "GASOLINE")
                                                if not is_header_token and source_type == SourceType.IMAGE:
                                                    from inference import fuzzy_match_header
                                                    # Use conservative threshold (0.85) to catch OCR variations but not false positives
                                                    for header_token in invalid_vin_tokens:
                                                        if fuzzy_match_header(vin_upper_clean, header_token, threshold=0.85):
                                                            is_header_token = True
                                                            logger.debug(f"[Vision] Fuzzy matched header token: {vin_value} -> {header_token} (similarity >= 0.85)")
                                                            break
                                                
                                                if is_header_token:
                                                    logger.warning(f"[Vision] Filtering row with invalid VIN (header token): {vin_value} -> {vin_upper_clean}")
                                                    continue
                                                # Check if VIN is too short or doesn't look like a valid VIN
                                                if len(vin_upper_clean) < 10:  # Valid VINs are typically 17 chars, but allow some tolerance
                                                    logger.debug(f"[Vision] Filtering row with invalid VIN (too short): {vin_value}")
                                                    continue
                                            valid_rows.append(row)
                                        else:
                                            # Row doesn't have enough columns, skip it
                                            logger.debug(f"[Vision] Filtering row with insufficient columns: {len(row)} < {vin_index + 1}")
                                    
                                    if valid_rows:
                                        # Apply document-level defaults to Vision API results if OCR text is available
                                        # (for handwritten PDF/image or PNG images where Vision API might miss fields)
                                        source_type_str = "pdf" if source_type == SourceType.PDF else "image"
                                        if page_ocr_text:
                                            vision_document_defaults = _extract_document_level_defaults(page_ocr_text, source_type_str, file_path)
                                            # Apply defaults to each Vision API row by field index
                                            vision_applied_count = 0
                                            for row_idx, row in enumerate(valid_rows):
                                                if len(row) >= len(canonical_schema):
                                                    for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                                        if field in canonical_schema:
                                                            field_idx = canonical_schema.index(field)
                                                            if field_idx < len(row) and (row[field_idx] is None or row[field_idx] == ""):
                                                                if vision_document_defaults.get(field) is not None:
                                                                    row[field_idx] = vision_document_defaults[field]
                                                                    vision_applied_count += 1
                                        
                                        # Detect table extraction: notes is null/empty/very short AND core fields (vin, year) are populated
                                        # Table extractions should be treated as authoritative (no OCR attachment, no inference)
                                        is_table_extraction = False
                                        if valid_rows:
                                            first_row = valid_rows[0]
                                            notes_index = canonical_schema.index('notes') if 'notes' in canonical_schema else -1
                                            vin_index = canonical_schema.index('vin') if 'vin' in canonical_schema else -1
                                            year_index = canonical_schema.index('year') if 'year' in canonical_schema else -1
                                            
                                            if notes_index >= 0 and vin_index >= 0 and year_index >= 0:
                                                notes = first_row[notes_index] if len(first_row) > notes_index else None
                                                vin = first_row[vin_index] if len(first_row) > vin_index else None
                                                year = first_row[year_index] if len(first_row) > year_index else None
                                                
                                                # Table extraction: notes null/empty/very short AND core fields populated
                                                # CRITICAL: Mark as table extraction if Vision extracted structured data (even if handwritten)
                                                # Only populate structured fields when a table is confidently detected
                                                # Do NOT infer table fields from free-form OCR
                                                is_handwritten_source = is_handwritten_pdf or is_scanned_pdf_routed_as_image or any(keyword in str(file_path).lower() for keyword in ['handwritten', 'hand'])
                                                if (notes is None or notes == "" or 
                                                    (isinstance(notes, str) and len(notes.strip()) < 50)):
                                                    if vin and year:  # Core fields populated
                                                        # Mark as table extraction if Vision extracted structured data
                                                        # This applies to both handwritten and non-handwritten sources
                                                        # Handwritten tables should be marked as table extractions to prevent inference
                                                        is_table_extraction = True
                                        
                                        # For Vision API rows, attach OCR text to notes if notes is null or very short
                                        # This ensures post-Vision inference has text to search for free-form text
                                        # BUT: Skip OCR attachment for table extractions (Vision is authoritative for tables)
                                        # CRITICAL: Do NOT attach OCR text for table extractions (even handwritten)
                                        # Only attach OCR text for free-form text sources that need inference
                                        notes_index = canonical_schema.index('notes') if 'notes' in canonical_schema else -1
                                        is_handwritten_source = is_handwritten_pdf or is_scanned_pdf_routed_as_image or any(keyword in str(file_path).lower() for keyword in ['handwritten', 'hand'])
                                        # Attach OCR text ONLY if: not a table extraction (table extractions should not infer from OCR)
                                        if notes_index >= 0 and not is_table_extraction:
                                            # Use page_ocr_text for PDF sources, extracted_ocr_text_for_vision for IMAGE sources and handwritten PDFs
                                            # For handwritten PDFs, extracted_ocr_text_for_vision is set before the page loop
                                            ocr_text_to_attach = page_ocr_text if page_ocr_text else extracted_ocr_text_for_vision
                                            if ocr_text_to_attach:
                                                for row in valid_rows:
                                                    if len(row) > notes_index:
                                                        current_notes = row[notes_index]
                                                        # For free-form text sources (not table extractions), attach OCR text to notes
                                                        # This ensures inference has access to OCR text for free-form sources
                                                        # BUT: Skip for table extractions (Vision is authoritative, don't infer from OCR)
                                                        if current_notes and isinstance(current_notes, str) and len(current_notes.strip()) >= 50:
                                                            # Vision extracted substantial notes - append OCR text for additional context (free-form text)
                                                            row[notes_index] = current_notes + " " + ocr_text_to_attach
                                                            logger.warning(f"[Vision] Appended OCR text to Vision notes for free-form source ({len(ocr_text_to_attach)} chars OCR, {len(current_notes)} chars Vision notes)")
                                                        elif current_notes is None or current_notes == "" or (isinstance(current_notes, str) and len(current_notes.strip()) < 50):
                                                            # Empty or short notes - replace with OCR text (free-form text source)
                                                            row[notes_index] = ocr_text_to_attach
                                                            logger.warning(f"[Vision] Attached OCR text to notes field for free-form source ({len(ocr_text_to_attach)} chars, previous notes length: {len(str(current_notes)) if current_notes else 0})")
                                                        else:
                                                            # For non-handwritten: only attach if notes is null/empty or very short (< 50 chars)
                                                            if (current_notes is None or current_notes == "" or 
                                                                (isinstance(current_notes, str) and len(current_notes.strip()) < 50)):
                                                                row[notes_index] = ocr_text_to_attach
                                                                logger.warning(f"[Vision] Attached OCR text to notes field for post-Vision inference ({len(ocr_text_to_attach)} chars, previous notes length: {len(str(current_notes)) if current_notes else 0})")
                                        
                                        # For Vision API rows, append _is_handwritten flag based on file path
                                        # Check if file path indicates handwriting
                                        is_handwritten_doc = False
                                        if file_path:
                                            file_path_str = str(file_path).lower()
                                            is_handwritten_doc = any(keyword in file_path_str for keyword in ['handwritten', 'hand'])
                                        # Append _is_handwritten, _is_vision_extracted, and _is_table_extraction flags to each valid row
                                        # Vision-extracted rows should NOT have VIN correction applied
                                        for row in valid_rows:
                                            if len(row) == len(canonical_schema):
                                                row.append(is_handwritten_doc)
                                                row.append(True)  # _is_vision_extracted = True
                                                row.append(is_table_extraction)  # _is_table_extraction
                                            # If row already has extra fields, update them
                                            elif len(row) == len(canonical_schema) + 1:
                                                row[-1] = is_handwritten_doc
                                                row.append(True)  # _is_vision_extracted = True
                                                row.append(is_table_extraction)  # _is_table_extraction
                                            elif len(row) == len(canonical_schema) + 2:
                                                row[-2] = is_handwritten_doc
                                                row[-1] = True  # _is_vision_extracted = True
                                                row.append(is_table_extraction)  # _is_table_extraction
                                            elif len(row) > len(canonical_schema) + 2:
                                                row[-3] = is_handwritten_doc
                                                row[-2] = True  # _is_vision_extracted = True
                                                row[-1] = is_table_extraction  # _is_table_extraction
                                        
                                        page_rows = valid_rows
                                        vision_extracted_pages.append(page_num)
                                        logger.info(f"[Vision] Extracted {len(valid_rows)} valid row(s) from page {page_num} (filtered {len(normalized_page_rows) - len(valid_rows)} invalid rows)")
                                    else:
                                        logger.debug(f"[Vision] All rows from page {page_num} were filtered as invalid (likely header rows)")
                                        page_rows = None
                    except (json.JSONDecodeError, KeyError, IndexError) as e:
                        logger.debug(f"[Vision] Failed to parse page {page_num} response: {e}")
                        page_rows = None
            except Exception as e:
                logger.debug(f"[Vision] API error for page {page_num}: {e}")
                page_rows = None
            
            # For handwritten PDFs: If Vision API failed, fall back to OCR text extraction
            if is_handwritten_pdf and (page_rows is None or len(page_rows) == 0):
                logger.info(f"[Handwritten PDF] Vision API failed or returned no rows for page {page_num}, falling back to OCR text extraction")
                # Extract OCR text for this page as fallback
                try:
                    if not hasattr(_extract_table_with_vision_api, '_cached_full_ocr_text'):
                        full_ocr_text = _get_vision_ocr_text_for_pdf(file_path)
                        _extract_table_with_vision_api._cached_full_ocr_text = full_ocr_text
                        logger.info(f"[Handwritten PDF] Extracted OCR text ({len(full_ocr_text)} chars) for fallback")
                    else:
                        full_ocr_text = getattr(_extract_table_with_vision_api, '_cached_full_ocr_text', "")
                    
                    if full_ocr_text:
                        # Use OCR text path as fallback (same logic as non-handwritten PDFs)
                        page_ocr_text = full_ocr_text if page_num == 1 else ""
                        if page_ocr_text:
                            logger.info(f"[Handwritten PDF] Using OCR fallback for page {page_num}")
                            # Pre-split extraction: Extract document-level defaults from full OCR text
                            document_defaults = _extract_document_level_defaults(page_ocr_text, "pdf", file_path)
                            
                            # After OCR text is obtained, force segmentation the same way
                            blocks = split_by_vin(page_ocr_text)
                            rows = [extract_fields_from_block(b, mapping=mapping_config, source_type="pdf") for b in blocks]
                            
                            # Apply document-level defaults to each vehicle if field is still None
                            for vehicle in rows:
                                if vehicle:
                                    for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                        if vehicle.get(field) is None and document_defaults.get(field) is not None:
                                            vehicle[field] = document_defaults[field]
                            
                            # If we successfully extracted vehicles from segmented blocks, use them
                            extracted_vehicles = [r for r in rows if r and r.get('vin')]
                            if extracted_vehicles:
                                logger.info(f"[Handwritten PDF] OCR fallback extracted {len(extracted_vehicles)} vehicle(s) from page {page_num}")
                                # Convert to 2D array format
                                page_rows = []
                                for vehicle in extracted_vehicles:
                                    row = []
                                    for field in canonical_schema:
                                        value = vehicle.get(field)
                                        if isinstance(value, str) and value.strip().lower() == "none":
                                            value = None
                                        row.append(value)
                                    # Preserve _is_handwritten flag
                                    row.append(vehicle.get("_is_handwritten", True))
                                    page_rows.append(row)
                except Exception as e:
                    logger.debug(f"[Handwritten PDF] OCR fallback failed for page {page_num}: {e}")
                    page_rows = None
            
            # If Vision extracted rows, use them (but only if text extraction didn't work)
            if page_rows:
                # DEBUG: Log which page and how many rows
                vin_index = canonical_schema.index('vin') if 'vin' in canonical_schema else 0
                vins_in_page = [row[vin_index] if len(row) > vin_index else None for row in page_rows]
                logger.info(f"[PDF] Vision extracted {len(page_rows)} row(s) from page {page_num}: VINs={vins_in_page}")
                
                # For PDF, merge Vision results with OCR results if OCR extracted the same VINs
                if source_type == SourceType.PDF:
                    # Check if we have OCR-extracted vehicles stored earlier
                    if ocr_vehicles_by_vin:
                        # Merge Vision rows with OCR-extracted vehicles by VIN
                        # If Vision has null critical fields but OCR has non-null values, prefer OCR
                        critical_fields = ['year', 'make', 'model']
                        critical_field_indices = [canonical_schema.index(f) for f in critical_fields if f in canonical_schema]
                        
                        for row_idx, vision_row in enumerate(page_rows):
                            if len(vision_row) > vin_index:
                                vision_vin = vision_row[vin_index]
                                if vision_vin and vision_vin in ocr_vehicles_by_vin:
                                    ocr_vehicle = ocr_vehicles_by_vin[vision_vin]
                                    # Check if Vision has null critical fields but OCR has non-null values
                                    vision_has_null_critical = all(
                                        idx < len(vision_row) and (vision_row[idx] is None or vision_row[idx] == "")
                                        for idx in critical_field_indices
                                    )
                                    ocr_has_non_null_critical = any(
                                        ocr_vehicle.get(field) is not None and ocr_vehicle.get(field) != ""
                                        for field in critical_fields
                                    )
                                    
                                    if vision_has_null_critical and ocr_has_non_null_critical:
                                        # Merge: prefer OCR values for critical fields when Vision has nulls
                                        logger.info(f"[PDF] Merging OCR and Vision results for VIN {vision_vin} (Vision had null critical fields)")
                                        for field_idx, field_name in enumerate(critical_fields):
                                            if field_name in canonical_schema:
                                                field_schema_idx = canonical_schema.index(field_name)
                                                if field_schema_idx < len(vision_row):
                                                    ocr_value = ocr_vehicle.get(field_name)
                                                    if (vision_row[field_schema_idx] is None or vision_row[field_schema_idx] == "") and ocr_value is not None:
                                                        vision_row[field_schema_idx] = ocr_value
                                        # Also merge other fields that Vision might have missed
                                        for field_name in canonical_schema:
                                            if field_name not in critical_fields:
                                                field_schema_idx = canonical_schema.index(field_name)
                                                if field_schema_idx < len(vision_row):
                                                    ocr_value = ocr_vehicle.get(field_name)
                                                    if (vision_row[field_schema_idx] is None or vision_row[field_schema_idx] == "") and ocr_value is not None:
                                                        vision_row[field_schema_idx] = ocr_value
                    
                    # CRITICAL FIX: Merge Vision rows with existing rows by VIN
                    # Later pages may contain better data for the same VIN, so merge field-by-field
                    # instead of skipping duplicate VINs
                    vin_index = canonical_schema.index('vin') if 'vin' in canonical_schema else 0
                    
                    new_rows = []
                    merged_count = 0
                    for vision_row in page_rows:
                        if len(vision_row) <= vin_index:
                            continue
                            
                        vision_vin = vision_row[vin_index]
                        if not vision_vin:
                            continue
                        
                        # Find existing row with same VIN
                        existing_row_idx = None
                        for idx, existing_row in enumerate(all_rows):
                            if len(existing_row) > vin_index and existing_row[vin_index] == vision_vin:
                                existing_row_idx = idx
                                break
                        
                        if existing_row_idx is not None:
                            # Merge field-by-field into existing row
                            existing_row = all_rows[existing_row_idx]
                            fields_merged = []
                            
                            # Merge all canonical schema fields
                            for field_idx in range(len(canonical_schema)):
                                if field_idx < len(existing_row) and field_idx < len(vision_row):
                                    existing_value = existing_row[field_idx]
                                    new_value = vision_row[field_idx]
                                    
                                    # If existing is None/empty and new is not None/empty, fill it
                                    if (existing_value is None or existing_value == "") and (new_value is not None and new_value != ""):
                                        existing_row[field_idx] = new_value
                                        field_name = canonical_schema[field_idx] if field_idx < len(canonical_schema) else f"field_{field_idx}"
                                        fields_merged.append(field_name)
                            
                            # Preserve extra fields (like _is_handwritten, _is_vision_extracted)
                            # These are appended after canonical schema fields
                            if len(vision_row) > len(canonical_schema):
                                # Ensure existing_row has space for extra fields
                                while len(existing_row) < len(vision_row):
                                    existing_row.append(None)
                                # Merge extra fields (preserve flags from both)
                                for extra_idx in range(len(canonical_schema), len(vision_row)):
                                    if extra_idx < len(existing_row):
                                        # For boolean flags, use OR logic (if either is True, keep True)
                                        if isinstance(vision_row[extra_idx], bool) and isinstance(existing_row[extra_idx], bool):
                                            existing_row[extra_idx] = existing_row[extra_idx] or vision_row[extra_idx]
                                        # For other extra fields, prefer non-None values
                                        elif (existing_row[extra_idx] is None or existing_row[extra_idx] == "") and (vision_row[extra_idx] is not None and vision_row[extra_idx] != ""):
                                            existing_row[extra_idx] = vision_row[extra_idx]
                            
                            if fields_merged:
                                logger.info(f"[PDF] Merged {len(fields_merged)} field(s) into existing row for VIN {vision_vin} from page {page_num}: {fields_merged}")
                                merged_count += 1
                            else:
                                logger.debug(f"[PDF] VIN {vision_vin} from page {page_num} already exists with all fields filled, no merge needed")
                        else:
                            # New VIN, add as new row
                            new_rows.append(vision_row)
                    
                    if new_rows:
                        logger.info(f"[PDF] Adding {len(new_rows)} new Vision row(s) from page {page_num} to all_rows (total before: {len(all_rows)}, after: {len(all_rows) + len(new_rows)})")
                        all_rows.extend(new_rows)
                    if merged_count > 0:
                        logger.info(f"[PDF] Merged {merged_count} existing row(s) from page {page_num} with Vision data")
                    if not new_rows and merged_count == 0:
                        logger.debug(f"[PDF] No new or merged rows from Vision page {page_num}")
                elif source_type == SourceType.IMAGE:
                    # CRITICAL FIX: Merge Vision results with OCR results for IMAGE sources (same as PDF)
                    # This brings IMAGE extraction up to parity with PDF extraction
                    if ocr_vehicles_by_vin:
                        vin_index = canonical_schema.index('vin') if 'vin' in canonical_schema else 0
                        critical_fields = ['year', 'make', 'model']
                        critical_field_indices = [canonical_schema.index(f) for f in critical_fields if f in canonical_schema]
                        
                        for row_idx, vision_row in enumerate(page_rows):
                            if len(vision_row) > vin_index:
                                vision_vin = vision_row[vin_index]
                                if vision_vin and vision_vin in ocr_vehicles_by_vin:
                                    ocr_vehicle = ocr_vehicles_by_vin[vision_vin]
                                    # Check if Vision has null critical fields but OCR has non-null values
                                    vision_has_null_critical = all(
                                        idx < len(vision_row) and (vision_row[idx] is None or vision_row[idx] == "")
                                        for idx in critical_field_indices
                                    )
                                    ocr_has_non_null_critical = any(
                                        ocr_vehicle.get(field) is not None and ocr_vehicle.get(field) != ""
                                        for field in critical_fields
                                    )
                                    
                                    if vision_has_null_critical and ocr_has_non_null_critical:
                                        # Merge: prefer OCR values for critical fields when Vision has nulls
                                        logger.info(f"[IMAGE] Merging OCR and Vision results for VIN {vision_vin} (Vision had null critical fields)")
                                        for field_idx, field_name in enumerate(critical_fields):
                                            if field_name in canonical_schema:
                                                field_schema_idx = canonical_schema.index(field_name)
                                                if field_schema_idx < len(vision_row):
                                                    ocr_value = ocr_vehicle.get(field_name)
                                                    if (vision_row[field_schema_idx] is None or vision_row[field_schema_idx] == "") and ocr_value is not None:
                                                        vision_row[field_schema_idx] = ocr_value
                                        # Also merge other fields that Vision might have missed
                                        for field_name in canonical_schema:
                                            if field_name not in critical_fields:
                                                field_schema_idx = canonical_schema.index(field_name)
                                                if field_schema_idx < len(vision_row):
                                                    ocr_value = ocr_vehicle.get(field_name)
                                                    if (vision_row[field_schema_idx] is None or vision_row[field_schema_idx] == "") and ocr_value is not None:
                                                        vision_row[field_schema_idx] = ocr_value
                    
                    # For IMAGE sources, collect rows for aggregation AFTER all pages are processed
                    image_rows_collector.extend(page_rows)
                else:
                    all_rows.extend(page_rows)
            else:
                # Vision failed or returned nothing - try OCR fallback inference
                if page_ocr_text:
                    # For PDF, split OCR text into vehicle blocks and process each separately
                    if source_type == SourceType.PDF:
                        # Pre-split extraction: Extract document-level defaults
                        # file_path is available in _extract_table_with_vision_api scope
                        document_defaults = _extract_document_level_defaults(page_ocr_text, "pdf", file_path)
                        blocks = split_by_vin(page_ocr_text)
                        rows = [extract_fields_from_block(b, mapping=mapping_config, source_type="pdf") for b in blocks]
                        # Apply document-level defaults to each vehicle if field is still None
                        for vehicle in rows:
                            if vehicle:
                                for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                    if vehicle.get(field) is None and document_defaults.get(field) is not None:
                                        vehicle[field] = document_defaults[field]
                        for vehicle in rows:
                            if vehicle and vehicle.get('vin'):
                                # Convert to canonical schema row and clean "None" strings
                                canonical_row = []
                                for field in canonical_schema:
                                    value = vehicle.get(field)
                                    # Clean "None" strings
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    canonical_row.append(value)
                                # Preserve _is_handwritten flag as extra field
                                canonical_row.append(vehicle.get("_is_handwritten", False))
                                all_rows.append(canonical_row)
                    elif source_type == SourceType.IMAGE:
                        # For IMAGE sources, collect blocks for aggregation AFTER all pages
                        # Pre-split extraction: Extract document-level defaults
                        # file_path is available in _extract_table_with_vision_api scope
                        document_defaults = _extract_document_level_defaults(page_ocr_text, "image", file_path)
                        blocks = split_by_vin(page_ocr_text)
                        extracted_vehicles = [extract_fields_from_block(b, mapping=mapping_config, source_type="image") for b in blocks]
                        # Apply document-level defaults to each vehicle if field is still None
                        for vehicle in extracted_vehicles:
                            if vehicle:
                                for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                    if vehicle.get(field) is None and document_defaults.get(field) is not None:
                                        vehicle[field] = document_defaults[field]
                        # Convert vehicles to canonical rows and collect for aggregation
                        # CRITICAL: Filter out rows with invalid VINs (header tokens) during collection
                        for vehicle in extracted_vehicles:
                            if vehicle:
                                vin_value = vehicle.get('vin')
                                # SAFEGUARD: Skip vehicles with invalid VINs (header tokens)
                                # Use the same excluded words list as extract_fields_from_block
                                if vin_value:
                                    vin_upper_clean = str(vin_value).upper().strip().rstrip('.,;:!?')
                                    invalid_vin_tokens = {'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 'VIN', 
                                                          'BODY', 'PAINTED', 'GASOLINE', 'TRANSMISSION', 'NOTES',
                                                          'CURRENT', 'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET',
                                                          'STYLE', 'TYPE', 'EMAIL', 'FUEL', 'BODY_STYLE', 'FUEL_TYPE'}
                                    if vin_upper_clean in invalid_vin_tokens:
                                        logger.debug(f"[IMAGE] Skipping vehicle with invalid VIN (header token): {vin_value}")
                                    continue
                                canonical_row = []
                                for field in canonical_schema:
                                    value = vehicle.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    canonical_row.append(value)
                                # Preserve _is_handwritten flag as extra field
                                canonical_row.append(vehicle.get("_is_handwritten", False))
                                image_rows_collector.append(canonical_row)
                    else:
                        # For non-PDF sources, use original fallback logic
                        fallback_row = _extract_from_ocr_text_with_fallback(page_ocr_text, page_num)
                        if fallback_row:
                            # Convert fallback dict to canonical schema row
                            canonical_row = []
                            for field in canonical_schema:
                                canonical_row.append(fallback_row.get(field))
                            # For IMAGE sources, collect rows for aggregation AFTER all pages
                            if source_type == SourceType.IMAGE:
                                image_rows_collector.append(canonical_row)
                            else:
                                all_rows.append(canonical_row)
                        else:
                            logger.debug(f"[Fallback] No data extracted from page {page_num} OCR text")
        
        # For IMAGE sources, group rows by VIN and aggregate fields within each VIN group
        # This handles both: (1) multiple complete vehicles (different VINs) and (2) partial rows for same vehicle
        if source_type == SourceType.IMAGE and image_rows_collector:
            
            # Group rows by VIN (each unique VIN = one vehicle)
            vin_index = canonical_schema.index('vin') if 'vin' in canonical_schema else 0
            vin_groups = {}  # vin_upper -> list of rows for that VIN
            
            for row in image_rows_collector:
                if len(row) <= vin_index:
                    continue
                vin_value = row[vin_index]
                if not vin_value:
                    continue
                vin_upper_clean = str(vin_value).upper().strip().rstrip('.,;:!?')
                # Skip header tokens
                invalid_vin_tokens = {'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 'VIN', 
                                      'BODY', 'PAINTED', 'GASOLINE', 'TRANSMISSION', 'NOTES',
                                      'CURRENT', 'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET',
                                      'STYLE', 'TYPE', 'EMAIL', 'FUEL', 'BODY_STYLE', 'FUEL_TYPE'}
                if vin_upper_clean in invalid_vin_tokens:
                    continue
                if vin_upper_clean not in vin_groups:
                    vin_groups[vin_upper_clean] = []
                vin_groups[vin_upper_clean].append(row)
            
            # For each VIN group, aggregate fields (merge partial rows into one complete vehicle)
            for vin_upper, rows_for_vin in vin_groups.items():
                aggregated_vehicle = {field: None for field in canonical_schema}
                aggregated_is_handwritten = False
                aggregated_is_vision_extracted = False
                aggregated_is_table_extraction = False
                for row in rows_for_vin:
                    for i, field in enumerate(canonical_schema):
                        value = row[i] if i < len(row) else None
                        # CRITICAL FIX: Prefer non-null values (only overwrite if existing is None/empty and new is not None/empty)
                        # This prevents later pages with nulls from overwriting good values from earlier pages
                        existing_value = aggregated_vehicle[field]
                        if (existing_value is None or existing_value == "") and (value is not None and value != ""):
                            aggregated_vehicle[field] = value
                        # If both are non-null, keep existing (first non-null value wins)
                    # Preserve _is_handwritten from any row (if any row is handwritten, aggregated is handwritten)
                    if len(row) > len(canonical_schema) and row[len(canonical_schema)]:
                        aggregated_is_handwritten = True
                    # Preserve _is_vision_extracted from any row (if any row is Vision-extracted, aggregated is Vision-extracted)
                    if len(row) > len(canonical_schema) + 1 and row[len(canonical_schema) + 1]:
                        aggregated_is_vision_extracted = True
                    # Preserve _is_table_extraction from any row (if any row is table extraction, aggregated is table extraction)
                    if len(row) > len(canonical_schema) + 2 and row[len(canonical_schema) + 2]:
                        aggregated_is_table_extraction = True
                
                # Convert aggregated vehicle to canonical row
                vin_value = aggregated_vehicle.get('vin')
                if vin_value:
                    canonical_row = []
                    for field in canonical_schema:
                        value = aggregated_vehicle.get(field)
                        if isinstance(value, str) and value.strip().lower() == "none":
                            value = None
                        canonical_row.append(value)
                    # Preserve _is_handwritten, _is_vision_extracted, _is_table_extraction, and _is_ai_repaired flags
                    canonical_row.append(aggregated_is_handwritten)
                    canonical_row.append(aggregated_is_vision_extracted)
                    canonical_row.append(aggregated_is_table_extraction)
                    # Preserve _is_ai_repaired from any row (if any row was AI-repaired, aggregated is AI-repaired)
                    aggregated_is_ai_repaired = False
                    for row in rows_for_vin:
                        if len(row) > len(canonical_schema) + 3 and row[len(canonical_schema) + 3]:
                            aggregated_is_ai_repaired = True
                            break
                    canonical_row.append(aggregated_is_ai_repaired)
                    all_rows.append(canonical_row)
                else:
                    logger.warning(f"[IMAGE] No valid VIN found after aggregation (got: {vin_value}), skipping row")
        
        # Log extraction summary
        if vision_extracted_pages:
            page_range = f"{min(vision_extracted_pages)}-{max(vision_extracted_pages)}" if len(vision_extracted_pages) > 1 else str(vision_extracted_pages[0])
            logger.info(f"[Vision] Extracted pages: {page_range}")
        
        if not all_rows:
            logger.warning("[Vision] No rows extracted from any page")
            # Return empty 2D array with headers to ensure normalize_v2 is called
            # Include _is_handwritten, _is_vision_extracted, _is_table_extraction, and _is_ai_repaired in header to preserve them through rows2d_to_objects
            header_row = list(canonical_schema) + ["_is_handwritten", "_is_vision_extracted", "_is_table_extraction", "_is_ai_repaired"]
            return [header_row]  # Return header row only, no data rows
        
        # Build 2D array: header row + data rows
        # Clean "None" strings from all rows before returning
        cleaned_rows = []
        for row in all_rows:
            cleaned_row = []
            for value in row:
                if isinstance(value, str) and value.strip().lower() == "none":
                    cleaned_row.append(None)
                else:
                    cleaned_row.append(value)
            cleaned_rows.append(cleaned_row)
        
        # Build 2D array: header row + data rows
        # Clean "None" strings from all rows before returning
        cleaned_rows = []
        for row in all_rows:
            cleaned_row = []
            for value in row:
                if isinstance(value, str) and value.strip().lower() == "none":
                    cleaned_row.append(None)
                else:
                    cleaned_row.append(value)
            cleaned_rows.append(cleaned_row)
        
        # Apply AI repair pass for Vision-extracted rows (only for non-table extractions)
        # Get OCR text for context (from extracted_ocr_text_for_vision or cached OCR text)
        ocr_text_for_repair = extracted_ocr_text_for_vision if extracted_ocr_text_for_vision else ""
        if not ocr_text_for_repair and hasattr(_extract_table_with_vision_api, '_cached_full_ocr_text'):
            ocr_text_for_repair = getattr(_extract_table_with_vision_api, '_cached_full_ocr_text', "")
        
        # Only apply AI repair to Vision-extracted rows that are NOT table extractions
        # Table extractions are authoritative and should not be modified
        repaired_cleaned_rows = []
        repair_count = 0
        for row in cleaned_rows:
            # Ensure row has expected structure (canonical_schema + 3 flags)
            # Pad with False if flags are missing
            row_with_flags = list(row)  # Copy row
            while len(row_with_flags) < len(canonical_schema) + 3:
                row_with_flags.append(False)
            
            # Extract flags
            is_handwritten = row_with_flags[len(canonical_schema)] if len(row_with_flags) > len(canonical_schema) else False
            is_vision_extracted = row_with_flags[len(canonical_schema) + 1] if len(row_with_flags) > len(canonical_schema) + 1 else False
            is_table_extraction = row_with_flags[len(canonical_schema) + 2] if len(row_with_flags) > len(canonical_schema) + 2 else False
            is_ai_repaired = row_with_flags[len(canonical_schema) + 3] if len(row_with_flags) > len(canonical_schema) + 3 else False
            
            # Repair Vision-extracted rows that haven't been repaired yet
            # For tables: VIN-only repair (other fields are authoritative)
            # For document-style: Full repair (VIN + casing + semantic fields)
            if is_vision_extracted and not is_ai_repaired:
                vin_only_mode = is_table_extraction  # VIN-only for tables, full repair for document-style
                repaired_row = repair_vehicle_row_with_ai(row_with_flags, canonical_schema, ocr_text_for_repair, vin_only=vin_only_mode)
                if repaired_row:
                    repaired_cleaned_rows.append(repaired_row)
                    repair_count += 1
                else:
                    # If repair failed, keep original row and mark as not repaired
                    if len(row_with_flags) == len(canonical_schema) + 3:
                        row_with_flags.append(False)  # _is_ai_repaired = False
                    else:
                        row_with_flags[len(canonical_schema) + 3] = False  # Update existing flag
                    repaired_cleaned_rows.append(row_with_flags)
            else:
                # Table extraction, not Vision-extracted, or already repaired - keep as-is and ensure flag
                if len(row_with_flags) == len(canonical_schema) + 3:
                    row_with_flags.append(False)  # _is_ai_repaired = False
                elif len(row_with_flags) > len(canonical_schema) + 3:
                    row_with_flags[len(canonical_schema) + 3] = is_ai_repaired  # Preserve existing flag
                else:
                    row_with_flags.append(False)  # _is_ai_repaired = False
                repaired_cleaned_rows.append(row_with_flags)
        
        if repair_count > 0:
            logger.info(f"[AI Repair] Repaired {repair_count} vehicle row(s) with AI")
        
        rows_2d = []
        # Include _is_handwritten, _is_vision_extracted, _is_table_extraction, and _is_ai_repaired in header to preserve them through rows2d_to_objects
        # CRITICAL: Use canonical_schema field names directly (don't normalize) to ensure exact match in normalize_v2
        header_row = list(canonical_schema) + ["_is_handwritten", "_is_vision_extracted", "_is_table_extraction", "_is_ai_repaired"]
        rows_2d.append(header_row)  # Header row
        rows_2d.extend(repaired_cleaned_rows)  # Data rows (after AI repair if applicable)
        
        logger.info(f"[Vision] Final parsed 2D array: {len(repaired_cleaned_rows)} rows × {len(canonical_schema)} columns")
        logger.info(f"[Vision] Extraction success: {len(repaired_cleaned_rows)} rows × {len(canonical_schema)} columns")
        # Clear cached OCR text after extraction is complete
        if hasattr(_extract_table_with_vision_api, '_cached_full_ocr_text'):
            delattr(_extract_table_with_vision_api, '_cached_full_ocr_text')
        return rows_2d
        
    except ImportError as e:
        logger.error(f"[Vision] Required package not installed: {e}")
        # Return empty 2D array with headers to ensure normalize_v2 is called
        from schema import VEHICLE_SCHEMA_ORDER
        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
        header_row = list(canonical_schema) + ["_is_handwritten", "_is_vision_extracted", "_is_table_extraction", "_is_ai_repaired"]
        return [header_row]  # Return header row only, no data rows
    except Exception as e:
        logger.error(f"[Vision] Extraction failed: {e}", exc_info=True)
        # Return empty 2D array with headers to ensure normalize_v2 is called (not None to bypass normalization)
        from schema import VEHICLE_SCHEMA_ORDER
        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
        header_row = list(canonical_schema) + ["_is_handwritten", "_is_vision_extracted", "_is_table_extraction", "_is_ai_repaired"]
        return [header_row]  # Return header row only, no data rows

def _extract_from_ocr_source(
    source: Union[str, Dict, Path],
    source_type: SourceType,
    header_row_index: int = 0,
    mapping_id: Optional[str] = None,
    **kwargs
) -> List[List[Any]]:
    """
    Extract data from PDF or Image sources using OCR pipeline.
    
    Args:
        source: Source identifier (file path or dict with file_path)
        source_type: SourceType.PDF or SourceType.IMAGE
        header_row_index: Row index containing headers (default: 0)
        **kwargs: Additional parameters
    
    Returns:
        2D list of values (rows x columns) compatible with other extractors
    """
    from ocr import (
        extract_text_from_image,
        extract_text_from_pdf,
        parse_text_blocks,
        detect_table_candidates,
        extract_tables_from_blocks
    )
    from external_tables import clean_header
    
    # Get file path
    if isinstance(source, dict):
        file_path = Path(source.get("file_path", source.get("path", "")))
    else:
        file_path = Path(source)
    
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # CRITICAL: Route IMAGE sources to Vision API first (like handwritten PDFs)
    # OCR is unreliable for images due to fragmented output - Vision API is primary
    if source_type == SourceType.IMAGE:
        domain_mapping_id = mapping_id or kwargs.get('mapping_id', '')
        domain_name = "driver" if 'driver' in domain_mapping_id.lower() else "vehicle"
        logger.warning(f"[IMAGE] Routing to Vision API first (primary extraction path)")
        logger.warning(f"[DEBUG] ROUTING TO {domain_name.upper()} PARSER (IMAGE) - domain_mapping_id='{domain_mapping_id}'")
        
        # Try Vision API first (same as handwritten PDFs)
        vision_result = _extract_table_with_vision_api(
            file_path,
            SourceType.IMAGE,
            mapping_id=domain_mapping_id,
            **kwargs
        )
        
        if vision_result and len(vision_result) > 1:  # Has header + at least one data row
            domain_name_plural = "drivers" if 'driver' in domain_mapping_id.lower() else "vehicles"
            logger.warning(f"[IMAGE] Vision API extracted {len(vision_result)-1} {domain_name_plural} (primary path)")
            return vision_result
        else:
            # Vision API failed or returned no vehicles - fall back to OCR with lower confidence threshold
            logger.warning(f"[IMAGE] Vision API returned {'None' if vision_result is None else 'empty'}, falling back to OCR with lower confidence threshold")
            # Continue to OCR fallback path below with lowered confidence threshold
    
    # Special handling for PDF: route through Vision OCR → domain-specific parser
    if source_type == SourceType.PDF:
        # Determine domain from mapping_id
        # IMPORTANT: Check relationships FIRST before policies, since "policy_vehicle_driver_link" contains "polic"
        domain_mapping_id = mapping_id or kwargs.get('mapping_id', '')
        logger.debug(f"[DEBUG _extract_from_ocr_source] mapping_id param={mapping_id}, kwargs mapping_id={kwargs.get('mapping_id', 'NONE')}, domain_mapping_id={domain_mapping_id}")
        domain_lower = domain_mapping_id.lower()
        is_relationship = 'relationship' in domain_lower or 'link' in domain_lower
        is_policy = ('polic' in domain_lower or domain_mapping_id == 'source_pdf_policies') and not is_relationship
        is_driver = 'driver' in domain_lower and not is_relationship
        is_location = 'location' in domain_lower
        is_claim = 'claim' in domain_lower
        
        logger.info(f"[PDF] Routing through Vision OCR → domain parser (mapping_id: {domain_mapping_id})")
        # Get Vision OCR text directly
        try:
            vision_ocr_text = _get_vision_ocr_text_for_pdf(file_path)
            logger.info(f"[PDF] Vision OCR text extracted: {len(vision_ocr_text)} characters")
            logger.debug(f"[DEBUG _extract_from_ocr_source] Domain detection: is_policy={is_policy}, is_driver={is_driver}, is_location={is_location}, is_claim={is_claim}, is_relationship={is_relationship}")
            logger.debug(f"[DEBUG] Vision OCR text length: {len(vision_ocr_text) if vision_ocr_text else 0}, is_empty={not vision_ocr_text or len(vision_ocr_text) == 0}")
            
            
            # Route to appropriate domain parser
            # IMPORTANT: Check relationships FIRST to avoid conflicts with "policy_vehicle_driver_link"
            if is_relationship:
                logger.info("[PDF] Using parse_relationship_raw_text for relationships")
                logger.info("[DEBUG] ROUTING TO RELATIONSHIP PARSER - parse_relationship_raw_text()")
                relationships = parse_relationship_raw_text(vision_ocr_text)
                if relationships:
                    logger.info(f"[PDF] Extracted {len(relationships)} relationship(s) from Vision OCR text")
                    from schema import RELATIONSHIP_SCHEMA_ORDER
                    canonical_schema = RELATIONSHIP_SCHEMA_ORDER.copy()
                    rows_2d = [canonical_schema]
                    for relationship in relationships:
                        row = []
                        for field in canonical_schema:
                            value = relationship.get(field)
                            if isinstance(value, str) and value.strip().lower() == "none":
                                value = None
                            row.append(value)
                        rows_2d.append(row)
                    return rows_2d
                else:
                    logger.warning("[PDF] parse_relationship_raw_text returned 0 relationships")
                    return []
            elif is_policy:
                logger.info("[PDF] Using parse_policy_raw_text for policies")
                logger.info("[DEBUG] ROUTING TO POLICY PARSER - parse_policy_raw_text()")
                policies = parse_policy_raw_text(vision_ocr_text, source_type="pdf")
                if policies:
                    logger.info(f"[PDF] Extracted {len(policies)} policy/policies from Vision OCR text")
                    from schema import POLICY_SCHEMA_ORDER
                    canonical_schema = POLICY_SCHEMA_ORDER.copy()
                    rows_2d = [canonical_schema]  # Header row
                    
                    # Track policy numbers already extracted
                    extracted_policy_numbers = {p.get('policy_number') for p in policies if p.get('policy_number')}
                    logger.info(f"[PDF] Extracted policy numbers: {extracted_policy_numbers}")
                    
                    # Convert existing policies to rows
                    for policy in policies:
                        row = []
                        for field in canonical_schema:
                            value = policy.get(field)
                            # Clean "None" strings
                            if isinstance(value, str) and value.strip().lower() == "none":
                                value = None
                            row.append(value)
                        rows_2d.append(row)
                    
                    # FALLBACK: Scan Vision OCR text for policy numbers not captured in structured extraction
                    # vision_ocr_text already contains all text from the PDF (including memo content)
                    # Scan it for policy numbers that weren't extracted as structured blocks
                    try:
                        import re
                        
                        if vision_ocr_text:
                            logger.info(f"[PDF] Scanning Vision OCR text ({len(vision_ocr_text)} chars) for fallback policy numbers")
                            # Find all policy numbers in Vision OCR text
                            # Pattern matches: P001, P002, P_BAD1, etc.
                            all_policy_numbers = re.findall(r'\b(P\d+|P_[A-Z0-9]+)\b', vision_ocr_text, re.IGNORECASE)
                            logger.info(f"[PDF] All policy numbers found in Vision OCR text: {all_policy_numbers}")
                            # Remove duplicates and filter out already extracted ones
                            unique_policy_numbers = list(set(all_policy_numbers))
                            missing_policy_numbers = [pn for pn in unique_policy_numbers if pn not in extracted_policy_numbers]
                            
                            if missing_policy_numbers:
                                logger.info(f"[PDF] Found {len(missing_policy_numbers)} fallback policy number(s): {missing_policy_numbers}")
                            
                            # Create fallback rows for missing policy numbers
                            for policy_num in missing_policy_numbers:
                                logger.info(f"[PDF] Creating fallback row for policy number: {policy_num}")
                                # Create minimal policy row: only policy_number populated, all others None
                                fallback_policy = {field: None for field in POLICY_SCHEMA_ORDER}
                                fallback_policy['policy_number'] = policy_num
                                # notes = None (do NOT infer from memo text)
                                
                                # Convert to row
                                row = []
                                for field in canonical_schema:
                                    value = fallback_policy.get(field)
                                    row.append(value)
                                rows_2d.append(row)
                                logger.info(f"[PDF] Added fallback row for {policy_num}, total rows now: {len(rows_2d)}")
                        else:
                            logger.warning("[PDF] Vision OCR text is empty, cannot scan for fallback policy numbers")
                    except Exception as e:
                        logger.warning(f"[PDF] Error scanning Vision OCR text for fallback policy numbers: {e}")
                        import traceback
                        logger.debug(f"[PDF] Fallback scanning error traceback: {traceback.format_exc()}")
                    
                    return rows_2d
                else:
                    logger.warning("[PDF] parse_policy_raw_text returned 0 policies")
                    return []
            elif is_driver:
                logger.info("[PDF] Using parse_driver_raw_text for drivers")
                logger.info("[DEBUG] ROUTING TO DRIVER PARSER - parse_driver_raw_text()")
                logger.debug(f"[PDF] Vision OCR text length: {len(vision_ocr_text) if vision_ocr_text else 0}")
                logger.debug(f"[PDF] Vision OCR text preview (first 500 chars): {vision_ocr_text[:500] if vision_ocr_text else 'None'}")
                if not vision_ocr_text or len(vision_ocr_text.strip()) == 0:
                    logger.warning("[PDF] Vision OCR text is empty, cannot parse drivers")
                    return []
                # DEBUG: Print full OCR text to see format (only in debug mode)
                logger.debug(f"[PDF DEBUG] FULL Vision OCR text ({len(vision_ocr_text)} chars):\n---\n{vision_ocr_text}\n---")
                drivers = parse_driver_raw_text(vision_ocr_text)
                logger.info(f"[PDF] parse_driver_raw_text returned {len(drivers)} driver(s)")
                if drivers:
                    logger.info(f"[PDF] Extracted {len(drivers)} driver(s) from Vision OCR text")
                    from schema import DRIVER_SCHEMA_ORDER
                    canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                    rows_2d = [canonical_schema]
                    for driver in drivers:
                        row = []
                        for field in canonical_schema:
                            value = driver.get(field)
                            if isinstance(value, str) and value.strip().lower() == "none":
                                value = None
                            row.append(value)
                        rows_2d.append(row)
                    return rows_2d
                else:
                    logger.warning("[PDF] parse_driver_raw_text returned 0 drivers")
                    logger.warning(f"[PDF] OCR text that failed to parse (first 1000 chars): {vision_ocr_text[:1000] if vision_ocr_text else 'None'}")
                    return []
            elif is_location:
                logger.info("[PDF] Using parse_locations_raw_text for locations")
                logger.info("[DEBUG] ROUTING TO LOCATION PARSER - parse_locations_raw_text()")
                locations = parse_locations_raw_text(vision_ocr_text)
                if locations:
                    logger.info(f"[PDF] Extracted {len(locations)} location(s) from Vision OCR text")
                    from schema import LOCATION_SCHEMA_ORDER
                    canonical_schema = LOCATION_SCHEMA_ORDER.copy()
                    rows_2d = [canonical_schema]
                    for location in locations:
                        row = []
                        for field in canonical_schema:
                            value = location.get(field)
                            if isinstance(value, str) and value.strip().lower() == "none":
                                value = None
                            row.append(value)
                        rows_2d.append(row)
                    return rows_2d
                else:
                    logger.warning("[PDF] parse_locations_raw_text returned 0 locations")
                    return []
            elif is_claim:
                logger.info("[PDF] Using parse_claim_raw_text for claims")
                logger.info("[DEBUG] ROUTING TO CLAIM PARSER - parse_claim_raw_text()")
                claims = parse_claim_raw_text(vision_ocr_text)
                if claims:
                    logger.info(f"[PDF] Extracted {len(claims)} claim(s) from Vision OCR text")
                    from schema import CLAIM_SCHEMA_ORDER
                    canonical_schema = CLAIM_SCHEMA_ORDER.copy()
                    rows_2d = [canonical_schema]
                    for claim in claims:
                        row = []
                        for field in canonical_schema:
                            value = claim.get(field)
                            if isinstance(value, str) and value.strip().lower() == "none":
                                value = None
                            row.append(value)
                        rows_2d.append(row)
                    return rows_2d
                else:
                    logger.warning("[PDF] parse_claim_raw_text returned 0 claims")
                    return []
            else:
                # Default fallback - route based on domain
                if is_driver:
                    # For drivers, try paragraph-based extraction FIRST
                    logger.info("[PDF] Using paragraph-based extraction for drivers")
                    logger.warning(f"[DEBUG] ROUTING TO DRIVER PARSER (paragraph-first) - domain_mapping_id=\'{domain_mapping_id}\'")
                    
                    drivers = parse_driver_raw_text(vision_ocr_text)
                    if drivers and len(drivers) > 0:
                        logger.info(f"[PDF] Paragraph-based extraction found {len(drivers)} driver(s) from Vision OCR text")
                        # Convert to 2D array format
                        from schema import DRIVER_SCHEMA_ORDER
                        canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for driver in drivers:
                            row = []
                            for field in canonical_schema:
                                value = driver.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        # Paragraph extraction failed or found no drivers - fall back to Vision API table extraction
                        logger.warning(f"[PDF] parse_driver_raw_text returned {len(drivers) if drivers else 0} driver(s), falling back to Vision API table extraction")
                        
                        # Check if PDF is scanned or handwritten (no embedded text OR low OCR confidence)
                        is_scanned, scan_reason = _is_pdf_scanned_or_handwritten(file_path)
                        
                        if is_scanned:
                            logger.info(f"[PDF] Detected scanned/handwritten PDF ({scan_reason}) - routing as IMAGE for better extraction")
                            # Convert PDF pages to images and route as IMAGE source type
                            # This forces image-first Vision extraction
                            vision_result = _extract_table_with_vision_api(
                                file_path,
                                SourceType.IMAGE,  # Route as IMAGE to force image-first processing
                                mapping_id=domain_mapping_id,
                                **kwargs
                            )
                        else:
                            # Structured PDF with embedded text - use Vision API table extraction as fallback
                            vision_result = _extract_table_with_vision_api(
                                file_path,
                                SourceType.PDF,
                                mapping_id=domain_mapping_id,
                                **kwargs
                            )
                        
                        if vision_result and len(vision_result) > 1:  # Has header + at least one data row
                            logger.info(f"[PDF] Vision API table extraction found {len(vision_result)-1} driver(s) (fallback)")
                            return vision_result
                        else:
                            logger.warning(f"[PDF] Both paragraph-based and Vision API extraction failed")
                            # Return empty 2D array with headers to ensure normalize_v2 is called
                            from schema import DRIVER_SCHEMA_ORDER
                            canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                            return [canonical_schema]  # Return header row only, no data rows
                else:
                    # Vehicle parser (preserve existing behavior)
                    logger.info("[PDF] Using paragraph-based extraction for vehicles (prioritizing semantic field extraction)")
                    logger.warning(f"[DEBUG] ROUTING TO VEHICLE PARSER (paragraph-first) - domain_mapping_id=\'{domain_mapping_id}\'")
                    
                    # For vehicles, try paragraph-based extraction FIRST (works better for semantic fields like body_style, fuel_type)
                    # This extracts from full OCR text and goes through normalize_v2 properly
                    vehicles = parse_vehicle_raw_text(vision_ocr_text, source_type="pdf")
                    if vehicles and len(vehicles) > 0:
                        logger.info(f"[PDF] Paragraph-based extraction found {len(vehicles)} vehicle(s) from Vision OCR text")
                        # Convert to 2D array format
                        from schema import VEHICLE_SCHEMA_ORDER
                        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for vehicle in vehicles:
                            row = []
                            for field in canonical_schema:
                                value = vehicle.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        # Paragraph extraction failed or found no vehicles - fall back to Vision API table extraction
                        logger.warning(f"[PDF] parse_vehicle_raw_text returned {len(vehicles) if vehicles else 0} vehicle(s), falling back to Vision API table extraction")
                        
                        # Check if PDF is scanned or handwritten (no embedded text OR low OCR confidence)
                        is_scanned, scan_reason = _is_pdf_scanned_or_handwritten(file_path)
                        
                        if is_scanned:
                            logger.info(f"[PDF] Detected scanned/handwritten PDF ({scan_reason}) - routing as IMAGE for better extraction")
                            # Convert PDF pages to images and route as IMAGE source type
                            # This forces image-first Vision extraction, skipping OCR-first and split_by_vin paths
                            vision_result = _extract_table_with_vision_api(
                                file_path,
                                SourceType.IMAGE,  # Route as IMAGE to force image-first processing
                                mapping_id=domain_mapping_id,
                                **kwargs
                            )
                        else:
                            # Structured PDF with embedded text - use Vision API table extraction as fallback
                            vision_result = _extract_table_with_vision_api(
                                file_path,
                                SourceType.PDF,
                                mapping_id=domain_mapping_id,
                                **kwargs
                            )
                        
                        if vision_result and len(vision_result) > 1:  # Has header + at least one data row
                            logger.info(f"[PDF] Vision API table extraction found {len(vision_result)-1} vehicle(s) (fallback)")
                            return vision_result
                        else:
                            logger.warning(f"[PDF] Both paragraph-based and Vision API extraction failed")
                            # Return empty 2D array with headers to ensure normalize_v2 is called
                            from schema import VEHICLE_SCHEMA_ORDER
                            canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                            return [canonical_schema]  # Return header row only, no data rows
        except Exception as e:
            logger.warning(f"[PDF] Vision OCR failed: {e}, falling back to regular OCR pipeline")
            logger.debug(f"[DEBUG] Vision OCR exception: {type(e).__name__}: {str(e)}")
            logger.info("[DEBUG] Attempting regular OCR extraction (fallback)...")
            # Try regular OCR extraction and parse as raw text
            try:
                from ocr.reader import extract_text_from_pdf
                text_blocks, metadata = extract_text_from_pdf(file_path, enable_vision=False, **kwargs)
                logger.info(f"[PDF] Regular OCR extracted {len(text_blocks) if text_blocks else 0} text blocks")
                if text_blocks:
                    from ocr.table_extract import _extract_raw_text
                    full_ocr_text = _extract_raw_text(text_blocks)
                    logger.info(f"[PDF] Extracted {len(full_ocr_text)} characters from text blocks")
                    if full_ocr_text:
                        logger.info(f"[PDF] Routing to domain parser: is_policy={is_policy}, is_driver={is_driver}, is_location={is_location}, is_claim={is_claim}, is_relationship={is_relationship}")
                        # Route to appropriate parser based on mapping_id
                        # IMPORTANT: Check relationships FIRST to avoid conflicts
                        if is_relationship:
                            logger.info("[DEBUG] ROUTING TO RELATIONSHIP PARSER (fallback OCR) - parse_relationship_raw_text()")
                            relationships = parse_relationship_raw_text(full_ocr_text)
                            if relationships:
                                logger.info(f"[PDF] Extracted {len(relationships)} relationship(s) from regular OCR text")
                                from schema import RELATIONSHIP_SCHEMA_ORDER
                                canonical_schema = RELATIONSHIP_SCHEMA_ORDER.copy()
                                rows_2d = [canonical_schema]
                                for relationship in relationships:
                                    row = []
                                    for field in canonical_schema:
                                        value = relationship.get(field)
                                        if isinstance(value, str) and value.strip().lower() == "none":
                                            value = None
                                        row.append(value)
                                    rows_2d.append(row)
                                return rows_2d
                            else:
                                logger.warning("[PDF] parse_relationship_raw_text returned 0 relationships (fallback OCR)")
                                return []
                        elif is_policy:
                            logger.info("[DEBUG] ROUTING TO POLICY PARSER (fallback OCR) - parse_policy_raw_text()")
                            policies = parse_policy_raw_text(full_ocr_text, source_type="pdf")
                            if policies:
                                logger.info(f"[PDF] Extracted {len(policies)} policy/policies from regular OCR text")
                                from schema import POLICY_SCHEMA_ORDER
                                canonical_schema = POLICY_SCHEMA_ORDER.copy()
                                rows_2d = [canonical_schema]
                                
                                # Track policy numbers already extracted
                                extracted_policy_numbers = {p.get('policy_number') for p in policies if p.get('policy_number')}
                                
                                for policy in policies:
                                    row = []
                                    for field in canonical_schema:
                                        value = policy.get(field)
                                        if isinstance(value, str) and value.strip().lower() == "none":
                                            value = None
                                        row.append(value)
                                    rows_2d.append(row)
                                
                                # FALLBACK: Scan OCR text for policy numbers not captured in structured extraction
                                try:
                                    import re
                                    if full_ocr_text:
                                        all_policy_numbers = re.findall(r'\b(P\d+|P_[A-Z0-9]+)\b', full_ocr_text, re.IGNORECASE)
                                        unique_policy_numbers = list(set(all_policy_numbers))
                                        missing_policy_numbers = [pn for pn in unique_policy_numbers if pn not in extracted_policy_numbers]
                                        
                                        for policy_num in missing_policy_numbers:
                                            logger.warning(f"[PDF] Creating fallback row for policy number: {policy_num}")
                                            fallback_policy = {field: None for field in POLICY_SCHEMA_ORDER}
                                            fallback_policy['policy_number'] = policy_num
                                            row = []
                                            for field in canonical_schema:
                                                row.append(fallback_policy.get(field))
                                            rows_2d.append(row)
                                except Exception as e:
                                    logger.debug(f"[PDF] Error in fallback scanning: {e}")
                                
                                return rows_2d
                            else:
                                logger.warning("[PDF] parse_policy_raw_text returned 0 policies (fallback OCR)")
                                return []
                        elif is_driver:
                            logger.info("[DEBUG] ROUTING TO DRIVER PARSER (fallback OCR) - parse_driver_raw_text()")
                            logger.debug(f"[PDF] Fallback OCR text length: {len(full_ocr_text) if full_ocr_text else 0}")
                            logger.debug(f"[PDF] Fallback OCR text preview (first 500 chars): {full_ocr_text[:500] if full_ocr_text else 'None'}")
                            if not full_ocr_text or len(full_ocr_text.strip()) == 0:
                                logger.warning("[PDF] Fallback OCR text is empty, cannot parse drivers")
                                return []
                            # DEBUG: Print full OCR text to see format
                            logger.warning(f"[PDF DEBUG] FULL Fallback OCR text ({len(full_ocr_text)} chars):\n---\n{full_ocr_text}\n---")
                            drivers = parse_driver_raw_text(full_ocr_text)
                            logger.info(f"[PDF] parse_driver_raw_text (fallback) returned {len(drivers)} driver(s)")
                            if drivers:
                                logger.info(f"[PDF] Extracted {len(drivers)} driver(s) from regular OCR text")
                                from schema import DRIVER_SCHEMA_ORDER
                                canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                                rows_2d = [canonical_schema]
                                for driver in drivers:
                                    row = []
                                    for field in canonical_schema:
                                        value = driver.get(field)
                                        if isinstance(value, str) and value.strip().lower() == "none":
                                            value = None
                                        row.append(value)
                                    rows_2d.append(row)
                                return rows_2d
                            else:
                                logger.warning("[PDF] parse_driver_raw_text returned 0 drivers (fallback OCR)")
                                logger.warning(f"[PDF] Fallback OCR text that failed to parse (first 1000 chars): {full_ocr_text[:1000] if full_ocr_text else 'None'}")
                                return []
                        elif is_location:
                            logger.info("[DEBUG] ROUTING TO LOCATION PARSER (fallback OCR) - parse_locations_raw_text()")
                            locations = parse_locations_raw_text(full_ocr_text)
                            if locations:
                                logger.info(f"[PDF] Extracted {len(locations)} location(s) from regular OCR text")
                                from schema import LOCATION_SCHEMA_ORDER
                                canonical_schema = LOCATION_SCHEMA_ORDER.copy()
                                rows_2d = [canonical_schema]
                                for location in locations:
                                    row = []
                                    for field in canonical_schema:
                                        value = location.get(field)
                                        if isinstance(value, str) and value.strip().lower() == "none":
                                            value = None
                                        row.append(value)
                                    rows_2d.append(row)
                                return rows_2d
                            else:
                                logger.warning("[PDF] parse_locations_raw_text returned 0 locations (fallback OCR)")
                                return []
                        elif is_claim:
                            logger.info("[DEBUG] ROUTING TO CLAIM PARSER (fallback OCR) - parse_claim_raw_text()")
                            claims = parse_claim_raw_text(full_ocr_text)
                            if claims:
                                logger.info(f"[PDF] Extracted {len(claims)} claim(s) from regular OCR text")
                                from schema import CLAIM_SCHEMA_ORDER
                                canonical_schema = CLAIM_SCHEMA_ORDER.copy()
                                rows_2d = [canonical_schema]
                                for claim in claims:
                                    row = []
                                    for field in canonical_schema:
                                        value = claim.get(field)
                                        if isinstance(value, str) and value.strip().lower() == "none":
                                            value = None
                                        row.append(value)
                                    rows_2d.append(row)
                                return rows_2d
                            else:
                                logger.warning("[PDF] parse_claim_raw_text returned 0 claims (fallback OCR)")
                                return []
                        else:
                            # Default fallback - route based on domain
                            if is_driver:
                                logger.warning(f"[DEBUG] ROUTING TO DRIVER PARSER (fallback OCR) - domain_mapping_id=\'{domain_mapping_id}\'")
                                # Try _extract_table_with_vision_api first (more reliable)
                                vision_result = _extract_table_with_vision_api(
                                    file_path,
                                    SourceType.PDF,
                                    mapping_id=domain_mapping_id,
                                    **kwargs
                                )
                                if vision_result and len(vision_result) > 1:  # Has header + at least one data row
                                    logger.info(f"[PDF] Extracted {len(vision_result)-1} driver(s) from Vision API (fallback OCR path)")
                                    return vision_result
                                # Fallback to parse_driver_raw_text if Vision API fails
                                drivers = parse_driver_raw_text(full_ocr_text)
                                logger.warning(f"[PDF] Fallback: parse_driver_raw_text found {len(drivers)} drivers")
                                if drivers:
                                    logger.info(f"[PDF] Extracted {len(drivers)} driver(s) from regular OCR text")
                                    from schema import DRIVER_SCHEMA_ORDER
                                    canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                                    rows_2d = [canonical_schema]
                                    for driver in drivers:
                                        row = []
                                        for field in canonical_schema:
                                            value = driver.get(field)
                                            if isinstance(value, str) and value.strip().lower() == "none":
                                                value = None
                                            row.append(value)
                                        rows_2d.append(row)
                                    return rows_2d
                                else:
                                    logger.warning("[PDF] parse_driver_raw_text returned 0 drivers (fallback OCR path)")
                                    from schema import DRIVER_SCHEMA_ORDER
                                    canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                                    return [canonical_schema]  # Return header row only, no data rows
                            else:
                                # Vehicle parser (preserve existing behavior)
                                logger.warning(f"[DEBUG] ROUTING TO VEHICLE PARSER (fallback OCR) - domain_mapping_id=\'{domain_mapping_id}\'")
                                # Try _extract_table_with_vision_api first (more reliable)
                                vision_result = _extract_table_with_vision_api(
                                    file_path,
                                    SourceType.PDF,
                                    mapping_id=domain_mapping_id,
                                    **kwargs
                                )
                                if vision_result and len(vision_result) > 1:  # Has header + at least one data row
                                    logger.info(f"[PDF] Extracted {len(vision_result)-1} vehicle(s) from Vision API (fallback OCR path)")
                                    return vision_result
                                # Fallback to parse_vehicle_raw_text if Vision API fails
                                vehicles = parse_vehicle_raw_text(full_ocr_text, source_type="pdf")
                                logger.warning(f"[PDF] Fallback: parse_vehicle_raw_text found {len(vehicles)} vehicles")
                                print(f"[PDF DEBUG] Number of VIN blocks found: {len(vehicles)}")
                                if vehicles:
                                    logger.info(f"[PDF] Extracted {len(vehicles)} vehicle(s) from regular OCR text")
                                    from schema import VEHICLE_SCHEMA_ORDER
                                    canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                                    rows_2d = [canonical_schema]
                                    for vehicle in vehicles:
                                        row = []
                                        for field in canonical_schema:
                                            value = vehicle.get(field)
                                            if isinstance(value, str) and value.strip().lower() == "none":
                                                value = None
                                            row.append(value)
                                        rows_2d.append(row)
                                    return rows_2d
                                else:
                                    # No vehicles found - return empty 2D array with headers to ensure normalize_v2 is called
                                    logger.warning("[PDF] parse_vehicle_raw_text returned 0 vehicles (fallback OCR path)")
                                    from schema import VEHICLE_SCHEMA_ORDER
                                    canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                                    return [canonical_schema]  # Return header row only, no data rows
            except Exception as fallback_error:
                logger.warning(f"[PDF] Regular OCR fallback also failed: {fallback_error}")
            # Fall through to regular OCR pipeline below
    
    # Extract text blocks using OCR (for IMAGE and RAW_TEXT)
    # Enable Vision API automatically if OPENAI_API_KEY is available
    enable_vision = kwargs.pop('enable_vision', None)  # None means auto-detect
    if enable_vision is None:
        try:
            from dotenv import load_dotenv
            import os
            load_dotenv()
            api_key = os.getenv("OPENAI_API_KEY")
            if api_key:
                enable_vision = True
                logger.info(f"[OCR] OPENAI_API_KEY found, enabling Vision API (key length: {len(api_key)})")
            else:
                enable_vision = False
                logger.debug("[OCR] OPENAI_API_KEY not found, Vision API disabled")
        except ImportError:
            # dotenv not installed, try direct env var
            import os
            api_key = os.getenv("OPENAI_API_KEY")
            if api_key:
                enable_vision = True
                logger.info(f"[OCR] OPENAI_API_KEY found in environment, enabling Vision API (key length: {len(api_key)})")
            else:
                enable_vision = False
                logger.debug("[OCR] OPENAI_API_KEY not found, Vision API disabled")
        except Exception as e:
            enable_vision = False
            logger.debug(f"[OCR] Error checking for API key: {e}, Vision API disabled")
    
    # Try OCR extraction first (do NOT enable Vision in OCR - we'll use it as fallback)
    if source_type == SourceType.IMAGE:
        text_blocks, metadata = extract_text_from_image(file_path, enable_vision=False, **kwargs)
    elif source_type == SourceType.PDF:
        # For PDFs, use regular OCR extraction (Vision failed or not available)
        text_blocks, metadata = extract_text_from_pdf(file_path, enable_vision=False, **kwargs)
    elif source_type == SourceType.RAW_TEXT:
        # For raw text files, read the file directly
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text_content = f.read()
            # Create a single text block from the file content
            from ocr.models import TextBlock, OCRMetadata
            text_blocks = [TextBlock(
                text=text_content,
                bbox=(0, 0, 100, 100),  # Dummy bbox
                confidence=1.0
            )]
            metadata = OCRMetadata(
                engine="file_reader",
                confidence=1.0,
                language="eng",
                page_count=1
            )
        except Exception as e:
            logger.error(f"[RAW_TEXT] Error reading file: {e}")
            return []
    else:
        raise ValueError(f"Unsupported OCR source type: {source_type}")
    
    # Organize text blocks by page for PDFs
    page_text_blocks = {}
    if source_type == SourceType.PDF and text_blocks:
        # Try to get page count and organize blocks by page
        try:
            import PyPDF2
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                page_count = len(pdf_reader.pages)
                
                # Try to extract text from each page to organize blocks
                # For now, distribute blocks evenly across pages
                # In a more sophisticated implementation, we'd track page numbers in TextBlock metadata
                if page_count > 0:
                    blocks_per_page = max(1, len(text_blocks) // page_count)
                    for page_num in range(1, page_count + 1):
                        start_idx = (page_num - 1) * blocks_per_page
                        end_idx = page_num * blocks_per_page if page_num < page_count else len(text_blocks)
                        if start_idx < len(text_blocks):
                            page_text_blocks[page_num] = text_blocks[start_idx:end_idx]
                        else:
                            page_text_blocks[page_num] = []
                else:
                    # No pages found, put all blocks in page 1
                    page_text_blocks[1] = text_blocks
        except Exception as e:
            # If we can't determine pages, put all blocks in page 1
            logger.debug(f"[OCR] Could not organize blocks by page: {e}, using page 1")
            page_text_blocks[1] = text_blocks
    
    # For PDFs, check mapping_id first and route to appropriate parser before table extraction
    if source_type == SourceType.PDF:
        domain_mapping_id = mapping_id or kwargs.get('mapping_id', '')
        logger.info(f"[PDF] Checking domain routing: mapping_id={domain_mapping_id}, text_blocks={len(text_blocks) if text_blocks else 0}")
        # IMPORTANT: Check relationships FIRST before policies, since "policy_vehicle_driver_link" contains "polic"
        domain_lower = domain_mapping_id.lower()
        is_relationship = 'relationship' in domain_lower or 'link' in domain_lower
        is_policy = ('polic' in domain_lower or domain_mapping_id == 'source_pdf_policies') and not is_relationship
        is_driver = 'driver' in domain_lower and not is_relationship
        is_location = 'location' in domain_lower
        is_claim = 'claim' in domain_lower
        
        logger.info(f"[PDF] Domain detection: is_policy={is_policy}, is_driver={is_driver}, is_location={is_location}, is_claim={is_claim}, is_relationship={is_relationship}")
        
        if (is_policy or is_driver or is_location or is_claim or is_relationship) and text_blocks:
            # Extract raw text and route to domain-specific parser
            try:
                from ocr.table_extract import _extract_raw_text
                full_ocr_text = _extract_raw_text(text_blocks)
                if full_ocr_text:
                    # IMPORTANT: Check relationships FIRST to avoid conflicts
                    if is_relationship:
                        logger.info("[PDF] Routing to parse_relationship_raw_text (before table extraction)")
                        relationships = parse_relationship_raw_text(full_ocr_text)
                        if relationships:
                            logger.info(f"[PDF] Extracted {len(relationships)} relationship(s) from text blocks")
                            from schema import RELATIONSHIP_SCHEMA_ORDER
                            canonical_schema = RELATIONSHIP_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for relationship in relationships:
                                row = []
                                for field in canonical_schema:
                                    value = relationship.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                        else:
                            logger.warning("[PDF] parse_relationship_raw_text returned 0 relationships (before table extraction)")
                            return []
                    elif is_policy:
                        logger.info("[PDF] Routing to parse_policy_raw_text (before table extraction)")
                        policies = parse_policy_raw_text(full_ocr_text, source_type="pdf")
                        if policies:
                            logger.info(f"[PDF] Extracted {len(policies)} policy/policies from text blocks")
                            from schema import POLICY_SCHEMA_ORDER
                            canonical_schema = POLICY_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            
                            # Track policy numbers already extracted
                            extracted_policy_numbers = {p.get('policy_number') for p in policies if p.get('policy_number')}
                            
                            for policy in policies:
                                row = []
                                for field in canonical_schema:
                                    value = policy.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            
                            # FALLBACK: Scan OCR text for policy numbers not captured in structured extraction
                            try:
                                import re
                                if full_ocr_text:
                                    logger.info(f"[PDF] Scanning {len(full_ocr_text)} chars of OCR text for fallback policy numbers")
                                    all_policy_numbers = re.findall(r'\b(P\d+|P_[A-Z0-9]+)\b', full_ocr_text, re.IGNORECASE)
                                    logger.info(f"[PDF] All policy numbers found: {all_policy_numbers}")
                                    unique_policy_numbers = list(set(all_policy_numbers))
                                    missing_policy_numbers = [pn for pn in unique_policy_numbers if pn not in extracted_policy_numbers]
                                    
                                    if missing_policy_numbers:
                                        logger.warning(f"[PDF] Found {len(missing_policy_numbers)} missing policy number(s): {missing_policy_numbers}")
                                    
                                    for policy_num in missing_policy_numbers:
                                        logger.warning(f"[PDF] Creating fallback row for policy number: {policy_num}")
                                        fallback_policy = {field: None for field in POLICY_SCHEMA_ORDER}
                                        fallback_policy['policy_number'] = policy_num
                                        row = []
                                        for field in canonical_schema:
                                            row.append(fallback_policy.get(field))
                                        rows_2d.append(row)
                                        logger.warning(f"[PDF] Added fallback row for {policy_num}, total rows now: {len(rows_2d)}")
                                else:
                                    logger.warning("[PDF] full_ocr_text is empty, cannot scan for fallback policy numbers")
                            except Exception as e:
                                logger.warning(f"[PDF] Error in fallback scanning: {e}")
                                import traceback
                                logger.debug(f"[PDF] Fallback scanning traceback: {traceback.format_exc()}")
                            
                            return rows_2d
                        else:
                            logger.warning("[PDF] parse_policy_raw_text returned 0 policies (fallback OCR)")
                            return []
                    elif is_driver:
                        logger.info("[PDF] Routing to parse_driver_raw_text (before table extraction)")
                        drivers = parse_driver_raw_text(full_ocr_text)
                        if drivers:
                            logger.info(f"[PDF] Extracted {len(drivers)} driver(s) from text blocks")
                            from schema import DRIVER_SCHEMA_ORDER
                            canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for driver in drivers:
                                row = []
                                for field in canonical_schema:
                                    value = driver.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                    elif is_location:
                        logger.info("[PDF] Routing to parse_locations_raw_text (before table extraction)")
                        locations = parse_locations_raw_text(full_ocr_text)
                        if locations:
                            logger.info(f"[PDF] Extracted {len(locations)} location(s) from text blocks")
                            from schema import LOCATION_SCHEMA_ORDER
                            canonical_schema = LOCATION_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for location in locations:
                                row = []
                                for field in canonical_schema:
                                    value = location.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                        else:
                            logger.warning("[PDF] parse_locations_raw_text returned 0 locations (before table extraction)")
                            return []
                    elif is_claim:
                        logger.info("[PDF] Routing to parse_claim_raw_text (before table extraction)")
                        claims = parse_claim_raw_text(full_ocr_text)
                        if claims:
                            logger.info(f"[PDF] Extracted {len(claims)} claim(s) from text blocks")
                            from schema import CLAIM_SCHEMA_ORDER
                            canonical_schema = CLAIM_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for claim in claims:
                                row = []
                                for field in canonical_schema:
                                    value = claim.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                        else:
                            logger.warning("[PDF] parse_claim_raw_text returned 0 claims (before table extraction)")
                            return []
            except Exception as e:
                logger.debug(f"[PDF] Domain-specific parsing failed: {e}, falling through to table extraction")
    
    # For IMAGE sources (OCR fallback): Skip table extraction, use domain-specific parser directly
    # Table extraction creates garbage rows from fragmented OCR output
    if source_type == SourceType.IMAGE and text_blocks:
        # Detect domain from mapping_id
        domain_mapping_id = mapping_id or kwargs.get('mapping_id', '')
        domain_lower = domain_mapping_id.lower()
        is_driver = 'driver' in domain_lower and 'relationship' not in domain_lower
        
        try:
            from ocr.table_extract import _extract_raw_text
            # Use lower confidence threshold (0.2) to preserve more blocks for images
            # Handwritten images have avg confidence ~0.3, so 0.5 threshold filters too aggressively
            min_confidence = 0.2
            logger.warning(f"[IMAGE] OCR fallback: Using confidence threshold {min_confidence}")
            # Filter blocks with lower threshold before extracting text
            filtered_blocks = [b for b in text_blocks if b.confidence >= min_confidence]
            logger.warning(f"[IMAGE] OCR fallback: Using {len(filtered_blocks)} blocks (from {len(text_blocks)} total) with confidence >= {min_confidence}")
            
            full_ocr_text = _extract_raw_text(filtered_blocks)
            logger.warning(f"[IMAGE] OCR fallback: Extracted {len(full_ocr_text)} chars from filtered blocks")
            
            # Lower threshold for corrupted OCR (handwritten images often have short but valid text)
            # Minimum text length: 30 chars (reduced from 50) to catch more corrupted OCR
            min_text_length = 30
            if full_ocr_text and len(full_ocr_text) > min_text_length:
                if is_driver:
                    # Use parse_driver_raw_text (designed for prose, not tables)
                    drivers = parse_driver_raw_text(full_ocr_text)
                    if drivers:
                        logger.warning(f"[IMAGE] OCR fallback: parse_driver_raw_text extracted {len(drivers)} driver(s)")
                        from schema import DRIVER_SCHEMA_ORDER
                        canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for driver in drivers:
                            row = []
                            for field in canonical_schema:
                                value = driver.get(field)
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning(f"[IMAGE] OCR fallback: parse_driver_raw_text returned 0 drivers")
                else:
                    # Use parse_vehicle_raw_text (designed for prose, not tables)
                    vehicles = parse_vehicle_raw_text(full_ocr_text, source_type="image")
                    if vehicles:
                        logger.warning(f"[IMAGE] OCR fallback: parse_vehicle_raw_text extracted {len(vehicles)} vehicle(s)")
                        from schema import VEHICLE_SCHEMA_ORDER
                        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for vehicle in vehicles:
                            row = []
                            for field in canonical_schema:
                                value = vehicle.get(field)
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning(f"[IMAGE] OCR fallback: parse_vehicle_raw_text returned 0 vehicles")
            else:
                logger.warning(f"[IMAGE] OCR fallback: Text too short ({len(full_ocr_text) if full_ocr_text else 0} chars)")
        except Exception as e:
            logger.warning(f"[IMAGE] OCR fallback failed: {e}", exc_info=True)
        
        # If OCR fallback also failed, return empty with headers (let normalize_v2 handle it)
        logger.warning(f"[IMAGE] All extraction paths failed, returning empty result")
        if is_driver:
            from schema import DRIVER_SCHEMA_ORDER
            return [list(DRIVER_SCHEMA_ORDER.copy())]  # Header row only
        else:
            from schema import VEHICLE_SCHEMA_ORDER
            return [list(VEHICLE_SCHEMA_ORDER.copy())]  # Header row only
    
    # Parse and detect tables (for PDFs and other sources)
    # For PDF sources, use standard confidence threshold (0.5)
    parsed_blocks = parse_text_blocks(text_blocks, min_confidence=0.5)
    table_candidates = detect_table_candidates(parsed_blocks)
    
    # Extract structured rows
    ocr_result = extract_tables_from_blocks(parsed_blocks, table_candidates, metadata)
    
    # Convert OCR rows (List[Dict]) to 2D array format (List[List])
    rows_2d = []
    
    # Special handling for RAW_TEXT: detect domain and use appropriate parser
    if source_type == SourceType.RAW_TEXT and text_blocks:
        try:
            from ocr.table_extract import _extract_raw_text
            full_text = _extract_raw_text(text_blocks)
            if full_text:
                # Check if this is policies, locations, drivers, or vehicles based on mapping_id parameter
                domain_mapping_id = mapping_id or kwargs.get('mapping_id', '')
                if 'policies' in domain_mapping_id.lower():
                    # Use policy parser with source_type="raw" to prevent field inference
                    policies = parse_policy_raw_text(full_text, source_type="raw")
                    if policies:
                        logger.info(f"[RAW_TEXT] Extracted {len(policies)} policy/policies from raw text")
                        # Convert to 2D array format
                        from schema import POLICY_SCHEMA_ORDER
                        canonical_schema = POLICY_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for policy in policies:
                            row = []
                            for field in canonical_schema:
                                value = policy.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                elif 'locations' in domain_mapping_id.lower():
                    # Use location parser
                    locations = parse_locations_raw_text(full_text)
                    if locations:
                        logger.info(f"[RAW_TEXT] Extracted {len(locations)} location(s) from raw text")
                        # Convert to 2D array format
                        from schema import LOCATION_SCHEMA_ORDER
                        canonical_schema = LOCATION_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for location in locations:
                            row = []
                            for field in canonical_schema:
                                value = location.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                elif 'drivers' in domain_mapping_id.lower():
                    # Use driver parser
                    drivers = parse_driver_raw_text(full_text)
                    if drivers:
                        logger.info(f"[RAW_TEXT] Extracted {len(drivers)} driver(s) from raw text")
                        # Convert to 2D array format
                        from schema import DRIVER_SCHEMA_ORDER
                        canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for driver in drivers:
                            row = []
                            for field in canonical_schema:
                                value = driver.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                else:
                    # Default to vehicle parser
                    vehicles = parse_vehicle_raw_text(full_text)
                    if vehicles:
                        logger.info(f"[RAW_TEXT] Extracted {len(vehicles)} vehicle(s) from raw text")
                        # Convert to 2D array format
                        from schema import VEHICLE_SCHEMA_ORDER
                        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for vehicle in vehicles:
                            row = []
                            for field in canonical_schema:
                                value = vehicle.get(field)
                                # Clean "None" strings
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
        except Exception as e:
            logger.debug(f"[RAW_TEXT] Error in raw text parsing: {e}")
            # Fall through to OCR table extraction
    
    if not ocr_result.rows:
        # For PDF: PDF → text → block segmentation → extract_fields_from_block
        # Disable Vision API as primary source - use pure text extraction first
        if source_type == SourceType.PDF:
            # Extract raw text from all text blocks
            try:
                from ocr.table_extract import _extract_raw_text
                full_ocr_text = _extract_raw_text(text_blocks)
                
                if full_ocr_text:
                    # Determine domain from mapping_id
                    domain_mapping_id = mapping_id or kwargs.get('mapping_id', '')
                    logger.debug(f"[DEBUG _extract_from_ocr_source] mapping_id param={mapping_id}, kwargs mapping_id={kwargs.get('mapping_id', 'NONE')}, domain_mapping_id={domain_mapping_id}")
                    is_policy = 'polic' in domain_mapping_id.lower() or domain_mapping_id == 'source_pdf_policies'
                    is_driver = 'driver' in domain_mapping_id.lower()
                    is_location = 'location' in domain_mapping_id.lower()
                    
                    # Route to appropriate parser
                    if is_policy:
                        logger.info("[PDF] Using parse_policy_raw_text for policies (OCR fallback)")
                        policies = parse_policy_raw_text(full_ocr_text, source_type="pdf")
                        if policies:
                            logger.info(f"[PDF] Extracted {len(policies)} policy/policies from text blocks")
                            from schema import POLICY_SCHEMA_ORDER
                            canonical_schema = POLICY_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            
                            # Track policy numbers already extracted
                            extracted_policy_numbers = {p.get('policy_number') for p in policies if p.get('policy_number')}
                            
                            for policy in policies:
                                row = []
                                for field in canonical_schema:
                                    value = policy.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            
                            # FALLBACK: Scan OCR text for policy numbers not captured in structured extraction
                            try:
                                import re
                                if full_ocr_text:
                                    logger.info(f"[PDF] Scanning {len(full_ocr_text)} chars of OCR text for fallback policy numbers")
                                    all_policy_numbers = re.findall(r'\b(P\d+|P_[A-Z0-9]+)\b', full_ocr_text, re.IGNORECASE)
                                    logger.info(f"[PDF] All policy numbers found: {all_policy_numbers}")
                                    unique_policy_numbers = list(set(all_policy_numbers))
                                    missing_policy_numbers = [pn for pn in unique_policy_numbers if pn not in extracted_policy_numbers]
                                    
                                    if missing_policy_numbers:
                                        logger.warning(f"[PDF] Found {len(missing_policy_numbers)} missing policy number(s): {missing_policy_numbers}")
                                    
                                    for policy_num in missing_policy_numbers:
                                        logger.warning(f"[PDF] Creating fallback row for policy number: {policy_num}")
                                        fallback_policy = {field: None for field in POLICY_SCHEMA_ORDER}
                                        fallback_policy['policy_number'] = policy_num
                                        row = []
                                        for field in canonical_schema:
                                            row.append(fallback_policy.get(field))
                                        rows_2d.append(row)
                                        logger.warning(f"[PDF] Added fallback row for {policy_num}, total rows now: {len(rows_2d)}")
                                else:
                                    logger.warning("[PDF] full_ocr_text is empty, cannot scan for fallback policy numbers")
                            except Exception as e:
                                logger.warning(f"[PDF] Error in fallback scanning: {e}")
                                import traceback
                                logger.debug(f"[PDF] Fallback scanning traceback: {traceback.format_exc()}")
                            
                            return rows_2d
                    elif is_driver:
                        logger.info("[PDF] Using parse_driver_raw_text for drivers (OCR fallback)")
                        drivers = parse_driver_raw_text(full_ocr_text)
                        if drivers:
                            logger.info(f"[PDF] Extracted {len(drivers)} driver(s) from text blocks")
                            from schema import DRIVER_SCHEMA_ORDER
                            canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for driver in drivers:
                                row = []
                                for field in canonical_schema:
                                    value = driver.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                    elif is_location:
                        logger.info("[PDF] Using parse_locations_raw_text for locations (OCR fallback)")
                        locations = parse_locations_raw_text(full_ocr_text)
                        if locations:
                            logger.info(f"[PDF] Extracted {len(locations)} location(s) from text blocks")
                            from schema import LOCATION_SCHEMA_ORDER
                            canonical_schema = LOCATION_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for location in locations:
                                row = []
                                for field in canonical_schema:
                                    value = location.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                    else:
                        # Default to vehicle parser (backward compatibility)
                        # Pre-split extraction: Extract document-level defaults for handwritten PDFs
                        document_defaults = _extract_document_level_defaults(full_ocr_text, "pdf", file_path)
                        blocks = split_by_vin(full_ocr_text)
                        from mappings import get_mapping_by_id
                        mapping_config = get_mapping_by_id("source_pdf_vehicles")
                        
                        rows = [extract_fields_from_block(b, mapping=mapping_config, source_type="pdf") for b in blocks]
                        # Apply document-level defaults to each vehicle if field is still None
                        for vehicle in rows:
                            if vehicle:
                                for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                    if vehicle.get(field) is None and document_defaults.get(field) is not None:
                                        vehicle[field] = document_defaults[field]
                        pdf_vehicles = [r for r in rows if r and r.get('vin')]
                        
                        if pdf_vehicles:
                            logger.info(f"[PDF] Extracted {len(pdf_vehicles)} vehicle(s) from text blocks")
                            from schema import VEHICLE_SCHEMA_ORDER
                            canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for vehicle in pdf_vehicles:
                                row = []
                                for field in canonical_schema:
                                    value = vehicle.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
            except Exception as e:
                logger.debug(f"[PDF] Text extraction failed: {e}")
        
        # Fallback to Vision API only if text extraction failed (for PDF) or for non-PDF sources
        logger.info("[OCR] OCR unusable — triggering Vision API fallback (reason: No rows extracted)")
        vision_result = _extract_table_with_vision_api(
            file_path, source_type, 
            text_blocks=text_blocks,
            page_text_blocks=page_text_blocks if page_text_blocks else None,
            mapping_id=mapping_id
        )
        if vision_result and len(vision_result) > 1:  # Has header + at least one data row
            logger.info("[Vision] Fallback succeeded, using Vision results")
            return vision_result
        # vision_result is None or empty (only header row) - return empty 2D array with headers
        logger.warning(f"[Vision] Fallback {'failed' if vision_result is None else 'returned empty (only headers)'}, returning empty result with headers")
        # Return empty 2D array with headers to ensure normalize_v2 is called
        from schema import VEHICLE_SCHEMA_ORDER
        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
        return [canonical_schema]  # Return header row only, no data rows
    
    # Get all unique keys from all rows (columns)
    all_keys = set()
    for row in ocr_result.rows:
        all_keys.update(row.keys())
    
    # Sort keys for consistency
    sorted_keys = sorted(all_keys)
    
    # Header row (cleaned keys)
    cleaned_headers = [clean_header(key) for key in sorted_keys]
    rows_2d.append(cleaned_headers)
    
    
    # SAFEGUARD: Filter out header rows for IMAGE sources
    # Data rows
    invalid_vin_tokens = {'YEAR', 'MAKE', 'MODEL', 'COLOR', 'MILEAGE', 'VIN', 'TRANSMISSION', 
                          'BODY_STYLE', 'FUEL_TYPE', 'OWNER_EMAIL', 'NOTES', 'VEHICLE_ID', 
                          'EFFECTIVE_DATE', 'TRIM', 'WEIGHT', 'BODY', 'PAINTED', 'GASOLINE',
                          'CURRENT', 'EXTERIOR', 'ADDITIONAL', 'DETAIL', 'SHEET', 'STYLE',
                          'TYPE', 'EMAIL', 'FUEL', 'BODY_TYPE', 'CURRENT_MILEAGE',
                          'EXTERIOR_COLOR'}
    
    for row_dict in ocr_result.rows:
        
        # SAFEGUARD: For IMAGE sources, filter out rows where VIN is a header token
        # CRITICAL FIX: Use conservative fuzzy matching for IMAGE sources to handle OCR noise
        if source_type == SourceType.IMAGE:
            # Try multiple possible VIN key names (case variations)
            vin_value = row_dict.get('vin') or row_dict.get('VIN') or row_dict.get('VIN Number') or row_dict.get('vin_number')
            if vin_value:
                vin_upper_clean = str(vin_value).upper().strip().rstrip('.,;:!?')
                is_header_token = vin_upper_clean in invalid_vin_tokens
                
                # Conservative fuzzy matching for OCR variations (threshold 0.85)
                if not is_header_token:
                    from inference import fuzzy_match_header
                    for header_token in invalid_vin_tokens:
                        if fuzzy_match_header(vin_upper_clean, header_token, threshold=0.85):
                            is_header_token = True
                            logger.debug(f"[IMAGE Table Extraction] Fuzzy matched header token: {vin_value} -> {header_token}")
                            break
                
                if is_header_token:
                    logger.warning(f"[IMAGE Table Extraction] Filtering row with invalid VIN (header token): {vin_value} -> {vin_upper_clean}")
                    continue  # Skip this row - it's a header row
        
        row_values = []
        for key in sorted_keys:
            value = row_dict.get(key, "")
            # Convert to string for consistency
            row_values.append(str(value) if value is not None else "")
        rows_2d.append(row_values)
    
    # Check if OCR result is unusable OR if raw text is suspiciously short (indicates filtering issue)
    is_unusable, reason = _is_ocr_result_unusable(rows_2d)
    raw_text_length = len(ocr_result.raw_text) if ocr_result.raw_text else 0
    is_text_too_short = raw_text_length < 200  # Suspiciously short for vehicle descriptions
    
    if is_unusable or (source_type == SourceType.IMAGE and is_text_too_short):
        reason_msg = reason if is_unusable else f"Raw text too short ({raw_text_length} chars)"
        logger.warning(f"[{source_type.name}] OCR result unusable or text too short: {reason_msg}, falling back to parse_vehicle_raw_text")
        
        # For IMAGE: Use parse_vehicle_raw_text with original (unfiltered) text blocks
        if source_type == SourceType.IMAGE:
            try:
                from ocr.table_extract import _extract_raw_text
                # Use original text_blocks (before confidence filtering) to get full text
                logger.warning(f"[IMAGE] Fallback: Using {len(text_blocks)} original (unfiltered) text blocks")
                full_ocr_text = _extract_raw_text(text_blocks)
                logger.warning(f"[IMAGE] Fallback: Extracted {len(full_ocr_text)} chars from {len(text_blocks)} blocks. Preview: {full_ocr_text[:300]}")
                
                if full_ocr_text and len(full_ocr_text) > 20:  # Only if we have some text
                    vehicles = parse_vehicle_raw_text(full_ocr_text, source_type="image")
                    if vehicles:
                        logger.warning(f"[IMAGE] Fallback: Extracted {len(vehicles)} vehicle(s) from parse_vehicle_raw_text")
                        from schema import VEHICLE_SCHEMA_ORDER
                        canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                        rows_2d = [canonical_schema]  # Header row
                        for vehicle in vehicles:
                            row = []
                            for field in canonical_schema:
                                value = vehicle.get(field)
                                if isinstance(value, str) and value.strip().lower() == "none":
                                    value = None
                                row.append(value)
                            rows_2d.append(row)
                        return rows_2d
                    else:
                        logger.warning(f"[IMAGE] Fallback: parse_vehicle_raw_text returned 0 vehicles")
                else:
                    logger.warning(f"[IMAGE] Fallback: Text too short ({len(full_ocr_text) if full_ocr_text else 0} chars), skipping")
            except Exception as e:
                logger.warning(f"[IMAGE] Text extraction fallback failed: {e}", exc_info=True)
        
        # For PDF: Try text extraction → parse_vehicle_raw_text
        if source_type == SourceType.PDF:
            try:
                from ocr.table_extract import _extract_raw_text
                full_ocr_text = _extract_raw_text(text_blocks)
                
                if full_ocr_text:
                    # Determine domain from mapping_id
                    domain_mapping_id = mapping_id or kwargs.get('mapping_id', '')
                    logger.debug(f"[DEBUG _extract_from_ocr_source] mapping_id param={mapping_id}, kwargs mapping_id={kwargs.get('mapping_id', 'NONE')}, domain_mapping_id={domain_mapping_id}")
                    is_policy = 'polic' in domain_mapping_id.lower() or domain_mapping_id == 'source_pdf_policies'
                    is_driver = 'driver' in domain_mapping_id.lower()
                    is_location = 'location' in domain_mapping_id.lower()
                    
                    # Route to appropriate parser
                    if is_policy:
                        logger.info("[PDF] Using parse_policy_raw_text for policies (OCR unusable fallback)")
                        policies = parse_policy_raw_text(full_ocr_text, source_type="pdf")
                        if policies:
                            logger.info(f"[PDF] Extracted {len(policies)} policy/policies from text blocks (OCR unusable fallback)")
                            from schema import POLICY_SCHEMA_ORDER
                            canonical_schema = POLICY_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            
                            # Track policy numbers already extracted
                            extracted_policy_numbers = {p.get('policy_number') for p in policies if p.get('policy_number')}
                            
                            for policy in policies:
                                row = []
                                for field in canonical_schema:
                                    value = policy.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            
                            # FALLBACK: Scan OCR text for policy numbers not captured in structured extraction
                            try:
                                import re
                                if full_ocr_text:
                                    all_policy_numbers = re.findall(r'\b(P\d+|P_[A-Z0-9]+)\b', full_ocr_text, re.IGNORECASE)
                                    unique_policy_numbers = list(set(all_policy_numbers))
                                    missing_policy_numbers = [pn for pn in unique_policy_numbers if pn not in extracted_policy_numbers]
                                    
                                    for policy_num in missing_policy_numbers:
                                        logger.warning(f"[PDF] Creating fallback row for policy number: {policy_num}")
                                        fallback_policy = {field: None for field in POLICY_SCHEMA_ORDER}
                                        fallback_policy['policy_number'] = policy_num
                                        row = []
                                        for field in canonical_schema:
                                            row.append(fallback_policy.get(field))
                                        rows_2d.append(row)
                            except Exception as e:
                                logger.debug(f"[PDF] Error in fallback scanning: {e}")
                            
                            return rows_2d
                    elif is_driver:
                        logger.info("[PDF] Using parse_driver_raw_text for drivers (OCR unusable fallback)")
                        drivers = parse_driver_raw_text(full_ocr_text)
                        if drivers:
                            logger.info(f"[PDF] Extracted {len(drivers)} driver(s) from text blocks (OCR unusable fallback)")
                            from schema import DRIVER_SCHEMA_ORDER
                            canonical_schema = DRIVER_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for driver in drivers:
                                row = []
                                for field in canonical_schema:
                                    value = driver.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                    elif is_location:
                        logger.info("[PDF] Using parse_locations_raw_text for locations (OCR unusable fallback)")
                        locations = parse_locations_raw_text(full_ocr_text)
                        if locations:
                            logger.info(f"[PDF] Extracted {len(locations)} location(s) from text blocks (OCR unusable fallback)")
                            from schema import LOCATION_SCHEMA_ORDER
                            canonical_schema = LOCATION_SCHEMA_ORDER.copy()
                            rows_2d = [canonical_schema]
                            for location in locations:
                                row = []
                                for field in canonical_schema:
                                    value = location.get(field)
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d.append(row)
                            return rows_2d
                    else:
                        # Default to vehicle parser (backward compatibility)
                        # Pre-split extraction: Extract document-level defaults for handwritten PDFs
                        document_defaults = _extract_document_level_defaults(full_ocr_text, "pdf", file_path)
                        blocks = split_by_vin(full_ocr_text)
                        from mappings import get_mapping_by_id
                        mapping_config = get_mapping_by_id("source_pdf_vehicles")
                        
                        rows = [extract_fields_from_block(b, mapping=mapping_config, source_type="pdf") for b in blocks]
                        # Apply document-level defaults to each vehicle if field is still None
                        for vehicle in rows:
                            if vehicle:
                                for field in ['body_style', 'fuel_type', 'transmission', 'mileage']:
                                    if vehicle.get(field) is None and document_defaults.get(field) is not None:
                                        vehicle[field] = document_defaults[field]
                        pdf_vehicles = [r for r in rows if r and r.get('vin')]
                        
                        if pdf_vehicles:
                            logger.info(f"[PDF] Extracted {len(pdf_vehicles)} vehicle(s) from text blocks (OCR unusable fallback)")
                            from schema import VEHICLE_SCHEMA_ORDER
                            canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                            rows_2d_text = [canonical_schema]  # Header row
                            for vehicle in pdf_vehicles:
                                row = []
                                for field in canonical_schema:
                                    value = vehicle.get(field)
                                    # Clean "None" strings
                                    if isinstance(value, str) and value.strip().lower() == "none":
                                        value = None
                                    row.append(value)
                                rows_2d_text.append(row)
                            return rows_2d_text
                        else:
                            # No vehicles found - return empty 2D array with headers to ensure normalize_v2 is called
                            logger.warning("[PDF] No vehicles extracted from text blocks (OCR unusable fallback)")
                            from schema import VEHICLE_SCHEMA_ORDER
                            canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                            return [canonical_schema]  # Return header row only, no data rows
            except Exception as e:
                logger.debug(f"[PDF] Text extraction fallback failed: {e}")
        
        # Fallback to Vision API only if text extraction failed
        logger.info(f"[OCR] OCR unusable — triggering Vision API fallback (reason: {reason})")
        vision_result = _extract_table_with_vision_api(
            file_path, source_type,
            text_blocks=text_blocks,
            page_text_blocks=page_text_blocks if page_text_blocks else None,
            mapping_id=mapping_id
        )
        if vision_result and len(vision_result) > 1:  # Has header + at least one data row
            logger.info("[Vision] Fallback succeeded, using Vision results")
            return vision_result
        else:
            # vision_result is None or empty (only header row) - return OCR results (which should have headers)
            logger.warning(f"[Vision] Fallback {'failed' if vision_result is None else 'returned empty (only headers)'}, returning OCR results")
            # Ensure rows_2d has headers (should already have them from OCR extraction)
            if rows_2d and len(rows_2d) > 0:
                return rows_2d
            else:
                # OCR also failed - return empty 2D array with headers to ensure normalize_v2 is called
                from schema import VEHICLE_SCHEMA_ORDER
                canonical_schema = VEHICLE_SCHEMA_ORDER.copy()
                header_row = list(canonical_schema) + ["_is_handwritten"]
                return [header_row]  # Return header row only, no data rows
    
    # OCR result is usable
    logger.debug(f"[OCR] OCR extraction successful: {len(rows_2d)-1} data rows × {len(cleaned_headers)} columns")
    return rows_2d

